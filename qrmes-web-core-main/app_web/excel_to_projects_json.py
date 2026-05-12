import openpyxl
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

# 定义项目名称列的匹配模式
HEADER_PATTERNS = [
    r'项目名称',
    r'项目',
    r'名称',
    r'project.*name',
    r'name',
]

# 定义项目编号列的匹配模式（需要排除的）
CODE_HEADER_PATTERNS = [
    r'项目号',
    r'项目编号',
    r'编号',
    r'code',
    r'id',
    r'number',
]

# 定义项目名称列的匹配模式（优先选择的）
NAME_HEADER_PATTERNS = [
    r'项目名称',
    r'项目',
    r'名称',
    r'project.*name',
    r'name',
]

def has_chinese(text):
    """检查文本是否包含中文字符"""
    if not isinstance(text, str):
        return False
    return bool(re.search(r'[\u4e00-\u9fff]', text))

def is_name_header(header):
    """检查是否为项目名称列标题"""
    if not isinstance(header, str):
        return False
    header_lower = header.lower().strip()
    for pattern in NAME_HEADER_PATTERNS:
        if re.search(pattern, header_lower, re.IGNORECASE):
            return True
    return False

def is_code_header(header):
    """检查是否为项目编号列标题"""
    if not isinstance(header, str):
        return False
    header_lower = header.lower().strip()
    for pattern in CODE_HEADER_PATTERNS:
        if re.search(pattern, header_lower, re.IGNORECASE):
            return True
    return False

def is_probable_code(value):
    """判断值是否可能是项目编号而不是项目名称"""
    if not isinstance(value, str):
        return True
    
    value = value.strip()
    if not value:
        return True
    
    # 如果包含中文，很可能是项目名称
    if has_chinese(value):
        return False
    
    # 如果是纯字母数字组合且长度较短，可能是编号
    if re.match(r'^[A-Za-z0-9/\-_]+$', value) and len(value) <= 15:
        return True
    
    # 如果以PC开头，很可能是项目编号
    if value.upper().startswith('PC'):
        return True
    
    return False

def is_chinese_project_name(value):
    """判断值是否为中文项目名称"""
    if not isinstance(value, str):
        return False
    
    value = value.strip()
    if not value:
        return False
    
    # 必须包含中文
    if not has_chinese(value):
        return False
    
    # 不应该是明显的编号格式
    if is_probable_code(value):
        return False
    
    return True

def read_excel_file(file_path):
    """读取Excel文件并返回所有工作表的数据"""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheets_data = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    sheet_data.append(row)
            
            if sheet_data:
                sheets_data[sheet_name] = sheet_data
        
        workbook.close()
        return sheets_data
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")
        return {}

def find_project_name_columns(sheet_data):
    """在工作表数据中找到项目名称列"""
    if not sheet_data:
        return []
    
    header_row = sheet_data[0]
    project_columns = []
    
    for col_idx, header in enumerate(header_row):
        if header is None:
            continue
        
        header_str = str(header).strip()
        
        # 优先选择明确的项目名称列
        if is_name_header(header_str):
            # 检查是否不是编号列
            if not is_code_header(header_str):
                score = 100  # 最高优先级
                project_columns.append((col_idx, header_str, score))
        # 如果包含"项目"但不是编号列，也考虑
        elif '项目' in header_str and not is_code_header(header_str):
            score = 50
            project_columns.append((col_idx, header_str, score))
    
    # 如果没找到明确的项目名称列，使用更宽泛的匹配
    if not project_columns:
        for col_idx, header in enumerate(header_row):
            if header is None:
                continue
            
            header_str = str(header).strip()
            
            # 检查是否匹配任何项目相关模式，但排除编号列
            for pattern in HEADER_PATTERNS:
                if re.search(pattern, header_str, re.IGNORECASE) and not is_code_header(header_str):
                    score = 30
                    project_columns.append((col_idx, header_str, score))
                    break
    
    # 按分数排序，选择最佳匹配
    project_columns.sort(key=lambda x: x[2], reverse=True)
    return project_columns

def extract_projects_from_sheet(sheet_data):
    """从单个工作表中提取项目名称"""
    project_columns = find_project_name_columns(sheet_data)
    projects = set()
    
    if not project_columns:
        print(f"未找到项目名称列")
        return projects
    
    # 使用得分最高的列
    best_column = project_columns[0]
    col_idx, header_name, score = best_column
    print(f"使用列: {header_name} (索引: {col_idx}, 得分: {score})")
    
    # 从第二行开始提取数据（跳过标题行）
    for row_idx, row in enumerate(sheet_data[1:], start=2):
        if col_idx < len(row) and row[col_idx] is not None:
            value = str(row[col_idx]).strip()
            
            # 只保留中文项目名称，过滤掉项目编号
            if is_chinese_project_name(value):
                projects.add(value)
            elif value and not is_probable_code(value):
                # 如果不是明显的编号，也加入（以防遗漏）
                print(f"可能的项目名称: {value}")
                projects.add(value)
    
    return projects

def extract_all_projects(file_path):
    """从Excel文件中提取所有项目名称"""
    sheets_data = read_excel_file(file_path)
    all_projects = set()
    
    for sheet_name, sheet_data in sheets_data.items():
        print(f"\n处理工作表: {sheet_name}")
        projects = extract_projects_from_sheet(sheet_data)
        print(f"从 {sheet_name} 提取到 {len(projects)} 个项目")
        all_projects.update(projects)
    
    return all_projects

def restore_from_backup(backup_file_path, output_file_path):
    """从备份文件恢复原始的中文项目名称"""
    try:
        with open(backup_file_path, 'r', encoding='utf-8') as f:
            backup_data = json.load(f)
        
        # 只保留中文项目名称
        chinese_projects = []
        for project in backup_data.get('projects', []):
            if is_chinese_project_name(project):
                chinese_projects.append(project)
        
        return set(chinese_projects)
    except Exception as e:
        print(f"从备份文件恢复时出错: {e}")
        return set()

def write_projects_json(projects, output_file_path, backup_file_path=None):
    """将项目列表写入JSON文件，并与备份文件合并"""
    try:
        # 从备份文件恢复原始的中文项目名称
        existing_projects = set()
        if backup_file_path and Path(backup_file_path).exists():
            existing_projects = restore_from_backup(backup_file_path, output_file_path)
            print(f"从备份文件恢复了 {len(existing_projects)} 个中文项目名称")
        
        # 合并项目（只保留中文项目名称）
        all_projects = existing_projects.copy()
        new_projects = set()
        
        for project in projects:
            if is_chinese_project_name(project) and project not in existing_projects:
                new_projects.add(project)
                all_projects.add(project)
        
        # 排序并写入文件
        sorted_projects = sorted(list(all_projects))
        
        projects_data = {
            "projects": sorted_projects
        }
        
        with open(output_file_path, 'w', encoding='utf-8') as f:
            json.dump(projects_data, f, ensure_ascii=False, separators=(',', ':'))
        
        print(f"成功写入 {len(sorted_projects)} 个项目到 {output_file_path}")
        print(f"其中新增 {len(new_projects)} 个项目")
        
        return len(sorted_projects), len(new_projects)
    
    except Exception as e:
        print(f"写入JSON文件时出错: {e}")
        return 0, 0

def main():
    # 文件路径
    excel_file_path = r"F:\GitHub\hours\QRTestScanner\app_web\预置基础资料3_2025102023294506_128558.xlsx"
    output_file_path = r"F:\GitHub\hours\QRTestScanner\app_web\projects.json"
    backup_file_path = r"F:\GitHub\hours\QRTestScanner\app_web\projects_backup_20251020_234419.json"
    
    print("开始处理Excel文件...")
    print(f"Excel文件: {excel_file_path}")
    print(f"输出文件: {output_file_path}")
    print(f"备份文件: {backup_file_path}")
    
    # 检查文件是否存在
    if not Path(excel_file_path).exists():
        print(f"错误: Excel文件不存在: {excel_file_path}")
        return
    
    if not Path(backup_file_path).exists():
        print(f"错误: 备份文件不存在: {backup_file_path}")
        return
    
    # 提取项目名称
    projects = extract_all_projects(excel_file_path)
    print(f"\n从Excel文件中提取到 {len(projects)} 个项目")
    
    # 写入JSON文件（与备份文件合并）
    total_count, new_count = write_projects_json(projects, output_file_path, backup_file_path)
    
    print(f"\n处理完成!")
    print(f"总项目数: {total_count}")
    print(f"新增项目数: {new_count}")

if __name__ == "__main__":
    main()