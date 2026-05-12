# -*- coding: utf-8 -*-
from __future__ import annotations
import re, sqlite3
from datetime import date, datetime
from pathlib import Path
from flask import Blueprint, current_app, jsonify, render_template, request
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None
try:
    from qrmes_shared_core.auth import login_required
except Exception:
    def login_required(func): return func

tool_overview_bp = Blueprint('tool_overview', __name__)
CATEGORY_LABELS = {'measuring_equipment': '计量设备', 'fixture': '工装夹具'}
RANGE_LABELS = {'all': '全部工具记录', '30d': '30 天送检工具', '7d': '7 天送检工具', 'expired': '已过期工具'}
ALIASES = {
    'tool_code': {'工具编号','工具编码','设备编号','工装型号','工装编号','工装编码','编号','编码','toolcode','tool_code','code'},
    'tool_name': {'工具名称','设备名称','设备或工序名称','工装名称','名称','name','toolname','tool_name'},
    'next_check_date': {'校检时间','校验时间','检定有效期','检定日期','检定有效期至','校定有效期','鉴定有效期','下次校验日期','下次校检日期','到期日期','有效期','有效期至','date','duedate','due_date'},
    'department': {'使用部门','部门','使用单位','所在部门','责任部门','所属部门','使用场所','使用地点','dept','department'},
    'purpose': {'主要用途','用途','工序用途','purpose'},
    'quantity': {'数量','qty','quantity'},
    'manufacturer': {'生产厂家','厂家','制造商','供应商','manufacturer'},
    'purchase_date': {'购置日期','购买日期','采购日期','purchase_date','purchasedate'},
    'use_status': {'使用状态','状态','status'},
    'remark': {'备注','说明','remark','remarks'}
}
EXTRA_FIELDS = ('purpose', 'quantity', 'manufacturer', 'purchase_date', 'use_status', 'remark')
EXTRA_FIELD_LABELS = {
    'purpose': '主要用途',
    'quantity': '数量',
    'manufacturer': '生产厂家',
    'purchase_date': '购置日期',
    'use_status': '使用状态',
    'remark': '备注',
}

def _today(): return date.today()
def _norm_header(v): return re.sub(r'[\s_\-:：/\\]+', '', str(v or '').strip().lower())
def _norm_code(v): return re.sub(r'\s+', '', str(v or '').strip()).upper()
def _db_path():
    base = current_app.config.get('TOOL_OVERVIEW_DATA_DIR') or current_app.config.get('DATA_DIR') or '/home/aiyan/QRMES'
    d = Path(base) / 'tool_overview'; d.mkdir(parents=True, exist_ok=True)
    return d / 'tool_overview.db'
def _conn():
    c = sqlite3.connect(_db_path()); c.row_factory = sqlite3.Row; return c

def init_tool_overview_db():
    with _conn() as c:
        c.execute('''CREATE TABLE IF NOT EXISTS tool_overview_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT, category TEXT NOT NULL,
            tool_code TEXT NOT NULL, normalized_tool_code TEXT NOT NULL,
            tool_name TEXT NOT NULL, next_check_date TEXT NOT NULL, department TEXT DEFAULT '',
            inspection_status TEXT DEFAULT 'pending', inspection_sent_at TEXT DEFAULT '',
            original_filename TEXT DEFAULT '', uploaded_at TEXT NOT NULL, updated_at TEXT NOT NULL,
            UNIQUE(category, normalized_tool_code))''')
        c.execute('CREATE INDEX IF NOT EXISTS idx_tool_overview_date ON tool_overview_records(next_check_date)')
        c.execute('CREATE INDEX IF NOT EXISTS idx_tool_overview_code ON tool_overview_records(normalized_tool_code)')
        columns = {row['name'] for row in c.execute('PRAGMA table_info(tool_overview_records)').fetchall()}
        if 'department' not in columns:
            c.execute("ALTER TABLE tool_overview_records ADD COLUMN department TEXT DEFAULT ''")
        if 'inspection_status' not in columns:
            c.execute("ALTER TABLE tool_overview_records ADD COLUMN inspection_status TEXT DEFAULT 'pending'")
        if 'inspection_sent_at' not in columns:
            c.execute("ALTER TABLE tool_overview_records ADD COLUMN inspection_sent_at TEXT DEFAULT ''")
        for field in EXTRA_FIELDS:
            if field not in columns:
                c.execute(f"ALTER TABLE tool_overview_records ADD COLUMN {field} TEXT DEFAULT ''")

def _parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    s = str(v).strip()
    if not s: return None
    if re.fullmatch(r'\d+(?:\.0)?', s):
        n = int(float(s))
        if 20000101 <= n <= 20991231:
            try: return datetime.strptime(str(n), '%Y%m%d').date()
            except ValueError: return None
    s2 = re.sub(r'[年月./]', '-', s).replace('日','').split()[0].rstrip('-')
    for f in ('%Y-%m-%d','%Y%m%d'):
        try: return datetime.strptime(s2, f).date()
        except ValueError: pass
    m = re.search(r'(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})', s)
    if m:
        try: return date(*map(int, m.groups()))
        except ValueError: return None
    return None

def _header_map(row):
    alias = {k: {_norm_header(x) for x in vs} for k, vs in ALIASES.items()}
    out = {}
    for i, h in enumerate(row):
        nh = _norm_header(h)
        for field, names in alias.items():
            if nh in names and field not in out: out[field] = i
    return out

def _row_text(row, i):
    return '' if i is None or i >= len(row) or row[i] is None else str(row[i]).strip()

def _optional_text(row, hm, field):
    return _row_text(row, hm.get(field))

def _days(s):
    d = _parse_date(s)
    return None if not d else (d - _today()).days

def _bucket(days):
    if days is None: return None
    if days < 0: return 'expired'
    if days <= 7: return '7d'
    if days <= 30: return '30d'
    return None

def _record(row):
    x = dict(row); x['days_left'] = _days(x['next_check_date']); x['range'] = _bucket(x['days_left'])
    x['inspection_status'] = x.get('inspection_status') or 'pending'
    x['inspection_sent_at'] = x.get('inspection_sent_at') or ''
    for field in EXTRA_FIELDS:
        x[field] = x.get(field) or ''
    x['category_label'] = CATEGORY_LABELS.get(x['category'], x['category'])
    return x

def _range_sql(name, params):
    t = _today().isoformat()
    if name == 'all': return '1=1'
    if name == 'expired': params.append(t); return 'date(next_check_date) < date(?)'
    if name == '7d': params += [t,t]; return "date(next_check_date) >= date(?) AND date(next_check_date) <= date(?, '+7 day')"
    if name == '30d': params += [t,t]; return "date(next_check_date) >= date(?, '+8 day') AND date(next_check_date) <= date(?, '+30 day')"
    return '1=1'

def _query(range_name=None, category=None, q=None, limit=500):
    init_tool_overview_db(); params=[]; where=[_range_sql(range_name, params)]
    if range_name:
        where.append("COALESCE(inspection_status, 'pending') != 'sent'")
    if category: where.append('category=?'); params.append(category)
    if q:
        where.append('''(normalized_tool_code LIKE ? OR tool_code LIKE ? OR tool_name LIKE ? OR department LIKE ?
            OR purpose LIKE ? OR manufacturer LIKE ? OR purchase_date LIKE ? OR use_status LIKE ? OR remark LIKE ?)''')
        params += ['%' + _norm_code(q) + '%'] + ['%' + q.strip() + '%'] * 8
    params.append(max(1, min(int(limit), 2000)))
    sql = 'SELECT * FROM tool_overview_records WHERE ' + ' AND '.join(where) + ' ORDER BY date(next_check_date), category, tool_code LIMIT ?'
    with _conn() as c: return [_record(r) for r in c.execute(sql, params).fetchall()]

@tool_overview_bp.route('/tool-overview')
@login_required
def tool_overview_page():
    return render_template('tool_overview.html', view_range='overview', range_title='工具总览')

@tool_overview_bp.route('/tool-overview/warnings/<range_name>')
@login_required
def tool_overview_warning_page(range_name):
    if range_name not in RANGE_LABELS: range_name = '30d'
    return render_template('tool_overview.html', view_range=range_name, range_title=RANGE_LABELS[range_name])

@tool_overview_bp.route('/api/tool-overview/upload/<category>', methods=['POST'])
@login_required
def upload_tool_overview(category):
    if category not in CATEGORY_LABELS: return jsonify(success=False, message='未知工具分类'), 400
    if load_workbook is None: return jsonify(success=False, message='服务器缺少 openpyxl，无法解析 Excel'), 500
    f = request.files.get('file')
    if not f or not f.filename: return jsonify(success=False, message='请选择 Excel 文件'), 400
    if not f.filename.lower().endswith(('.xlsx','.xlsm')): return jsonify(success=False, message='仅支持 .xlsx / .xlsm 文件'), 400
    try:
        rows = list(load_workbook(f, data_only=True, read_only=True).active.iter_rows(values_only=True))
    except Exception as e:
        current_app.logger.exception('[工具总览] Excel 解析失败')
        detail = str(e)
        friendly = 'Excel 解析失败：文件不是标准 .xlsx/.xlsm，可能是加密文件、旧版 .xls、CSV/网页表格改后缀，或文件损坏。请用 Excel 另存为标准 .xlsx 后再上传。'
        if detail:
            friendly += ' 原始错误：' + detail
        return jsonify(success=False, message=friendly), 400
    if not rows: return jsonify(success=False, message='Excel 内容为空'), 400
    required = ('tool_code', 'tool_name') if category == 'fixture' else ('tool_code', 'tool_name', 'next_check_date')
    labels = {
        'tool_code': '工装型号' if category == 'fixture' else '工具编号/设备编号',
        'tool_name': '设备或工序名称' if category == 'fixture' else '工具名称/设备名称',
        'next_check_date': '校检时间/检定有效期',
    }
    hm = {}
    header_row_index = None
    best_hm = {}
    best_row_index = 0
    for idx, row in enumerate(rows[:20]):
        candidate = _header_map(row or ())
        if len(candidate) > len(best_hm):
            best_hm = candidate
            best_row_index = idx
        if all(name in candidate for name in required):
            hm = candidate
            header_row_index = idx
            break
    if header_row_index is None:
        missing = [x for x in required if x not in best_hm]
        scanned = min(len(rows), 20)
        expected = '工装型号、设备或工序名称' if category == 'fixture' else '设备编号、设备名称、检定有效期'
        return jsonify(success=False, message='缺少必填表头：' + '、'.join(labels[x] for x in missing) + f'。已扫描前 {scanned} 行，未找到完整表头；请确认表头行包含{expected}。'), 400
    current_app.logger.info('[工具总览] Excel 表头识别成功：第 %s 行，字段=%s', header_row_index + 1, sorted(hm.keys()))
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S'); imported = 0; failed = []
    init_tool_overview_db()
    with _conn() as c:
        for n, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            if not row or not any(cell is not None and str(cell).strip() for cell in row): continue
            code = _row_text(row, hm.get('tool_code')); name = _row_text(row, hm.get('tool_name')); dept = _row_text(row, hm.get('department'))
            due = _parse_date(row[hm['next_check_date']] if hm.get('next_check_date') is not None and hm['next_check_date'] < len(row) else None)
            extra = {field: _optional_text(row, hm, field) for field in EXTRA_FIELDS}
            if not code and not name and not due:
                continue
            errs = []
            if not code: errs.append('工具编号为空')
            if not name: errs.append('工具名称为空')
            if category != 'fixture' and not due: errs.append('校检时间无效')
            if errs: failed.append({'row': n, 'message': '；'.join(errs)}); continue
            c.execute('''INSERT INTO tool_overview_records(category,tool_code,normalized_tool_code,tool_name,next_check_date,department,purpose,quantity,manufacturer,purchase_date,use_status,remark,inspection_status,inspection_sent_at,original_filename,uploaded_at,updated_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(category,normalized_tool_code) DO UPDATE SET
                tool_code=excluded.tool_code, tool_name=excluded.tool_name, next_check_date=excluded.next_check_date,
                department=excluded.department, purpose=excluded.purpose, quantity=excluded.quantity,
                manufacturer=excluded.manufacturer, purchase_date=excluded.purchase_date, use_status=excluded.use_status,
                remark=excluded.remark, inspection_status='pending', inspection_sent_at='',
                original_filename=excluded.original_filename, updated_at=excluded.updated_at''',
                (category, code, _norm_code(code), name, due.isoformat() if due else '', dept,
                 extra['purpose'], extra['quantity'], extra['manufacturer'], extra['purchase_date'],
                 extra['use_status'], extra['remark'], 'pending', '', f.filename, now, now))
            imported += 1
    msg = '已导入 %s 条%s记录' % (imported, CATEGORY_LABELS[category])
    if failed: msg += '，%s 行失败' % len(failed)
    return jsonify(success=True, message=msg, imported=imported, failed_rows=failed)

@tool_overview_bp.route('/api/tool-overview/summary')
@login_required
def summary():
    init_tool_overview_db()
    data = {'total': 0, 'categories': {k: 0 for k in CATEGORY_LABELS}, 'ranges': {'30d': 0, '7d': 0, 'expired': 0}}
    with _conn() as c:
        data['total'] = c.execute('SELECT COUNT(*) FROM tool_overview_records').fetchone()[0]
        for row in c.execute('SELECT category, COUNT(*) AS n FROM tool_overview_records GROUP BY category').fetchall():
            if row['category'] in data['categories']:
                data['categories'][row['category']] = row['n']
        for name in data['ranges']:
            params = []
            where_sql = _range_sql(name, params)
            data['ranges'][name] = c.execute("SELECT COUNT(*) FROM tool_overview_records WHERE COALESCE(inspection_status, 'pending') != 'sent' AND " + where_sql, params).fetchone()[0]
    return jsonify(success=True, summary=data, today=_today().isoformat())

@tool_overview_bp.route('/api/tool-overview/tools')
@login_required
def tools():
    r = request.args.get('range') or None; c = request.args.get('category') or None
    if r and r not in RANGE_LABELS: return jsonify(success=False, message='未知预警范围'), 400
    if c and c not in CATEGORY_LABELS: return jsonify(success=False, message='未知工具分类'), 400
    return jsonify(success=True, tools=_query(r, c, (request.args.get('q') or '').strip() or None, request.args.get('limit', 500)))


@tool_overview_bp.route('/api/tool-overview/tools/<int:record_id>/mark-sent', methods=['POST'])
@login_required
def mark_tool_sent(record_id):
    init_tool_overview_db()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with _conn() as c:
        cur = c.execute(
            """UPDATE tool_overview_records
               SET inspection_status='sent', inspection_sent_at=?, updated_at=?
               WHERE id=?""",
            (now, now, record_id),
        )
        if cur.rowcount == 0:
            return jsonify(success=False, message='未找到工具记录'), 404
    return jsonify(success=True, message='已标记为已送检', inspection_sent_at=now)


@tool_overview_bp.route('/api/tool-overview/tools/<int:record_id>', methods=['DELETE'])
@login_required
def delete_tool_record(record_id):
    init_tool_overview_db()
    with _conn() as c:
        cur = c.execute('DELETE FROM tool_overview_records WHERE id=?', (record_id,))
        if cur.rowcount == 0:
            return jsonify(success=False, message='未找到工具记录'), 404
    return jsonify(success=True, message='工具记录已删除')

@tool_overview_bp.route('/api/tool-overview/search')
@login_required
def search():
    q = (request.args.get('q') or '').strip()
    return jsonify(success=True, tools=[] if not q else _query(q=q, limit=200))
