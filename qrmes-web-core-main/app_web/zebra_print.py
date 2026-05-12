from __future__ import annotations

import io
import re

from flask import Blueprint, Response, jsonify, redirect, render_template, render_template_string, request, url_for
from PIL import Image

from qrmes_shared_core.auth import login_required

zebra_print_bp = Blueprint("zebra_print", __name__)


ZEBRA_PRINT_ENTRY_TEMPLATE = """
{% extends "base.html" %}

{% block title %}标签打印 - MESAPP{% endblock %}

{% block extra_css %}
<style>
.zebra-entry-shell{max-width:1180px;margin:0 auto;padding:1.25rem 0}
.zebra-entry-card{border:1px solid #dbe3ea;border-radius:8px;background:linear-gradient(180deg,#ffffff 0%,#f8fafc 100%);box-shadow:0 18px 40px rgba(15,23,42,.08);overflow:hidden}
.zebra-entry-hero{padding:1.5rem 1.6rem 1.25rem;border-bottom:1px solid #e2e8f0;background:linear-gradient(135deg,#ffffff 0%,#f8fafc 58%,#ecfeff 100%)}
.zebra-entry-hero h2{margin:0;font-size:2.1rem;line-height:1.1;color:#0f172a}
.zebra-entry-hero p{margin:.75rem 0 0;color:#475569;font-size:1rem;line-height:1.7;max-width:620px}
.zebra-entry-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:1rem;padding:1.35rem}
.zebra-entry-link{display:flex;flex-direction:column;gap:.8rem;padding:1.2rem;border:1px solid #dbe3ea;border-radius:8px;background:#fff;text-decoration:none;color:#0f172a;box-shadow:0 10px 24px rgba(15,23,42,.06);transition:transform .16s ease, box-shadow .16s ease, border-color .16s ease}
.zebra-entry-link:hover{transform:translateY(-2px);box-shadow:0 18px 32px rgba(15,23,42,.11);border-color:#94a3b8}
.zebra-entry-badge{display:inline-flex;align-items:center;align-self:flex-start;min-height:1.9rem;padding:.2rem .7rem;border-radius:999px;font-size:.82rem;font-weight:800}
.zebra-entry-badge.product{background:#eef2ff;color:#4f46e5}
.zebra-entry-badge.device{background:#fff7ed;color:#c2410c}
.zebra-entry-badge.material{background:#ecfeff;color:#0f766e}
.zebra-entry-title{font-size:1.35rem;font-weight:800;line-height:1.2}
.zebra-entry-size{font-size:.95rem;font-weight:700;color:#334155}
.zebra-entry-desc{font-size:.95rem;line-height:1.7;color:#64748b}
.zebra-entry-action{margin-top:auto;font-size:.95rem;font-weight:800;color:#0f172a}
@media (max-width:900px){
    .zebra-entry-hero{padding:1.2rem 1.1rem}
    .zebra-entry-hero h2{font-size:1.8rem}
    .zebra-entry-grid{grid-template-columns:1fr;padding:1rem}
}
</style>
{% endblock %}

{% block content %}
<div class="zebra-entry-shell">
    <div class="zebra-entry-card">
        <div class="zebra-entry-hero">
            <h2>标签打印</h2>
            <p>选择要打开的打印窗口，按标签方向进入对应页面，直接录入、排版并输出。</p>
        </div>
        <div class="zebra-entry-grid">
        <a class="zebra-entry-link" href="{{ url_for('zebra_print.zebra_print_product') }}" target="_blank" rel="noopener">
            <span class="zebra-entry-badge product">产品标签</span>
            <div class="zebra-entry-title">产品编码打印</div>
            <div class="zebra-entry-size">60 x 80 mm 纵向标签</div>
            <div class="zebra-entry-desc">适合整机与产品级标签，支持字段编辑、排版预览和二维码打印。</div>
            <div class="zebra-entry-action">打开产品打印窗口</div>
        </a>
        <a class="zebra-entry-link" href="{{ url_for('zebra_print.zebra_print_device') }}" target="_blank" rel="noopener">
            <span class="zebra-entry-badge device">设备标签</span>
            <div class="zebra-entry-title">设备编码打印</div>
            <div class="zebra-entry-size">60 x 80 mm 纵向标签</div>
            <div class="zebra-entry-desc">适合设备、工装和资产编码标签，支持字段编辑、排版预览和二维码打印。</div>
            <div class="zebra-entry-action">打开设备打印窗口</div>
        </a>
        <a class="zebra-entry-link" href="{{ url_for('zebra_print.zebra_print_material') }}" target="_blank" rel="noopener">
            <span class="zebra-entry-badge material">物料标签</span>
            <div class="zebra-entry-title">物料编码打印</div>
            <div class="zebra-entry-size">80 x 60 mm 横向标签</div>
            <div class="zebra-entry-desc">适合 Excel 批量粘贴、动态字段识别和物料标签连续打印。</div>
            <div class="zebra-entry-action">打开物料打印窗口</div>
        </a>
        </div>
    </div>
</div>
{% endblock %}
"""


PRODUCT_FIELD_DEFS_DEFAULT = [
    {
        "key": "project",
        "dataKey": "projectName",
        "labelInputId": "productProjectLabel",
        "valueInputId": "productProjectName",
        "defaultLabel": "项目",
        "placeholder": "输入项目",
    },
    {
        "key": "name",
        "dataKey": "productName",
        "labelInputId": "productNameLabel",
        "valueInputId": "productName",
        "defaultLabel": "产品名称",
        "placeholder": "输入产品名称",
    },
    {
        "key": "code",
        "dataKey": "productCode",
        "labelInputId": "productCodeLabel",
        "valueInputId": "productCode",
        "defaultLabel": "产品型号",
        "placeholder": "输入产品型号",
    },
    {
        "key": "serial",
        "dataKey": "productSerialNumber",
        "labelInputId": "productSerialLabel",
        "valueInputId": "productSerialNumber",
        "defaultLabel": "序列号",
        "placeholder": "输入序列号",
    },
    {
        "key": "qty",
        "dataKey": "workOrderQuantity",
        "labelInputId": "productQuantityLabel",
        "valueInputId": "workOrderQuantity",
        "defaultLabel": "工单数量",
        "placeholder": "输入工单数量",
    },
    {
        "key": "order",
        "dataKey": "workOrderNumber",
        "labelInputId": "productWorkOrderLabel",
        "valueInputId": "workOrderNumber",
        "defaultLabel": "工单号",
        "placeholder": "输入工单号",
    },
    {
        "key": "process",
        "dataKey": "processName",
        "labelInputId": "productProcessLabel",
        "valueInputId": "processName",
        "defaultLabel": "工序",
        "placeholder": "输入工序说明",
        "span": 2,
    },
]

PRODUCT_PASTE_FIELD_ALIASES_DEFAULT = {
    "项目": "project",
    "项目名称": "project",
    "project": "project",
    "projectname": "project",
    "产品": "name",
    "产品名": "name",
    "产品名称": "name",
    "品名": "name",
    "name": "name",
    "productname": "name",
    "产品型号": "code",
    "产品编码": "code",
    "型号": "code",
    "编码": "code",
    "code": "code",
    "model": "code",
    "productcode": "code",
    "序列号": "serial",
    "sn": "serial",
    "serial": "serial",
    "serialnumber": "serial",
    "产品序列号": "serial",
    "二维码": "qr",
    "二维码内容": "qr",
    "qrcode": "qr",
    "qrvalue": "qr",
    "qr": "qr",
    "工单数量": "qty",
    "数量": "qty",
    "qty": "qty",
    "quantity": "qty",
    "workorderquantity": "qty",
    "工单号": "order",
    "工单": "order",
    "mo": "order",
    "order": "order",
    "workorder": "order",
    "workordernumber": "order",
    "工序": "process",
    "工序说明": "process",
    "process": "process",
    "processname": "process",
    "打印份数": "copies",
    "份数": "copies",
    "copies": "copies",
}

DEVICE_FIELD_DEFS = [
    {
        "key": "deviceName",
        "dataKey": "deviceName",
        "labelInputId": "deviceNameLabel",
        "valueInputId": "deviceName",
        "defaultLabel": "设备名称",
        "placeholder": "输入设备名称",
    },
    {
        "key": "deviceModel",
        "dataKey": "deviceModel",
        "labelInputId": "deviceModelLabel",
        "valueInputId": "deviceModel",
        "defaultLabel": "设备型号",
        "placeholder": "输入设备型号",
    },
    {
        "key": "assetNumber",
        "dataKey": "assetNumber",
        "labelInputId": "assetNumberLabel",
        "valueInputId": "assetNumber",
        "defaultLabel": "资产编号",
        "placeholder": "输入资产编号",
    },
    {
        "key": "manufacturer",
        "dataKey": "manufacturer",
        "labelInputId": "manufacturerLabel",
        "valueInputId": "manufacturer",
        "defaultLabel": "生产厂家",
        "placeholder": "输入生产厂家",
    },
    {
        "key": "purchaseDate",
        "dataKey": "purchaseDate",
        "labelInputId": "purchaseDateLabel",
        "valueInputId": "purchaseDate",
        "defaultLabel": "购置日期",
        "placeholder": "输入购置日期",
    },
    {
        "key": "usageStatus",
        "dataKey": "usageStatus",
        "labelInputId": "usageStatusLabel",
        "valueInputId": "usageStatus",
        "defaultLabel": "使用状态",
        "placeholder": "输入使用状态",
    },
    {
        "key": "usageLocation",
        "dataKey": "usageLocation",
        "labelInputId": "usageLocationLabel",
        "valueInputId": "usageLocation",
        "defaultLabel": "使用场所",
        "placeholder": "输入使用场所",
    },
    {
        "key": "monthlyMaintenance",
        "dataKey": "monthlyMaintenance",
        "labelInputId": "monthlyMaintenanceLabel",
        "valueInputId": "monthlyMaintenance",
        "defaultLabel": "月度保养",
        "placeholder": "输入月度保养",
    },
    {
        "key": "nextMonthlyMaintenance",
        "dataKey": "nextMonthlyMaintenance",
        "labelInputId": "nextMonthlyMaintenanceLabel",
        "valueInputId": "nextMonthlyMaintenance",
        "defaultLabel": "下次月度保养",
        "placeholder": "输入下次月度保养",
    },
]

DEVICE_PASTE_FIELD_ALIASES = {
    "设备名称": "deviceName",
    "设备": "deviceName",
    "设备名": "deviceName",
    "devicename": "deviceName",
    "设备型号": "deviceModel",
    "型号": "deviceModel",
    "devicemodel": "deviceModel",
    "资产编号": "assetNumber",
    "资产编码": "assetNumber",
    "设备编码": "assetNumber",
    "设备编号": "assetNumber",
    "assetnumber": "assetNumber",
    "assetcode": "assetNumber",
    "生产厂家": "manufacturer",
    "厂家": "manufacturer",
    "manufacturer": "manufacturer",
    "购置日期": "purchaseDate",
    "purchasedate": "purchaseDate",
    "使用状态": "usageStatus",
    "status": "usageStatus",
    "usagestatus": "usageStatus",
    "使用场所": "usageLocation",
    "场所": "usageLocation",
    "location": "usageLocation",
    "usagelocation": "usageLocation",
    "月度保养": "monthlyMaintenance",
    "monthlymaintenance": "monthlyMaintenance",
    "下次月度保养": "nextMonthlyMaintenance",
    "nextmonthlymaintenance": "nextMonthlyMaintenance",
    "季度保养": "quarterlyMaintenance",
    "quarterlymaintenance": "quarterlyMaintenance",
    "下次季度保养": "nextQuarterlyMaintenance",
    "nextquarterlymaintenance": "nextQuarterlyMaintenance",
    "年度保养": "yearlyMaintenance",
    "yearlymaintenance": "yearlyMaintenance",
    "下次年度保养": "nextYearlyMaintenance",
    "nextyearlymaintenance": "nextYearlyMaintenance",
    "二维码": "qr",
    "二维码内容": "qr",
    "qrcode": "qr",
    "qrvalue": "qr",
    "qr": "qr",
    "打印份数": "copies",
    "份数": "copies",
    "copies": "copies",
}

PRODUCT_FIELD_ALIASES = {
    "项目": "projectName",
    "产品名称": "productName",
    "物料名称": "productName",
    "产品型号": "productCode",
    "物料编码": "productCode",
    "产品编码": "productCode",
    "序列号": "productSerialNumber",
    "工单数量": "workOrderQuantity",
    "工单号": "workOrderNumber",
    "工序": "processName",
    "工序说明": "processName",
    "打印份数": "copies",
    "copies": "copies",
    "projectname": "projectName",
    "productname": "productName",
    "productcode": "productCode",
    "productserialnumber": "productSerialNumber",
    "workorderquantity": "workOrderQuantity",
    "workordernumber": "workOrderNumber",
    "processname": "processName",
    "设备名称": "deviceName",
    "设备型号": "deviceModel",
    "资产编号": "assetNumber",
    "资产编码": "assetNumber",
    "设备编码": "assetNumber",
    "设备编号": "assetNumber",
    "生产厂家": "manufacturer",
    "购置日期": "purchaseDate",
    "使用状态": "usageStatus",
    "使用场所": "usageLocation",
    "月度保养": "monthlyMaintenance",
    "下次月度保养": "nextMonthlyMaintenance",
    "季度保养": "quarterlyMaintenance",
    "下次季度保养": "nextQuarterlyMaintenance",
    "年度保养": "yearlyMaintenance",
    "下次年度保养": "nextYearlyMaintenance",
    "devicename": "deviceName",
    "devicemodel": "deviceModel",
    "assetnumber": "assetNumber",
    "assetcode": "assetNumber",
    "manufacturer": "manufacturer",
    "purchasedate": "purchaseDate",
    "usagestatus": "usageStatus",
    "usagelocation": "usageLocation",
    "monthlymaintenance": "monthlyMaintenance",
    "nextmonthlymaintenance": "nextMonthlyMaintenance",
    "quarterlymaintenance": "quarterlyMaintenance",
    "nextquarterlymaintenance": "nextQuarterlyMaintenance",
    "yearlymaintenance": "yearlyMaintenance",
    "nextyearlymaintenance": "nextYearlyMaintenance",
}

PRODUCT_PRINT_TEMPLATE_DEFAULTS = {
    "product_data_tab_title": "产品数据",
    "product_switch_url": None,
    "product_switch_text": "",
    "product_section_title": "产品编码标签",
    "product_section_description": "对应第一张大标签，通常用于整机或产品级流转。",
    "product_print_button_text": "打印产品码",
    "product_name_label_default": "产品名称",
    "product_name_placeholder": "输入产品名称",
    "product_code_label_default": "产品型号",
    "product_code_placeholder": "输入产品型号",
    "product_serial_label_default": "序列号",
    "product_serial_placeholder": "输入序列号",
    "product_qr_placeholder": "自动按产品型号+序列号生成",
    "product_qr_separator": "",
    "product_preview_title": "产品码排版预览",
    "product_name_font_size_label": "产品名称字号",
    "product_code_font_size_label": "产品型号/序列号字号",
    "product_rotation_label": "产品编码方向",
    "product_print_preview_title": "产品码打印",
    "sample_product": {
        "projectName": "柳工1.8T挖机",
        "productName": "柳工1.8T挖机",
        "productCode": "Genesis-LiGEx-F22-A",
        "productSerialNumber": "2026040300002",
        "productQrValue": "Genesis-LiGEx-F22-A2026040300002",
        "workOrderQuantity": "5-1",
        "workOrderNumber": "MO0001089",
        "processName": "工序汇报入库-普通生产",
        "productCopies": "1",
    },
    "product_field_defs": PRODUCT_FIELD_DEFS_DEFAULT,
    "product_paste_field_aliases": PRODUCT_PASTE_FIELD_ALIASES_DEFAULT,
    "product_qr_source_keys": ["productCode", "productSerialNumber"],
}

DEVICE_PRINT_TEMPLATE_OVERRIDES = {
    "product_data_tab_title": "设备数据",
    "product_section_title": "设备编码标签",
    "product_section_description": "适用于设备、工装或固定资产编码标签打印。",
    "product_print_button_text": "打印设备码",
    "product_name_label_default": "设备名称",
    "product_name_placeholder": "输入设备名称",
    "product_code_label_default": "资产编号",
    "product_code_placeholder": "输入资产编号",
    "product_serial_label_default": "设备型号",
    "product_serial_placeholder": "输入设备型号",
    "product_qr_placeholder": "默认按设备型号##资产编号##购置日期生成，可手动修改",
    "product_qr_separator": "##",
    "product_preview_title": "设备码排版预览",
    "product_name_font_size_label": "设备名称字号",
    "product_code_font_size_label": "设备信息字号",
    "product_rotation_label": "设备编码方向",
    "product_print_preview_title": "设备码打印",
    "sample_product": {
        "deviceName": "激光打标机",
        "deviceModel": "LC-LM20",
        "assetNumber": "EQP-20260427-01",
        "manufacturer": "深圳某某自动化",
        "purchaseDate": "2026-03-15",
        "usageStatus": "在用",
        "usageLocation": "一号装配线",
        "monthlyMaintenance": "2026-05-30",
        "nextMonthlyMaintenance": "2026-06-30",
        "productQrValue": "LC-LM20##RT3820-03##2026-03-15",
        "productCopies": "1",
    },
    "product_field_defs": DEVICE_FIELD_DEFS,
    "product_paste_field_aliases": DEVICE_PASTE_FIELD_ALIASES,
    "product_qr_source_keys": ["deviceModel", "assetNumber", "purchaseDate"],
}

MATERIAL_FIELD_ALIASES = {
    "单号": "ticketNo",
    "工单号": "ticketNo",
    "品名": "itemName",
    "品名称": "itemName",
    "产品名称": "itemName",
    "品名名称": "itemName",
    "物料编码": "materialCode",
    "物料名称": "materialName",
    "物料规格": "materialSpec",
    "数量": "quantity",
    "qty": "quantity",
    "批号": "batchNo",
    "批次": "batchNo",
    "lot": "batchNo",
    "lotno": "batchNo",
    "batch": "batchNo",
    "batchno": "batchNo",
    "二维码": "qrValue",
    "二维码内容": "qrValue",
    "项目": "projectName",
    "打印份数": "copies",
    "ticketno": "ticketNo",
    "itemname": "itemName",
    "materialcode": "materialCode",
    "materialname": "materialName",
    "materialspec": "materialSpec",
    "quantity": "quantity",
    "qrvalue": "qrValue",
    "qrcode": "qrValue",
    "projectname": "projectName",
    "copies": "copies",
}


def _normalize_key(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("：", ":")
    text = re.sub(r"[\s_\-]+", "", text)
    return text


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_int(value: object, default: int = 1) -> int:
    try:
        return max(1, int(float(str(value or default).strip())))
    except Exception:
        return default


def _match_sheet(workbook, names: tuple[str, ...]):
    normalized_names = {_normalize_key(name) for name in names}
    for sheet in workbook.worksheets:
        if _normalize_key(sheet.title) in normalized_names:
            return sheet
    return None


def _product_sheet_score(sheet) -> int:
    rows = [
        [_cell_text(cell) for cell in row]
        for row in sheet.iter_rows(values_only=True, max_row=8, max_col=4)
        if any(_cell_text(cell) for cell in row)
    ]
    if not rows:
        return 0

    score = 0
    for row in rows:
        if not row:
            continue
        normalized_first = _normalize_key(row[0])
        if PRODUCT_FIELD_ALIASES.get(normalized_first):
            score += 2
        for cell in row:
            normalized = _normalize_key(cell)
            if normalized in PRODUCT_FIELD_ALIASES:
                score += 1
    return score


def _material_sheet_score(sheet) -> int:
    rows = [
        [_cell_text(cell) for cell in row]
        for row in sheet.iter_rows(values_only=True, max_row=3)
        if any(_cell_text(cell) for cell in row)
    ]
    if not rows:
        return 0

    score = 0
    for cell in rows[0]:
        normalized = _normalize_key(cell)
        if MATERIAL_FIELD_ALIASES.get(normalized):
            score += 2
        elif re.match(r"^(自定义|custom)\s*:\s*(.+)$", str(cell or ""), flags=re.I):
            score += 1

    if len(rows) > 1:
        sample_row = rows[1]
        if any(_cell_text(value) for value in sample_row):
            score += 1
    return score


def _best_sheet_by_score(workbook, scorer):
    best_sheet = None
    best_score = 0
    for sheet in workbook.worksheets:
        score = scorer(sheet)
        if score > best_score:
            best_sheet = sheet
            best_score = score
    return best_sheet if best_score > 0 else None


def _parse_product_sheet(sheet) -> dict:
    rows = [
        [_cell_text(cell) for cell in row]
        for row in sheet.iter_rows(values_only=True)
        if any(_cell_text(cell) for cell in row)
    ]
    if not rows:
        return {}

    header_tokens = [_normalize_key(cell) for cell in rows[0]]
    key_index = 0
    value_index = 1 if len(rows[0]) > 1 else 0
    label_index = 2 if len(rows[0]) > 2 else None

    if any(token in {"字段", "字段键", "field", "key"} for token in header_tokens):
        key_index = next((idx for idx, token in enumerate(header_tokens) if token in {"字段", "字段键", "field", "key"}), 0)
        value_index = next((idx for idx, token in enumerate(header_tokens) if token in {"值", "value", "内容"}), value_index)
        label_index = next((idx for idx, token in enumerate(header_tokens) if token in {"标签名", "label", "显示名"}), None)
        rows = rows[1:]

    product: dict = {"fieldLabels": {}, "customFields": []}
    for row in rows:
        if key_index >= len(row):
            continue
        raw_key = _cell_text(row[key_index])
        if not raw_key:
            continue
        raw_value = _cell_text(row[value_index]) if value_index < len(row) else ""
        raw_label = _cell_text(row[label_index]) if label_index is not None and label_index < len(row) else ""
        custom_match = re.match(r"^(自定义|custom)\s*:\s*(.+)$", raw_key, flags=re.I)
        if custom_match:
            product["customFields"].append({
                "label": custom_match.group(2).strip(),
                "value": raw_value,
            })
            continue
        canonical = PRODUCT_FIELD_ALIASES.get(_normalize_key(raw_key))
        if not canonical:
            continue
        if canonical == "copies":
            product[canonical] = _to_int(raw_value, 1)
        else:
            product[canonical] = raw_value
        if canonical != "copies" and raw_label:
            product["fieldLabels"][canonical] = raw_label

    if not product["fieldLabels"]:
        product.pop("fieldLabels", None)
    if not product["customFields"]:
        product.pop("customFields", None)
    return product


def _parse_material_sheet(sheet) -> list[dict]:
    rows = [
        [_cell_text(cell) for cell in row]
        for row in sheet.iter_rows(values_only=True)
        if any(_cell_text(cell) for cell in row)
    ]
    if not rows:
        return []

    headers = rows[0]
    normalized_headers = [_normalize_key(header) for header in headers]
    custom_columns: list[tuple[int, str]] = []
    mapped_columns: dict[int, str] = {}

    for index, header in enumerate(headers):
        custom_match = re.match(r"^(自定义|custom)\s*:\s*(.+)$", header or "", flags=re.I)
        if custom_match:
            custom_columns.append((index, custom_match.group(2).strip()))
            continue
        canonical = MATERIAL_FIELD_ALIASES.get(normalized_headers[index])
        if canonical:
            mapped_columns[index] = canonical

    items: list[dict] = []
    for row in rows[1:]:
        item = {
            "ticketNo": "",
            "itemName": "",
            "materialCode": "",
            "materialName": "",
            "materialSpec": "",
            "quantity": "",
            "batchNo": "",
            "qrValue": "",
            "projectName": "",
            "copies": 1,
            "customFields": [],
        }
        for index, canonical in mapped_columns.items():
            if index >= len(row):
                continue
            value = _cell_text(row[index])
            if canonical == "copies":
                item[canonical] = _to_int(value, 1)
            else:
                item[canonical] = value
        for index, label in custom_columns:
            if index >= len(row):
                continue
            value = _cell_text(row[index])
            if value:
                item["customFields"].append({
                    "label": label,
                    "value": value,
                })
        if any(item[key] for key in ("ticketNo", "itemName", "materialCode", "materialName", "materialSpec", "quantity", "batchNo", "qrValue", "projectName")) or item["customFields"]:
            if not item["customFields"]:
                item.pop("customFields", None)
            items.append(item)
    return items


def _build_product_print_context(**overrides) -> dict:
    context = dict(PRODUCT_PRINT_TEMPLATE_DEFAULTS)
    context.update(overrides)
    return context


@zebra_print_bp.route("/zebra-print", methods=["GET"])
@zebra_print_bp.route("/zebra-print/", methods=["GET"])
@login_required
def zebra_print_index():
    return render_template_string(ZEBRA_PRINT_ENTRY_TEMPLATE)


@zebra_print_bp.route("/_disabled/zebra-print-all", methods=["GET"])
@zebra_print_bp.route("/_disabled/zebra-print-all/", methods=["GET"])
@login_required
def zebra_print_all_disabled():
    return redirect(url_for("zebra_print.zebra_print_index"))
    return render_template(
        "zebra_print.html",
        zebra_qr_endpoint=url_for("zebra_print.zebra_print_qr"),
        print_mode="all",
        print_mode_title="标签打印",
        print_mode_description="可同时查看并编辑产品编码排版与物料编码排版。",
    )


@zebra_print_bp.route("/zebra-print/entry", methods=["GET"])
@zebra_print_bp.route("/zebra-print/entry/", methods=["GET"])
@login_required
def zebra_print_entry():
    return redirect(url_for("zebra_print.zebra_print_index"))


@zebra_print_bp.route("/zebra-print/product", methods=["GET"])
@zebra_print_bp.route("/zebra-print/product/", methods=["GET"])
@login_required
def zebra_print_product():
    return render_template(
        "zebra_print.html",
        zebra_qr_endpoint=url_for("zebra_print.zebra_print_qr"),
        print_mode="product",
        print_mode_title="产品编码打印",
        print_mode_description="产品编码打印窗口，按 60 x 80 竖向标签输出。",
        **_build_product_print_context(
            product_switch_url=url_for("zebra_print.zebra_print_material"),
            product_switch_text="打开物料编码打印",
        ),
    )


@zebra_print_bp.route("/zebra-print/device", methods=["GET"])
@zebra_print_bp.route("/zebra-print/device/", methods=["GET"])
@login_required
def zebra_print_device():
    return render_template(
        "zebra_print.html",
        zebra_qr_endpoint=url_for("zebra_print.zebra_print_qr"),
        print_mode="product",
        print_mode_title="设备编码打印",
        print_mode_description="设备编码打印窗口，按 60 x 80 竖向标签输出。",
        is_device_print=True,
        **_build_product_print_context(
            **DEVICE_PRINT_TEMPLATE_OVERRIDES,
            product_switch_url=url_for("zebra_print.zebra_print_product"),
            product_switch_text="打开产品编码打印",
        ),
    )


@zebra_print_bp.route("/zebra-print/material", methods=["GET"])
@zebra_print_bp.route("/zebra-print/material/", methods=["GET"])
@login_required
def zebra_print_material():
    return render_template(
        "zebra_print.html",
        zebra_qr_endpoint=url_for("zebra_print.zebra_print_qr"),
        print_mode="material",
        print_mode_title="物料编码打印",
        print_mode_description="物料编码打印窗口，按 80 x 60 横向标签输出。",
        **_build_product_print_context(),
    )


@zebra_print_bp.route("/api/zebra/qr", methods=["GET"])
@login_required
def zebra_print_qr():
    value = str(request.args.get("value") or "").strip()
    if not value:
        return jsonify({"error": "VALUE_REQUIRED", "message": "二维码内容不能为空"}), 400

    try:
        box_size = max(2, min(int(request.args.get("box_size", "8") or "8"), 20))
    except ValueError:
        box_size = 8
    try:
        border = max(1, min(int(request.args.get("border", "2") or "2"), 8))
    except ValueError:
        border = 2

    try:
        import qrcode
    except Exception as exc:  # pragma: no cover - runtime dependency
        return jsonify({
            "error": "QRCODE_LIBRARY_MISSING",
            "message": f"二维码组件未安装: {exc}",
        }), 503

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=box_size,
        border=border,
    )
    qr.add_data(value)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white")
    if hasattr(image, "get_image"):
        image = image.get_image()
    if not isinstance(image, Image.Image):
        image = image.convert("RGB")

    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    buffer.seek(0)
    return Response(buffer.getvalue(), mimetype="image/png")


@zebra_print_bp.route("/api/zebra/import-excel", methods=["POST"])
@login_required
def zebra_print_import_excel():
    uploaded = request.files.get("file")
    if not uploaded:
        return jsonify({"error": "FILE_REQUIRED", "message": "请上传 Excel 文件"}), 400

    filename = str(uploaded.filename or "").lower()
    if filename and not filename.endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "INVALID_FILE_TYPE", "message": "仅支持 .xlsx / .xlsm 文件"}), 400

    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        return jsonify({"error": "OPENPYXL_MISSING", "message": f"Excel 组件不可用: {exc}"}), 503

    try:
        workbook = load_workbook(io.BytesIO(uploaded.read()), data_only=True)
    except Exception as exc:
        return jsonify({"error": "INVALID_EXCEL", "message": f"Excel 解析失败: {exc}"}), 400

    import_mode = _normalize_key(request.form.get("mode") or "all")

    product_sheet = _match_sheet(workbook, ("产品", "产品码", "产品编码", "设备", "设备码", "设备编码", "product", "device"))
    material_sheet = _match_sheet(workbook, ("物料", "物料码", "物料编码", "materials", "material"))

    if not product_sheet:
        product_sheet = _best_sheet_by_score(workbook, _product_sheet_score)
    if not material_sheet:
        material_sheet = _best_sheet_by_score(workbook, _material_sheet_score)

    if import_mode == "product" and not product_sheet and workbook.worksheets:
        product_sheet = workbook.worksheets[0]
    elif import_mode in {"material", "materials"} and not material_sheet and workbook.worksheets:
        material_sheet = workbook.worksheets[0]
    else:
        if not product_sheet and workbook.worksheets:
            product_sheet = workbook.worksheets[0]
        if not material_sheet:
            if len(workbook.worksheets) > 1:
                material_sheet = workbook.worksheets[1]
            elif workbook.worksheets and _material_sheet_score(workbook.worksheets[0]) > 0:
                material_sheet = workbook.worksheets[0]

    product = _parse_product_sheet(product_sheet) if product_sheet else {}
    materials = _parse_material_sheet(material_sheet) if material_sheet else []

    if import_mode in {"material", "materials"} and not materials:
        return jsonify({
            "error": "MATERIAL_DATA_NOT_FOUND",
            "message": "未识别到物料数据，请确认 Excel 表头包含工单号、产品名称、物料编码、物料名称、物料规格、数量、批号、项目、二维码内容等字段。",
        }), 400
    if import_mode == "product" and not product:
        return jsonify({
            "error": "PRODUCT_DATA_NOT_FOUND",
            "message": "未识别到产品数据，请确认 Excel 使用“字段 / 值 / 标签名”结构，或包含可识别的产品字段。",
        }), 400

    return jsonify({
        "success": True,
        "message": "Excel 导入成功",
        "data": {
            "product": product,
            "materials": materials or [{}],
        },
    })
