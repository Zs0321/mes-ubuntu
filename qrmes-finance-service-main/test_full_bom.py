#!/usr/bin/env python3
import os
os.environ["DASHSCOPE_API_KEY"] = ""
os.environ["PRICING_QWEN_TIMEOUT"] = "5"
os.environ["PRICING_QWEN_TIMEOUT_RETRY"] = "0"

import sys
sys.path.insert(0, '/Volumes/172.16.30.10/volume2/mes_ubuntu_split_result/qrmes-finance-service/app_web')
sys.path.insert(0, '/Volumes/172.16.30.10/volume2/mes_ubuntu_split_result/qrmes-finance-service')

from backend.services.finance_skill_quote_service import FinanceSkillQuoteService
import csv
import openpyxl
import tempfile
from pathlib import Path

class FakeAIRouteService:
    is_ready = False
    max_workers = 1
    def load_skill_workflow_hint(self):
        return "price_bom mandatory; gap/format/volume on demand."
    def plan_script_usage(self, plan_input):
        return {"selected_scripts": ["price_bom_xlsx.py", "model_volume_pricing.py"], "registry": ["price_bom_xlsx.py", "model_volume_pricing.py"]}
    def prepare_staged_pricing_input(self, item):
        return item
    def estimate_item(self, item):
        # 给某些零件模拟 AI 报价
        name = item.get("name", "")
        weight = item.get("weight_kg", 0)
        if "外壳" in name:
            return {"unit_price": 120, "confidence": 0.8, "reasoning": "ADC12高压铸造", "source": "fake"}
        if "机壳" in name:
            return {"unit_price": 300, "confidence": 0.8, "reasoning": "6063拉伸", "source": "fake"}
        if "电机轴" in name:
            return {"unit_price": 80, "confidence": 0.8, "reasoning": "20CrMnTiH机加工", "source": "fake"}
        if "转子挡板" in name:
            return {"unit_price": 50, "confidence": 0.8, "reasoning": "6061机加工", "source": "fake"}
        if "定子组件" in name:
            return {"unit_price": 800, "confidence": 0.8, "reasoning": "定子组件", "source": "fake"}
        if "定子总成" in name:
            return {"unit_price": 900, "confidence": 0.8, "reasoning": "定子总成", "source": "fake"}
        if "轴套" in name:
            return {"unit_price": 40, "confidence": 0.8, "reasoning": "40Cr机加工", "source": "fake"}
        if "轴承压板" in name:
            return {"unit_price": 25, "confidence": 0.8, "reasoning": "Q235A发黑", "source": "fake"}
        if "前端盖" in name:
            return {"unit_price": 150, "confidence": 0.8, "reasoning": "A356低压铸造", "source": "fake"}
        if "水嘴" in name or "弯头水嘴" in name:
            return {"unit_price": 30, "confidence": 0.8, "reasoning": "6061铸造", "source": "fake"}
        if "永磁体" in name:
            return {"unit_price": 5, "confidence": 0.8, "reasoning": "48UH磁钢", "source": "fake"}
        return {"unit_price": 0, "confidence": 0, "reasoning": "", "source": "fake"}

rows_data = [
    ("轴承压板","T45170281.A0","内外径51×108mm","Q235A发黑",0.107,"",1.0),
    ("前端盖","T46110282.A0","OD220mm前端盖（028-B电机）","A356-T6",3.82,"低压模具",1.0),
    ("旋变定子压板","T46170021.A1","厚度2 mm","Q235A",0.014,"",1.0),
    ("轴承","T19100018.A0","6206-2Z/AEM3/C3GJN","轴承钢",0.3,"",1.0),
    ("轴承","T19100019.A0","6309-2Z/AEM3/C3GJN","轴承钢",0.36,"",1.0),
    ("油封","T27080282.A0","SA1J 45×68×12 GJ2668F2","日本NOK油封","","",1.0),
    ("波形弹簧","T19170001.A1","JB/T 7590-2005 （D62/发蓝）","","","",1.0),
    ("内六角圆柱头螺钉","T19020102.A1","GB/T 70.1-2008（M4X12/8.8级/达克罗）","","","",4.0),
    ("内六角圆柱头螺钉","T19020104.A1","GB/T 70.1-2008（M5x16/8.8级/达克罗）","","","",4.0),
    ("内六角圆柱头螺钉","T19020111.A1","GB/T 70.1-2008（M8x30/8.8级/达克罗）","","","",20.0),
    ("六角法兰面螺钉 加大系列 B级","T19020414.A1","GB/T 5789（ M5×20/10.9级/达克罗）","","","",5.0),
    ("六角法兰面螺钉","T19020303.A1","GB/T 5787-2000（M8X20/8.8级/达克罗/带齿）","","","",2.0),
    ("弹簧垫圈","T19030101.A1","GB/T 93-1987（4/达克罗）","","","",4.0),
    ("弹簧垫圈","T19030102.A1","GB/T 93-1987（5/达克罗）","","","",4.0),
    ("弹簧垫圈","T19030104.A1","GB/T 93-1987（8/达克罗）","","","",20.0),
    ("轴用弹性挡圈","T19140004.A0","GB/T 894-2017（45）","","","",1.0),
    ("水嘴","T46140281.A0","D20直水嘴（028-B电机）","6061-T6",0.07,"低压模具",1.0),
    ("弯头水嘴","T46140371.A0","D20弯管，6061","6061-T6",0.07,"低压模具",1.0),
    ("普通平键","T19180001.A1","GB/T 1096-2003（A2X2X8）","45","","",1.0),
    ("O型圈","T27080019.A0","φ221×φ2.65","氟橡胶","","",2.0),
    ("堵头","T49060281.A0","M20×1.5","","","",1.0),
    ("线束总成","U22170112.A0","高速油泵旋变线束","","","",1.0),
    ("线束总成","W22170112.A0","改制临工测试 B+B-电源线2米长安德森插头","","","",1.0),
    ("旋转变压器","T47210801.A1","J52XU9734-L49","","","",1.0),
    ("0型密封圈","T27080017.A0","JB7757.2(20*2)","NBR70","","",2.0),
    ("电控外壳","J11160008.A0","ADC12 HPD","",4.3,"高压模具",1.0),
    ("特大垫圈C级","T19030302.A1","GB/T 5287-2002（5/镀白锌）","","","",1.0),
    ("吊耳","T46170061.A1","304不锈钢","","","",2.0),
    ("定子总成","T44001177","OD220定子总成（028）","","","",1.0),
    ("定子组件","T44200282.A0","OD220mm，L135mm浸漆定子组件","扁线4.12*1.92",5.6,"",1.0),
    ("定子铁芯","T44210291.A1","OD220mm，L135mm，三V","B30AHV1500",14.3,"",1.0),
    ("机壳","T44100282.A0","拉伸机壳OD220mm，L243mm","6063-T5",9.49,"",1.0),
    ("转子总成","T45001148","OD220转子总成（028）","","","",1.0),
    ("转子铁芯","T45120291.A1","OD220mm，L22.7mm，三V","B30AHV1500",2.03,"",6.0),
    ("电机轴","T45110282.A0","L=279.8mm，20CrMnTiH","20CrMnTiH",2.78,"",1.0),
    ("永磁体","T45130291.A1","24×3.5×22.7mm","48UH",0.014583333,"",96.0),
    ("永磁体","T45130292.A1","18×3×22.7mm","48UH",0.009375,"",96.0),
    ("永磁体","T45130293.A1","9.7×3×22.7mm","48UH",0.005052083,"",96.0),
    ("转子挡板","T45140281.A0","内外径50×159mm","6061-T6",0.44,"",2.0),
    ("转子轴套","T45150291.A1","内外径50×82mm","40Cr",0.28,"",1.0),
    ("金属水嘴","J19160007.A0","D20×1.5mm","不锈钢",0.35,"",2.0),
    ("橡胶软管","W27010001.A0","19x27mm EPDM","",0.04,"",0.140),
    ("弹簧管卡","W19010001.A0","锦钢，WN27","",0.02,"",2.0),
    ("内六角平圆头螺钉","T19020703.A0","GB/T 70.2—2000 （M5x16 8.8级/达克罗）","","","",1.0),
]

svc = FinanceSkillQuoteService(ai_route_service=FakeAIRouteService())

with tempfile.TemporaryDirectory(prefix="finance_test_") as temp_dir:
    workdir = Path(temp_dir)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["产品", "产品编码", "物料", "物料编码", "规格", "材质", "重量（kg）", "工艺", "含税单价", "数量"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, it in enumerate(rows_data, 2):
        ws.cell(row=row_idx, column=1, value="高速油泵")
        ws.cell(row=row_idx, column=2, value="Leviticus-Motor-WP-105-D")
        ws.cell(row=row_idx, column=3, value=it[0])
        ws.cell(row=row_idx, column=4, value=it[1])
        ws.cell(row=row_idx, column=5, value=it[2])
        ws.cell(row=row_idx, column=6, value=it[3] if it[3] else "")
        ws.cell(row=row_idx, column=7, value=it[4] if it[4] else "")
        ws.cell(row=row_idx, column=8, value=it[5] if it[5] else "")
        ws.cell(row=row_idx, column=9, value="")
        ws.cell(row=row_idx, column=10, value=it[6])
    input_xlsx = workdir / "input.xlsx"
    wb.save(input_xlsx)
    
    line_csv, grouped_csv, summary_md, snapshot_json = svc._run_price_bom(input_xlsx, workdir)
    
    for av in [1000, 10000, 50000]:
        print(f"\n=== annual_volume={av} ===")
        volume_outputs = svc._run_volume_pricing(line_csv, workdir, av)
        detail_rows = svc._read_volume_detail_rows(volume_outputs)
        
        # 读取 line_csv 并与 detail_rows 关联
        with open(line_csv, 'r', encoding='utf-8-sig') as f:
            line_reader = list(csv.DictReader(f))
        
        detail_map = {r.get("物料"): r for r in detail_rows}
        
        total_sample = 0
        total_baseline = 0
        total_conservative = 0
        for line in line_reader:
            name = line.get("物料", "")
            qty = float(line.get("数量", 0) or 0)
            sample_price = float(line.get("行总价", 0) or 0) * qty
            total_sample += sample_price
            
            dr = detail_map.get(name)
            if dr:
                baseline = float(dr.get("基准总价", 0) or 0) * qty
                conservative = float(dr.get("保守总价", 0) or 0) * qty
                total_baseline += baseline
                total_conservative += conservative
        
        print(f"  样品总价: {total_sample:.2f}")
        print(f"  量产基准总价: {total_baseline:.2f}")
        print(f"  量产保守总价: {total_conservative:.2f}")
