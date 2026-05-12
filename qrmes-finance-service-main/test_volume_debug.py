#!/usr/bin/env python3
import os
os.environ["DASHSCOPE_API_KEY"] = ""
os.environ["PRICING_QWEN_TIMEOUT"] = "5"
os.environ["PRICING_QWEN_TIMEOUT_RETRY"] = "0"

import sys
sys.path.insert(0, '/Volumes/172.16.30.10/volume2/mes_ubuntu_split_result/qrmes-finance-service/app_web')
sys.path.insert(0, '/Volumes/172.16.30.10/volume2/mes_ubuntu_split_result/qrmes-finance-service')

from backend.services.finance_skill_quote_service import FinanceSkillQuoteService
import csv

class FakeAIRouteService:
    is_ready = False
    max_workers = 1
    def load_skill_workflow_hint(self):
        return "price_bom mandatory; gap/format/volume on demand."
    def plan_script_usage(self, plan_input):
        return {"selected_scripts": ["price_bom_xlsx.py", "model_volume_pricing.py"], "registry": ["price_bom_xlsx.py", "model_volume_pricing.py"]}
    def prepare_staged_pricing_input(self, item):
        return item
    def estimate_item(self, item):
        return {"unit_price": 0, "confidence": 0, "reasoning": "", "source": "fake"}

items = [
    {"name":"电控外壳","code":"J11160008.A0","spec":"ADC12 HPD","material":"","weight_kg":4.3,"process":"高压模具","qty":1},
    {"name":"定子组件","code":"T44200282.A0","spec":"OD220mm，L135mm浸漆定子组件","material":"扁线4.12*1.92","weight_kg":5.6,"process":"","qty":1},
    {"name":"前端盖","code":"T46110282.A0","spec":"OD220mm前端盖（028-B电机）","material":"A356-T6","weight_kg":3.82,"process":"低压模具","qty":1},
    {"name":"机壳","code":"T44100282.A0","spec":"拉伸机壳OD220mm，L243mm","material":"6063-T5","weight_kg":9.49,"process":"","qty":1},
    {"name":"电机轴","code":"T45110282.A0","spec":"L=279.8mm，20CrMnTiH","material":"20CrMnTiH","weight_kg":2.78,"process":"","qty":1},
]

svc = FinanceSkillQuoteService(ai_route_service=FakeAIRouteService())

# 调用内部方法生成 line_csv
from pathlib import Path
import tempfile

with tempfile.TemporaryDirectory(prefix="finance_test_") as temp_dir:
    workdir = Path(temp_dir)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["产品", "产品编码", "物料", "物料编码", "规格", "材质", "重量（kg）", "工艺", "含税单价", "数量"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, it in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value="高速油泵")
        ws.cell(row=row_idx, column=2, value="Leviticus-Motor-WP-105-D")
        ws.cell(row=row_idx, column=3, value=it["name"])
        ws.cell(row=row_idx, column=4, value=it["code"])
        ws.cell(row=row_idx, column=5, value=it["spec"])
        ws.cell(row=row_idx, column=6, value=it.get("material") or "")
        ws.cell(row=row_idx, column=7, value=it.get("weight_kg") or "")
        ws.cell(row=row_idx, column=8, value=it.get("process") or "")
        ws.cell(row=row_idx, column=9, value="")
        ws.cell(row=row_idx, column=10, value=it["qty"])
    input_xlsx = workdir / "input.xlsx"
    wb.save(input_xlsx)
    
    line_csv, grouped_csv, summary_md, snapshot_json = svc._run_price_bom(input_xlsx, workdir)
    print(f"line_csv: {line_csv}")
    with open(line_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            print(row)
    
    for av in [1000, 10000, 50000]:
        print(f"\n=== annual_volume={av} ===")
        volume_outputs = svc._run_volume_pricing(line_csv, workdir, av)
        print(f"volume_outputs: {volume_outputs}")
        
        # 读取明细
        detail_rows = svc._read_volume_detail_rows(volume_outputs)
        for r in detail_rows:
            print(f"  {r.get('物料')}: 基准总价={r.get('基准总价')} 保守总价={r.get('保守总价')} 激进总价={r.get('激进总价')}")
