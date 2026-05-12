from __future__ import annotations

import copy
import io
import json
import re
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from backend.config import AppConfig
from backend.services.ai_route_quote_service import AIRouteQuoteService
from backend.services.finance_skill_quote_service import FinanceSkillQuoteService
from backend.services.kingdee_service import KingdeeService


class ExcelQuoteService:
    RULES_SOURCE = "mes_ubuntu/changjiang-bom-pricing"

    HEADER_ALIASES = {
        "code": (
            "物料编码", "子项物料编码", "子件物料编码", "子项编码", "组件编码", "编码", "料号", "物料号",
            "物资编码", "存货编码", "零件号", "图号", "materialcode", "code", "itemcode",
        ),
        "name": (
            "物料名称", "子项名称", "子件物料名称", "零件名称", "名称", "品名", "品名规格",
            "物资名称", "物料长描述", "materialname", "name", "itemname",
        ),
        "spec": ("规格", "规格型号", "型号", "型号规格", "规格型号图号", "规格/型号", "物料规格", "spec", "specification", "itemmodel"),
        "material": ("材质", "材料", "材质/材料", "材质牌号", "原材料", "材料名称", "material", "materialtype"),
        "weight_kg": (
            "单件重量kg", "重量kg", "单重kg", "净重kg", "重量", "单件重量", "单件净重",
            "每件重量", "单重", "weightkg", "weight", "netweight",
        ),
        "qty": ("数量", "用量", "qty", "quantity", "单位用量", "单台用量", "件数", "需求数量", "数量/台"),
        "process": ("工艺", "工序", "process", "processtype", "工艺路线", "加工工艺", "制造工艺", "工序名称", "工艺名称"),
        "loss": ("损耗率", "损耗", "loss", "lossrate", "损耗系数", "损耗(%)", "损耗百分比"),
        "current_unit_price": (
            "采购单价", "当前单价", "参考单价", "含税单价", "未税单价", "最近采购价", "参考采购价",
            "单价", "单价(元)", "price", "currentunitprice",
        ),
        "target_unit_price": ("目标单价", "目标价", "targetunitprice", "targetprice", "目标采购价", "目标成本单价"),
        "material_price": ("材料单价", "原材单价", "原材价", "materialprice", "rawprice", "原材料单价", "材料价", "材料价格", "单公斤材料价"),
        "extra": ("附加费用", "外协附加", "额外费用", "extra", "extracost", "工序附加", "辅料附加", "包装物流"),
        "vendor": ("供应商", "vendor", "supplier", "厂家", "供应商名称", "供方"),
        "unit": ("单位", "unit"),
    }

    PROCESS_DEFAULTS = {
        "定子绕线": {"lossRate": 0.03, "processFactor": 1.2},
        "定子绕线总成": {"lossRate": 0.05, "processFactor": 2.35},
        "铁芯冲压叠压": {"lossRate": 0.05, "processFactor": 1.55},
        "冲压开模件": {"lossRate": 0.08, "processFactor": 1.08},
        "拉伸件": {"lossRate": 0.06, "processFactor": 1.12},
        "拉伸开模件": {"lossRate": 0.06, "processFactor": 1.12},
        "低压铸造": {"lossRate": 0.06, "processFactor": 1.15},
        "高压铸造": {"lossRate": 0.05, "processFactor": 1.14},
        "机加工": {"lossRate": 0.04, "processFactor": 1.18},
    }

    PROCESS_RULES = (
        {"keywords": ("定子总成", "定子组件", "扁线定子"), "process": "定子绕线总成", "material": "铜"},
        {"keywords": ("定子铁芯", "转子铁芯", "铁芯"), "process": "铁芯冲压叠压", "material": "硅钢"},
        {"keywords": ("机壳",), "process": "拉伸件", "material": "铝"},
        {"keywords": ("端盖",), "process": "高压铸造", "material": "铝"},
        {"keywords": ("电机轴", "主轴", "转轴"), "process": "机加工", "material": "钢材"},
        {"keywords": ("永磁体", "磁钢", "磁瓦"), "process": "烧结", "material": "磁材"},
    )

    def __init__(
        self,
        config: AppConfig,
        kingdee_service: KingdeeService | None = None,
        ai_route_service: AIRouteQuoteService | None = None,
        skill_quote_service: FinanceSkillQuoteService | None = None,
    ):
        self.config = config
        self.kingdee_service = kingdee_service or KingdeeService(config.kingdee)
        self.ai_route_service = ai_route_service or AIRouteQuoteService()
        self.skill_quote_service = skill_quote_service or FinanceSkillQuoteService(
            config=config,
            kingdee_service=self.kingdee_service,
            ai_route_service=self.ai_route_service,
        )
        self._demo_payload: dict | None = None

    @property
    def rules_path(self) -> Path:
        return Path(__file__).resolve().parents[3] / "changjiang-bom-pricing"

    def quote_workbook(
        self,
        file_bytes: bytes,
        filename: str = "",
        model_label: str = "",
        production_mode: str = "sample",
        annual_volume: int = 0,
        progress_callback: Callable[[dict], None] | None = None,
    ) -> dict:
        if not file_bytes:
            raise ValueError("请先上传 Excel 文件")

        workbook = self._load_uploaded_workbook(file_bytes, filename)
        worksheet, header_row_idx, column_map = self._select_sheet_and_columns(workbook)
        if not column_map:
            raise ValueError("未识别出 Excel 表头，请至少包含物料编码/名称/数量等列")

        items = []
        for row in worksheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            item = self._row_to_item(row, column_map)
            if item:
                items.append(item)

        if not items:
            raise ValueError("未在 Excel 中解析出有效 BOM 行")

        label = model_label.strip() or Path(filename or "Excel鎶ヤ环").stem or "Excel鎶ヤ环"
        payload = self.skill_quote_service.quote_items(
            items,
            model={
                "label": label,
                "filename": filename,
                "sheet_name": worksheet.title,
                "production_mode": production_mode,
                "annual_volume": annual_volume,
            },
            progress_callback=progress_callback,
        )
        payload["dataset"] = "excel_quote"
        payload["source"] = "excel_upload"
        payload["rules_source"] = self.RULES_SOURCE
        payload["rules_path"] = str(self.rules_path)
        payload["rules_available"] = self.rules_path.exists()
        payload.setdefault("model", {})
        payload["model"]["sheet_name"] = worksheet.title
        payload["exports"] = self.describe_quote_exports(payload)
        return payload

    @staticmethod
    def _load_uploaded_workbook(file_bytes: bytes, filename: str = ""):
        normalized_name = str(filename or "").strip()
        suffix = Path(normalized_name).suffix.lower()
        try:
            return load_workbook(io.BytesIO(file_bytes), data_only=True)
        except (zipfile.BadZipFile, InvalidFileException, KeyError) as exc:
            if suffix == ".xls":
                raise ValueError("上传的文件是 .xls 老格式，当前只支持 .xlsx，请先另存为 Excel 工作簿(.xlsx)后再上传") from exc
            raise ValueError("上传的文件不是有效的 Excel 工作簿(.xlsx)，请检查文件是否损坏，或先另存为 .xlsx 后再上传") from exc

        total_items = len(items)
        self._notify_progress(
            progress_callback,
            stage="preparing",
            processed=0,
            total=total_items,
            message=f"正在准备报价，已处理 0 / {total_items}",
        )

        purchase_refs = self._fetch_purchase_refs([item["code"] for item in items if item["code"]])
        for index, item in enumerate(items, start=1):
            purchase = purchase_refs.get(item["code"], {})
            if purchase:
                item["kingdee_reference_price"] = self._to_number(purchase.get("price"))
                item["kingdee_reference_tax_price"] = self._to_number(purchase.get("tax_price"))
                item["kingdee_supplier_name"] = purchase.get("supplier_name", "")
                item["kingdee_bill_no"] = purchase.get("bill_no", "")
                item["kingdee_reference_date"] = purchase.get("date", "")
                if not item["vendor"]:
                    item["vendor"] = purchase.get("supplier_name", "")
            else:
                item["kingdee_reference_price"] = 0.0
                item["kingdee_reference_tax_price"] = 0.0
                item["kingdee_supplier_name"] = ""
                item["kingdee_bill_no"] = ""
                item["kingdee_reference_date"] = ""

            route = self._estimate_changjiang_route(item)
            item.update(route)

            finance_unit_price, finance_source, finance_has_reference = self._pick_finance_route(item)
            item["reference_unit_price"] = finance_unit_price
            item["reference_source"] = finance_source
            item["finance_route_unit_price"] = finance_unit_price
            item["finance_route_source"] = finance_source
            item["finance_route_has_reference"] = finance_has_reference
            item["finance_route_status"] = "传统参考命中" if finance_has_reference else "缺传统参考"
            item["source_tag"] = self._source_tag(item)

            self._notify_progress(
                progress_callback,
                stage="pricing",
                processed=index,
                total=total_items,
                message=f"正在计算财务传统路线，已处理 {index} / {total_items}",
            )

        ai_route_results = self._estimate_ai_routes(items, progress_callback)
        for item, ai_result in zip(items, ai_route_results):
            item.update(ai_result)
            finance_unit_price = self._to_number(item.get("finance_route_unit_price"))
            ai_unit_price = self._to_number(item.get("ai_route_unit_price"))
            qty = self._to_number(item.get("qty")) or 1.0
            item["route_gap_unit_price"] = ai_unit_price - finance_unit_price
            item["route_gap_total"] = (ai_unit_price - finance_unit_price) * qty
            item["comparison_reason_summary"] = self._analyze_price_gap(item)
            item["status"] = self._item_status(item)

        self._notify_progress(
            progress_callback,
            stage="finalizing",
            processed=total_items,
            total=total_items,
            message=f"正在整理结果，已处理 {total_items} / {total_items}",
        )

        summary = self._build_summary(items)
        label = model_label.strip() or Path(filename or "Excel报价").stem or "Excel报价"
        payload = {
            "dataset": "excel_quote",
            "source": "excel_upload",
            "rules_source": self.RULES_SOURCE,
            "rules_path": str(self.rules_path),
            "rules_available": self.rules_path.exists(),
            "model": {
                "label": label,
                "filename": filename,
                "sheet_name": worksheet.title,
                "item_count": len(items),
            },
            "summary": summary,
            "items": items,
        }
        payload["exports"] = self.describe_quote_exports(payload)
        return payload

    @staticmethod
    def _notify_progress(
        progress_callback: Callable[[dict], None] | None,
        *,
        stage: str,
        processed: int,
        total: int,
        message: str,
        **extra: object,
    ) -> None:
        if progress_callback is None:
            return
        progress = {
            "stage": stage,
            "processed": processed,
            "total": total,
            "message": message,
        }
        progress.update(extra)
        progress_callback(progress)

    def describe_quote_exports(self, payload: dict) -> list[dict]:
        items = payload.get("items", [])
        summary = payload.get("summary", {}) or self._build_summary(items)
        total_items = len(items)
        gap_review_count = len(self._collect_gap_review_items(items))
        prefix = self._export_file_prefix(payload)
        exports = [
            {
                "id": "quote_summary",
                "label": "系统行级报价总表",
                "filename": f"{prefix}_系统行级报价总表.xlsx",
                "description": "基于 skills 结果汇总的总价概览与完整行级报价",
                "item_count": total_items,
            },
            {
                "id": "bom_detail",
                "label": "BOM明细与异常清单",
                "filename": f"{prefix}_BOM明细与异常清单.xlsx",
                "description": "完整 BOM 明细、状态与异常复核信息",
                "item_count": total_items,
            },
            {
                "id": "ai_routes",
                "label": "系统AI分析明细",
                "filename": f"{prefix}_系统AI分析明细.xlsx",
                "description": "基于 skills 知识内容生成的 AI 报价、建议工艺与分析说明",
                "item_count": int(summary.get("ai_ready_count", 0)),
            },
            {
                "id": "finance_references",
                "label": "系统采购参考对照",
                "filename": f"{prefix}_系统采购参考对照.xlsx",
                "description": "Excel、金蝶与财务参考来源对照",
                "item_count": int(summary.get("finance_reference_count", 0)),
            },
            {
                "id": "gap_review",
                "label": "系统差异复核工作簿",
                "filename": f"{prefix}_系统差异复核工作簿.xlsx",
                "description": "高价差、缺传统参考、AI 未形成报价的重点复核项",
                "item_count": gap_review_count,
            },
        ]
        skill_exports = self.skill_quote_service.describe_skill_exports(payload) if payload.get("skill_outputs") else []
        for item in skill_exports:
            exports.append({
                "id": item.get("id"),
                "label": item.get("label"),
                "filename": item.get("filename"),
                "description": "Skill 原始输出文件",
                "item_count": item.get("item_count", total_items),
                "source_path": str((payload.get("skill_outputs") or {}).get(item.get("id")) or ""),
            })
        return exports

    def export_quote_workbook(self, payload: dict) -> bytes:
        normalized_payload = self._normalize_export_payload(payload)
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "总价概览"
        detail_sheet = workbook.create_sheet("行级报价总表")

        summary_sheet.append(["价格口径说明", "默认展示为含税价格；仅 AI 报价与最终选用价拆分未税价和税额，传统报价/金蝶参考价保持原始口径。"])
        summary_sheet.append([])

        model = normalized_payload.get("model", {})
        summary = normalized_payload.get("summary", {})
        items = normalized_payload.get("items", [])

        summary_rows = [
            ("报价名称", model.get("label", "")),
            ("来源文件", model.get("filename", "")),
            ("工作表", model.get("sheet_name", "")),
            ("规则来源", normalized_payload.get("rules_source", self.RULES_SOURCE)),
            ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("物料条数", len(items)),
            ("财务参考总价(元)", summary.get("finance_total", 0)),
            ("AI报价总价(含税,元)", summary.get("ai_total_tax_inclusive", summary.get("ai_total", 0))),
            ("AI报价总价(未税,元)", summary.get("ai_total_tax_exclusive", 0)),
            ("AI报价总税额(元)", summary.get("ai_total_tax_amount", 0)),
            ("量产基准总价(元)", summary.get("volume_baseline_total", 0)),
            ("量产保守总价(元)", summary.get("volume_conservative_total", 0)),
            ("量产激进总价(元)", summary.get("volume_aggressive_total", 0)),
            ("最终选用总价(含税,元)", summary.get("selected_quote_total_tax_inclusive", summary.get("selected_quote_total", 0))),
            ("最终选用总价(未税,元)", summary.get("selected_quote_total_tax_exclusive", 0)),
            ("最终选用总税额(元)", summary.get("selected_quote_total_tax_amount", 0)),
            ("AI-财务价差(元)", summary.get("route_gap_total", 0)),
            ("总重量(kg)", summary.get("total_weight", 0)),
            ("Excel 上传价条数", summary.get("uploaded_price_count", 0)),
            ("金蝶参考价条数", summary.get("kingdee_reference_count", 0)),
            ("Excel目标价条数", summary.get("target_price_count", 0)),
            ("财务参考价命中条数", summary.get("finance_reference_count", 0)),
            ("传统参考缺失条数", summary.get("finance_missing_count", 0)),
            ("AI有效报价条数", summary.get("ai_ready_count", 0)),
            ("AI不可用条数", summary.get("ai_unavailable_count", 0)),
            ("高差异条数", summary.get("high_gap_count", 0)),
            ("待补参数条数", summary.get("pending_count", 0)),
        ]
        for key, value in summary_rows:
            summary_sheet.append([key, value])

        detail_sheet.append([
            "物料编码",
            "物料名称",
            "规格",
            "材质",
            "材质分类",
            "工艺",
            "数量",
            "单位",
            "单件重量(kg)",
            "财务参考单价(元)",
            "财务参考来源",
            "AI报价单价(含税,元)",
            "AI报价单价(未税,元)",
            "AI报价税额(元)",
            "样品机加工单价(元)",
            "量产开模单价(元)",
            "开模费(元)",
            "开模收益平衡点(套/年)",
            f"量产年产量(套/年)",
            f"量产基准单价({int(self._to_number(normalized_payload.get('model', {}).get('annual_volume')) or 0)}套/年,元)",
            f"量产保守单价({int(self._to_number(normalized_payload.get('model', {}).get('annual_volume')) or 0)}套/年,元)",
            f"量产激进单价({int(self._to_number(normalized_payload.get('model', {}).get('annual_volume')) or 0)}套/年,元)",
            "AI置信度",
            "AI分析摘要",
            "AI分析说明",
            "AI-财务价差(元)",
            "差异分析",
            "最终选用单价(含税,元)",
            "最终选用单价(未税,元)",
            "最终选用税额(元)",
            "skills行级估算单价(元)",
            "skills材料成本(元)",
            "skills工艺成本(元)",
            "金蝶参考价(元)",
            "金蝶供应商",
            "Excel采购价(元)",
            "目标价(元)",
            "材料单价(元/kg)",
            "损耗率",
            "附加费用(元)",
            "财务参考小计(元)",
            "AI报价小计(元)",
            "状态",
            "参考项",
        ])

        for item in items:
            qty = self._to_number(item.get("qty")) or 1.0
            finance_unit = self._to_number(item.get("financeRouteUnitPrice", item.get("finance_route_unit_price", item.get("reference_unit_price", 0))))
            ai_unit = self._to_number(item.get("aiRouteUnitPrice", item.get("ai_route_unit_price", 0)))
            changjiang_unit = self._to_number(item.get("changjiangRouteUnitPrice", item.get("changjiang_route_unit_price", item.get("rule_estimate_unit_price", 0))))
            detail_sheet.append([
                item.get("code", ""),
                item.get("name", ""),
                item.get("spec", ""),
                item.get("material", ""),
                item.get("material_alias", ""),
                item.get("process", ""),
                qty,
                item.get("unit", ""),
                item.get("weight_kg", 0),
                finance_unit,
                item.get("financeRouteSource", item.get("finance_route_source", item.get("reference_source", ""))),
                ai_unit,
                item.get("ai_route_tax_exclusive_unit_price", 0),
                item.get("ai_route_tax_amount_unit_price", 0),
                item.get("sample_machining_unit_price", 0),
                item.get("mass_tooling_unit_price", 0),
                item.get("tooling_cost", 0),
                item.get("mass_break_even_volume", 0),
                self._to_number(item.get("annual_volume", payload.get("model", {}).get("annual_volume", 0))),
                item.get("volume_baseline_unit_price", 0),
                item.get("volume_conservative_unit_price", 0),
                item.get("volume_aggressive_unit_price", 0),
                item.get("aiRouteConfidence", item.get("ai_route_confidence", 0)),
                self._compose_ai_analysis_summary(item),
                self._compose_ai_analysis_explanation(item),
                item.get("routeGapUnitPrice", item.get("route_gap_unit_price", 0)),
                item.get("comparisonReasonSummary", item.get("comparison_reason_summary", "")),
                item.get("selected_quote_tax_inclusive_unit_price", item.get("selected_quote_unit_price", 0)),
                item.get("selected_quote_tax_exclusive_unit_price", 0),
                item.get("selected_quote_tax_amount_unit_price", 0),
                changjiang_unit,
                item.get("changjiangMaterialCost", item.get("material_cost_est", 0)),
                item.get("changjiangProcessCost", item.get("process_cost_est", 0)),
                item.get("kingdeeReferencePrice", item.get("kingdee_reference_price", 0)),
                item.get("kingdee_supplier_name", ""),
                item.get("current_unit_price", 0),
                item.get("target_unit_price", 0),
                item.get("rawPrice", item.get("material_price_used", 0)),
                item.get("lossRate", item.get("loss", 0)),
                item.get("extra", 0),
                finance_unit * qty,
                ai_unit * qty,
                item.get("status", item.get("finance_route_status", "")),
                item.get("sourceTag", item.get("source_tag", "")),
            ])

        self._apply_summary_layout(summary_sheet)
        self._apply_table_layout(
            detail_sheet,
            [
                16, 20, 22, 18, 14, 16, 10, 10, 14, 16, 18, 16, 12, 12, 16, 16, 16, 18,
                14, 16, 16, 16, 12, 24, 40, 16, 34, 16, 16, 16, 16, 16, 16, 18, 16, 16,
                14, 16, 16, 16, 16, 18,
            ],
        )

        tooling_items = [
            item for item in items
            if self._to_number(item.get("tooling_cost", 0)) > 0 or self._to_number(item.get("mass_tooling_unit_price", 0)) > 0
        ]
        if tooling_items:
            tooling_sheet = workbook.create_sheet("开模类物料")
            tooling_sheet.append([
                "物料编码",
                "物料名称",
                "规格",
                "数量",
                "AI报价单价(元)",
                "样品机加工单价(元)",
                "量产开模单价(元)",
                "开模费(元)",
                "开模收益平衡点(套/年)",
                "量产年产量(套/年)",
                "AI分析摘要",
                "AI分析说明",
            ])
            for item in tooling_items:
                qty = self._to_number(item.get("qty")) or 1.0
                tooling_sheet.append([
                    item.get("code", ""),
                    item.get("name", ""),
                    item.get("spec", ""),
                    qty,
                    self._to_number(item.get("aiRouteUnitPrice", item.get("ai_route_unit_price", 0))),
                    self._to_number(item.get("sample_machining_unit_price", 0)),
                    self._to_number(item.get("mass_tooling_unit_price", 0)),
                    self._to_number(item.get("tooling_cost", 0)),
                    self._to_number(item.get("mass_break_even_volume", 0)),
                    self._to_number(item.get("annual_volume", payload.get("model", {}).get("annual_volume", 0))),
                    self._compose_ai_analysis_summary(item),
                    self._compose_ai_analysis_explanation(item),
                ])
            self._apply_table_layout(tooling_sheet, [16, 20, 24, 10, 16, 16, 16, 16, 18, 14, 24, 42])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _normalize_export_payload(self, payload: dict) -> dict:
        normalized_payload = copy.deepcopy(payload or {})
        items = list(normalized_payload.get("items") or [])
        model = dict(normalized_payload.get("model") or {})
        annual_volume = self._to_number(model.get("annual_volume"))
        production_mode = str(model.get("production_mode") or "").strip().lower()
        for item in items:
            if production_mode and not item.get("production_mode"):
                item["production_mode"] = production_mode
            if annual_volume > 0 and not self._to_number(item.get("annual_volume")):
                item["annual_volume"] = annual_volume
            FinanceSkillQuoteService._ensure_mass_volume_prices(
                item,
                qty=self._to_number(item.get("qty")) or 1.0,
                manual_surcharge_unit=self._to_number(item.get("manual_surcharge_unit")),
            )
        normalized_payload["items"] = items
        recalculated_summary = FinanceSkillQuoteService._build_summary(items)
        merged_summary = dict(normalized_payload.get("summary") or {})
        merged_summary.update(recalculated_summary)
        normalized_payload["summary"] = merged_summary
        return normalized_payload

    def export_ai_quote_workbook(self, payload: dict) -> bytes:
        normalized_payload = self._normalize_export_payload(payload)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "AI分析明细"
        sheet.append([
            "物料编码",
            "物料名称",
            "规格",
            "数量",
            "AI报价单价(元)",
            "AI报价小计(元)",
            "样品机加工单价(元)",
            "量产开模单价(元)",
            "开模费(元)",
            "开模收益平衡点(套/年)",
            "AI置信度",
            "AI分析摘要",
            "AI建议工艺",
            "AI建议材质",
            "AI分析说明",
            "差异分析摘要",
        ])

        items = normalized_payload.get("items", [])
        for item in items:
            qty = self._to_number(item.get("qty")) or 1.0
            ai_unit = self._to_number(item.get("aiRouteUnitPrice", item.get("ai_route_unit_price", 0)))
            sheet.append([
                item.get("code", ""),
                item.get("name", ""),
                item.get("spec", ""),
                qty,
                ai_unit,
                ai_unit * qty,
                self._to_number(item.get("sample_machining_unit_price", 0)),
                self._to_number(item.get("mass_tooling_unit_price", 0)),
                self._to_number(item.get("tooling_cost", 0)),
                self._to_number(item.get("mass_break_even_volume", 0)),
                self._to_number(item.get("aiRouteConfidence", item.get("ai_route_confidence", 0))),
                self._compose_ai_analysis_summary(item),
                item.get("ai_route_process_guess", item.get("aiRouteProcessGuess", "")),
                item.get("ai_route_material_guess", item.get("aiRouteMaterialGuess", "")),
                self._compose_ai_analysis_explanation(item),
                item.get("comparisonReasonSummary", item.get("comparison_reason_summary", "")),
            ])

        self._apply_table_layout(sheet, [16, 20, 22, 10, 16, 16, 16, 16, 16, 18, 12, 24, 20, 24, 42, 32])

        detail_sheet = workbook.create_sheet("BOM明细与异常清单")
        detail_sheet.append([
            "物料编码",
            "物料名称",
            "规格",
            "材质",
            "数量",
            "财务参考单价(元)",
            "AI报价单价(元)",
            "AI-财务价差(元)",
            "状态",
            "差异分析",
            "量产基准单价(元)",
            "量产保守单价(元)",
            "量产激进单价(元)",
        ])
        for item in items:
            finance_unit = self._to_number(item.get("financeRouteUnitPrice", item.get("finance_route_unit_price", item.get("reference_unit_price", 0))))
            ai_unit = self._to_number(item.get("aiRouteUnitPrice", item.get("ai_route_unit_price", 0)))
            detail_sheet.append([
                item.get("code", ""),
                item.get("name", ""),
                item.get("spec", ""),
                item.get("material", ""),
                self._to_number(item.get("qty")) or 1.0,
                finance_unit,
                ai_unit,
                ai_unit - finance_unit,
                item.get("status", item.get("finance_route_status", "")),
                item.get("comparisonReasonSummary", item.get("comparison_reason_summary", "")),
                self._to_number(item.get("volume_baseline_unit_price", 0)),
                self._to_number(item.get("volume_conservative_unit_price", 0)),
                self._to_number(item.get("volume_aggressive_unit_price", 0)),
            ])
        self._apply_table_layout(detail_sheet, [16, 20, 22, 18, 10, 16, 16, 16, 14, 32, 16, 16, 16])

        gap_sheet = workbook.create_sheet("差异复核工作簿")
        gap_sheet.append([
            "物料编码",
            "物料名称",
            "状态",
            "财务参考单价(元)",
            "AI报价单价(元)",
            "AI-财务价差(元)",
            "差异比例",
            "量产年产量(套/年)",
            "量产基准单价(元)",
            "量产保守单价(元)",
            "量产激进单价(元)",
            "差异分析",
            "建议动作",
        ])
        for item in self._collect_gap_review_items(items):
            finance_unit = self._to_number(item.get("financeRouteUnitPrice", item.get("finance_route_unit_price", item.get("reference_unit_price", 0))))
            ai_unit = self._to_number(item.get("aiRouteUnitPrice", item.get("ai_route_unit_price", 0)))
            gap = ai_unit - finance_unit
            gap_ratio = abs(gap) / max(finance_unit, ai_unit, 1.0)
            gap_sheet.append([
                item.get("code", ""),
                item.get("name", ""),
                item.get("status", item.get("finance_route_status", "")),
                finance_unit,
                ai_unit,
                gap,
                gap_ratio,
                self._to_number(item.get("annual_volume", normalized_payload.get("model", {}).get("annual_volume", 0))),
                self._to_number(item.get("volume_baseline_unit_price", 0)),
                self._to_number(item.get("volume_conservative_unit_price", 0)),
                self._to_number(item.get("volume_aggressive_unit_price", 0)),
                item.get("comparisonReasonSummary", item.get("comparison_reason_summary", "")),
                self._review_action(item),
            ])
        self._apply_table_layout(gap_sheet, [16, 20, 14, 16, 16, 16, 12, 14, 16, 16, 16, 38, 22])
        return self._workbook_to_bytes(workbook)

    def export_finance_reference_workbook(self, payload: dict) -> bytes:
        normalized_payload = self._normalize_export_payload(payload)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "采购参考对照"
        sheet.append([
            "物料编码",
            "物料名称",
            "数量",
            "财务参考单价(元)",
            "财务参考来源",
            "金蝶参考价(元)",
            "金蝶供应商",
            "Excel采购价(元)",
            "Excel目标价(元)",
            "附加费用(元)",
            "财务参考状态",
        ])

        for item in normalized_payload.get("items", []):
            sheet.append([
                item.get("code", ""),
                item.get("name", ""),
                self._to_number(item.get("qty")) or 1.0,
                self._to_number(item.get("financeRouteUnitPrice", item.get("finance_route_unit_price", item.get("reference_unit_price", 0)))),
                item.get("financeRouteSource", item.get("finance_route_source", item.get("reference_source", ""))),
                self._to_number(item.get("kingdeeReferencePrice", item.get("kingdee_reference_price", 0))),
                item.get("kingdee_supplier_name", ""),
                self._to_number(item.get("current_unit_price", 0)),
                self._to_number(item.get("target_unit_price", 0)),
                self._to_number(item.get("extra", 0)),
                item.get("finance_route_status", item.get("financeRouteStatus", item.get("status", ""))),
            ])

        self._apply_table_layout(sheet, [16, 20, 10, 16, 18, 16, 18, 16, 16, 14, 16])
        return self._workbook_to_bytes(workbook)

    def export_bom_detail_workbook(self, payload: dict) -> bytes:
        normalized_payload = self._normalize_export_payload(payload)
        workbook = Workbook()
        detail_sheet = workbook.active
        detail_sheet.title = "BOM明细与异常清单"
        detail_sheet.append([
            "物料编码",
            "物料名称",
            "规格",
            "材质",
            "数量",
            "财务参考单价(元)",
            "AI报价单价(元)",
            "AI-财务价差(元)",
            "状态",
            "差异分析",
            "量产基准单价(元)",
            "量产保守单价(元)",
            "量产激进单价(元)",
        ])
        items = normalized_payload.get("items", [])
        for item in items:
            finance_unit = self._to_number(item.get("financeRouteUnitPrice", item.get("finance_route_unit_price", item.get("reference_unit_price", 0))))
            ai_unit = self._to_number(item.get("aiRouteUnitPrice", item.get("ai_route_unit_price", 0)))
            detail_sheet.append([
                item.get("code", ""),
                item.get("name", ""),
                item.get("spec", ""),
                item.get("material", ""),
                self._to_number(item.get("qty")) or 1.0,
                finance_unit,
                ai_unit,
                ai_unit - finance_unit,
                item.get("status", item.get("finance_route_status", "")),
                item.get("comparisonReasonSummary", item.get("comparison_reason_summary", "")),
                self._to_number(item.get("volume_baseline_unit_price", 0)),
                self._to_number(item.get("volume_conservative_unit_price", 0)),
                self._to_number(item.get("volume_aggressive_unit_price", 0)),
            ])
        self._apply_table_layout(detail_sheet, [16, 20, 22, 18, 10, 16, 16, 16, 14, 32, 16, 16, 16])

        gap_sheet = workbook.create_sheet("差异复核工作簿")
        gap_sheet.append([
            "物料编码",
            "物料名称",
            "状态",
            "财务参考单价(元)",
            "AI报价单价(元)",
            "AI-财务价差(元)",
            "差异比例",
            "量产年产量(套/年)",
            "量产基准单价(元)",
            "量产保守单价(元)",
            "量产激进单价(元)",
            "差异分析",
            "建议动作",
        ])
        for item in self._collect_gap_review_items(items):
            finance_unit = self._to_number(item.get("financeRouteUnitPrice", item.get("finance_route_unit_price", item.get("reference_unit_price", 0))))
            ai_unit = self._to_number(item.get("aiRouteUnitPrice", item.get("ai_route_unit_price", 0)))
            gap = ai_unit - finance_unit
            gap_ratio = abs(gap) / max(finance_unit, ai_unit, 1.0)
            gap_sheet.append([
                item.get("code", ""),
                item.get("name", ""),
                item.get("status", item.get("finance_route_status", "")),
                finance_unit,
                ai_unit,
                gap,
                gap_ratio,
                self._to_number(item.get("annual_volume", normalized_payload.get("model", {}).get("annual_volume", 0))),
                self._to_number(item.get("volume_baseline_unit_price", 0)),
                self._to_number(item.get("volume_conservative_unit_price", 0)),
                self._to_number(item.get("volume_aggressive_unit_price", 0)),
                item.get("comparisonReasonSummary", item.get("comparison_reason_summary", "")),
                self._review_action(item),
            ])
        self._apply_table_layout(gap_sheet, [16, 20, 14, 16, 16, 16, 12, 14, 16, 16, 16, 38, 22])
        return self._workbook_to_bytes(workbook)

    def export_gap_review_workbook(self, payload: dict) -> bytes:
        normalized_payload = self._normalize_export_payload(payload)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "差异复核工作簿"
        sheet.append([
            "物料编码",
            "物料名称",
            "状态",
            "财务参考单价(元)",
            "AI报价单价(元)",
            "AI-财务价差(元)",
            "差异比例",
            "量产年产量(套/年)",
            "量产基准单价(元)",
            "量产保守单价(元)",
            "量产激进单价(元)",
            "差异分析",
            "建议动作",
        ])

        for item in self._collect_gap_review_items(normalized_payload.get("items", [])):
            finance_unit = self._to_number(item.get("financeRouteUnitPrice", item.get("finance_route_unit_price", item.get("reference_unit_price", 0))))
            ai_unit = self._to_number(item.get("aiRouteUnitPrice", item.get("ai_route_unit_price", 0)))
            gap = ai_unit - finance_unit
            gap_ratio = abs(gap) / max(finance_unit, ai_unit, 1.0)
            sheet.append([
                item.get("code", ""),
                item.get("name", ""),
                item.get("status", item.get("finance_route_status", "")),
                finance_unit,
                ai_unit,
                gap,
                gap_ratio,
                self._to_number(item.get("annual_volume", payload.get("model", {}).get("annual_volume", 0))),
                item.get("volume_baseline_unit_price", 0),
                item.get("volume_conservative_unit_price", 0),
                item.get("volume_aggressive_unit_price", 0),
                item.get("comparisonReasonSummary", item.get("comparison_reason_summary", "")),
                self._review_action(item),
            ])

        self._apply_table_layout(sheet, [16, 20, 14, 16, 16, 16, 12, 14, 16, 16, 16, 38, 22])
        return self._workbook_to_bytes(workbook)

    def export_quote_package(self, payload: dict) -> bytes:
        manifest = self.describe_quote_exports(payload)
        builders = {
            "quote_summary": self.export_quote_workbook,
            "bom_detail": self.export_bom_detail_workbook,
            "ai_routes": self.export_ai_quote_workbook,
            "finance_references": self.export_finance_reference_workbook,
            "gap_review": self.export_gap_review_workbook,
        }
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for item in manifest:
                builder = builders.get(item.get("id"))
                if builder is not None:
                    archive.writestr(item["filename"], builder(payload))
                    continue
                source_path = Path(str(item.get("source_path") or ""))
                if source_path.exists() and source_path.is_file():
                    archive.write(source_path, arcname=item["filename"])
            archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8"))
        buffer.seek(0)
        return buffer.getvalue()

    def export_quote_package_batch(self, payloads: list[dict]) -> bytes:
        valid_payloads = [payload for payload in payloads if isinstance(payload, dict) and (payload.get("items") or [])]
        if not valid_payloads:
            raise ValueError("当前没有可下载的 AI 报价结果")

        builders = {
            "quote_summary": self.export_quote_workbook,
            "bom_detail": self.export_bom_detail_workbook,
            "ai_routes": self.export_ai_quote_workbook,
            "finance_references": self.export_finance_reference_workbook,
            "gap_review": self.export_gap_review_workbook,
        }
        batch_manifest: list[dict] = []
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for index, payload in enumerate(valid_payloads, start=1):
                folder = self._export_file_prefix(payload)
                manifest = self.describe_quote_exports(payload)
                batch_manifest.append({
                    "index": index,
                    "folder": folder,
                    "model": payload.get("model", {}),
                    "exports": manifest,
                })
                for item in manifest:
                    arcname = f"{folder}/{item['filename']}"
                    builder = builders.get(item.get("id"))
                    if builder is not None:
                        archive.writestr(arcname, builder(payload))
                        continue
                    source_path = Path(str(item.get("source_path") or ""))
                    if source_path.exists() and source_path.is_file():
                        archive.write(source_path, arcname=arcname)
                archive.writestr(
                    f"{folder}/manifest.json",
                    json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8"),
                )
            archive.writestr("manifest.json", json.dumps(batch_manifest, ensure_ascii=False, indent=2).encode("utf-8"))
        buffer.seek(0)
        return buffer.getvalue()


    def _collect_gap_review_items(self, items: list[dict]) -> list[dict]:
        rows: list[dict] = []
        for item in items:
            finance_unit = self._to_number(item.get("financeRouteUnitPrice", item.get("finance_route_unit_price", item.get("reference_unit_price", 0))))
            ai_unit = self._to_number(item.get("aiRouteUnitPrice", item.get("ai_route_unit_price", 0)))
            gap_ratio = abs(ai_unit - finance_unit) / max(finance_unit, ai_unit, 1.0) if (finance_unit > 0 or ai_unit > 0) else 0.0
            if finance_unit <= 0 or ai_unit <= 0 or gap_ratio >= 0.15:
                rows.append(item)
        return rows

    @staticmethod
    def _review_action(item: dict) -> str:
        finance_unit = ExcelQuoteService._to_number(item.get("financeRouteUnitPrice", item.get("finance_route_unit_price", item.get("reference_unit_price", 0))))
        ai_unit = ExcelQuoteService._to_number(item.get("aiRouteUnitPrice", item.get("ai_route_unit_price", 0)))
        if finance_unit <= 0 and ai_unit <= 0:
            return "补充重量、材质、工艺后重新生成报价"
        if finance_unit <= 0:
            return "补充金蝶或 Excel 采购参考，再复核 AI 报价"
        if ai_unit <= 0:
            return "查看 AI 说明并重试，必要时人工补价"
        return "优先复核差异原因，确认供应商、批量与工艺口径"

    @staticmethod
    def _apply_summary_layout(sheet) -> None:
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 26
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2):
            row[0].alignment = copy.copy(row[0].alignment)
            row[0].alignment = row[0].alignment.copy(wrap_text=True, vertical="center")
            row[1].alignment = copy.copy(row[1].alignment)
            row[1].alignment = row[1].alignment.copy(wrap_text=True, vertical="center")

    @staticmethod
    def _apply_table_layout(sheet, widths: list[float], freeze_panes: str = "A2") -> None:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = freeze_panes
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        for cell in sheet[1]:
            cell.alignment = copy.copy(cell.alignment)
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="center")
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = copy.copy(cell.alignment)
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")


    @classmethod
    def _compose_ai_analysis_explanation(cls, item: dict) -> str:
        reasoning = cls._to_text(item.get("aiRouteReasoning", item.get("ai_route_reasoning", "")))
        material_cost = cls._to_number(item.get("ai_material_cost_reference", item.get("aiMaterialCostReference", 0)))
        process_cost = cls._to_number(item.get("ai_process_cost_reference", item.get("aiProcessCostReference", 0)))
        process_rule = cls._to_text(item.get("ai_process_rule_reference", item.get("aiProcessRuleReference", "")))
        process_rule_label = cls._to_text(item.get("ai_process_rule_label", item.get("aiProcessRuleLabel", "")))
        parts: list[str] = []
        if reasoning:
            parts.append(reasoning)
        breakdown: list[str] = []
        if material_cost > 0:
            breakdown.append(f"材料费：{material_cost:.2f} 元")
        if process_cost > 0:
            breakdown.append(f"工艺费：{process_cost:.2f} 元")
        if process_rule:
            rule_text = f"命中的工艺规则：{process_rule}"
            if process_rule_label:
                rule_text += f"（计价口径：{process_rule_label}）"
            breakdown.append(rule_text)
        if breakdown:
            parts.append("\n".join(breakdown))
        return "\n".join(part for part in parts if part).strip()

    @classmethod
    def _compose_ai_analysis_summary(cls, item: dict) -> str:
        summary_parts: list[str] = []
        route_source = cls._to_text(item.get("ai_route_source", item.get("aiRouteSource", "")))
        if route_source == "mass-tooling-route":
            summary_parts.append("已按量产开模口径计价")
        confidence = cls._to_number(item.get("aiRouteConfidence", item.get("ai_route_confidence", 0)))
        if confidence > 0:
            summary_parts.append(f"AI置信度 {confidence:.2f}")
        comparison = cls._to_text(item.get("comparisonReasonSummary", item.get("comparison_reason_summary", "")))
        if comparison:
            summary_parts.append(comparison)
        tooling_cost = cls._to_number(item.get("tooling_cost", 0))
        if tooling_cost > 0:
            summary_parts.append(f"开模费 {tooling_cost:.2f} 元")
        break_even = cls._to_number(item.get("mass_break_even_volume", 0))
        if break_even > 0:
            summary_parts.append(f"平衡点 {int(round(break_even))} 套/年")
        text = "；".join(part for part in summary_parts if part).strip()
        if text:
            return text
        reasoning = cls._compose_ai_analysis_explanation(item)
        if not reasoning:
            return ""
        first_line = reasoning.splitlines()[0].strip()
        return first_line[:120] + ("..." if len(first_line) > 120 else "")

    @staticmethod
    def _workbook_to_bytes(workbook: Workbook) -> bytes:
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _export_file_prefix(self, payload: dict) -> str:
        model = payload.get("model", {})
        label = self._to_text(model.get("original_label")) or self._to_text(model.get("label")) or self._to_text(model.get("filename")) or "AI报价"
        volume_label = self._export_volume_label(model)
        if volume_label and volume_label not in label:
            label = f"{label}_{volume_label}"
        safe = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", label).strip("_")
        return safe or "AI报价"

    @classmethod
    def _export_volume_label(cls, model: dict | None) -> str:
        model = model or {}
        requested_label = cls._to_text(model.get("requested_volume_label"))
        requested_match = re.search(r"(\d+)套/年", requested_label) if requested_label else None
        if requested_match:
            return f"{requested_match.group(1)}套/年"
        if requested_label:
            return requested_label
        annual_volume = int(cls._to_number(model.get("requested_annual_volume") or model.get("annual_volume") or 0))
        if annual_volume > 0:
            return f"{annual_volume}套/年"
        return ""

    def _select_sheet_and_columns(self, workbook):
        best_sheet = workbook.worksheets[0]
        best_row_idx = 1
        best_map: dict[str, int] = {}
        best_score = -1

        for worksheet in workbook.worksheets:
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row), values_only=True), start=1):
                column_map = self._match_columns(row)
                score = len(column_map)
                if score > best_score:
                    best_sheet = worksheet
                    best_row_idx = row_idx
                    best_map = column_map
                    best_score = score
        return best_sheet, best_row_idx, best_map

    def _match_columns(self, row_values) -> dict[str, int]:
        mapping: dict[str, int] = {}
        normalized_cells = [self._normalize_header(value) for value in row_values]
        for field, aliases in self.HEADER_ALIASES.items():
            normalized_aliases = {self._normalize_header(alias) for alias in aliases}
            matched_idx = self._find_alias_index(normalized_cells, normalized_aliases)
            if matched_idx is not None:
                mapping[field] = matched_idx
        return mapping

    @staticmethod
    def _find_alias_index(normalized_cells: list[str], normalized_aliases: set[str]) -> int | None:
        for idx, cell in enumerate(normalized_cells):
            if cell and cell in normalized_aliases:
                return idx

        for idx, cell in enumerate(normalized_cells):
            if not cell:
                continue
            for alias in normalized_aliases:
                if len(alias) < 2:
                    continue
                if alias in cell or cell in alias:
                    return idx
        return None

    @staticmethod
    def _normalize_header(value) -> str:
        text = str(value or "").strip().lower()
        text = re.sub(r"[\s_\-/（）()【】\[\]：:，,\.]+", "", text)
        return text

    def _row_to_item(self, row_values, column_map: dict[str, int]) -> dict | None:
        def raw(field: str):
            idx = column_map.get(field)
            if idx is None or idx >= len(row_values):
                return None
            return row_values[idx]

        code = self._to_text(raw("code"))
        name = self._to_text(raw("name"))
        spec = self._to_text(raw("spec"))
        material_text = self._to_text(raw("material"))
        if not any((code, name, spec, material_text)):
            return None

        qty = self._to_number(raw("qty")) or 1.0
        weight_kg = self._to_number(raw("weight_kg"))
        process_original = self._to_text(raw("process"))
        inferred_process, inferred_material = self._infer_process_and_material(name, spec, material_text, process_original)
        process = process_original or inferred_process
        process_inferred = bool(not process_original and inferred_process)
        process_inference_note = ""
        if process_inferred:
            process_inference_note = f"Excel未填写工艺，系统按物料特征默认使用{process}口径"
        material_alias = self._material_alias(material_text)
        if not material_alias and inferred_material:
            material_alias = inferred_material
        if material_alias == "未识别" and inferred_material:
            material_alias = inferred_material
        if not material_alias:
            material_alias = self._material_alias(" ".join(part for part in (name, spec, process) if part))
        if material_alias == "未识别" and inferred_material:
            material_alias = inferred_material
        defaults = self.PROCESS_DEFAULTS.get(process, {"lossRate": 0.05, "processFactor": 1.1})
        loss = self._to_ratio(raw("loss"), defaults["lossRate"])
        material_price_raw = self._to_number(raw("material_price"))
        market_lookup = self._latest_raw_price(material_text or material_alias or name)
        material_price_used = material_price_raw or market_lookup["price"] or 0.0

        return {
            "code": code,
            "name": name or code or "未命名物料",
            "spec": spec,
            "vendor": self._to_text(raw("vendor")),
            "qty": qty,
            "unit": self._to_text(raw("unit")) or "Pcs",
            "current_unit_price": self._to_number(raw("current_unit_price")),
            "target_unit_price": self._to_number(raw("target_unit_price")),
            "material": material_text or material_alias or "未识别",
            "material_original": material_text or material_alias or "未识别",
            "material_alias": material_alias or inferred_material or "",
            "material_price": material_price_raw,
            "weight_kg": weight_kg,
            "process": process,
            "process_original": process_original,
            "process_inferred": process_inferred,
            "process_inference_note": process_inference_note,
            "material_cost_est": 0.0,
            "process_cost_est": 0.0,
            "material_price_used": material_price_used,
            "material_price_source": market_lookup["source"] if material_price_used and not material_price_raw else ("excel" if material_price_raw else "pending"),
            "loss": loss,
            "extra": self._to_number(raw("extra")),
        }

    def _pick_finance_route(self, item: dict) -> tuple[float, str, bool]:
        extra = self._to_number(item.get("extra"))
        uploaded_price = self._to_number(item.get("current_unit_price"))
        if uploaded_price > 0:
            return uploaded_price + extra, "Excel表格采购价", True

        kingdee_price = self._to_number(item.get("kingdee_reference_price"))
        if kingdee_price > 0:
            return kingdee_price + extra, "金蝶最近采购价", True

        target_price = self._to_number(item.get("target_unit_price"))
        if target_price > 0:
            return target_price + extra, "Excel目标价", True

        return 0.0, "待补传统参考", False

    def _estimate_ai_routes(
        self,
        items: list[dict],
        progress_callback: Callable[[dict], None] | None = None,
    ) -> list[dict]:
        if not items:
            return []

        total_items = len(items)
        if not self.ai_route_service.is_ready:
            return [self._empty_ai_route_result("Qwen API key 未配置，AI 路线未执行") for _ in items]

        results = [self._empty_ai_route_result("AI 未执行") for _ in items]
        max_workers = min(self.ai_route_service.max_workers, total_items) or 1
        completed = 0

        self._notify_progress(
            progress_callback,
            stage="ai_pricing",
            processed=0,
            total=total_items,
            parallel_workers=max_workers,
            mode="parallel",
            message=f"正在并行生成 AI 报价，已启动 {max_workers} 路任务",
        )

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_map = {
                executor.submit(self.ai_route_service.estimate_item, item): index
                for index, item in enumerate(items)
            }
            for future in as_completed(future_map):
                index = future_map[future]
                item = items[index]
                try:
                    ai_result = future.result()
                    results[index] = self._compose_ai_route_result(item, ai_result)
                except Exception as exc:  # pragma: no cover
                    reason, status = self._classify_ai_route_exception(str(exc))
                    results[index] = self._empty_ai_route_result(reason, status=status)
                completed += 1
                self._notify_progress(
                    progress_callback,
                    stage="ai_pricing",
                    processed=completed,
                    total=total_items,
                    parallel_workers=max_workers,
                    mode="parallel",
                    message=f"正在计算 AI 路线，已处理 {completed} / {total_items}",
                )

        return results

    def _compose_ai_route_result(self, item: dict, ai_result: dict) -> dict:
        unit_price = self._to_number(ai_result.get("unit_price"))
        confidence = self._to_number(ai_result.get("confidence"))
        reasoning = self._sanitize_ai_reasoning(self._to_text(ai_result.get("reasoning")) or "AI 未返回原因说明")
        original_process = self._to_text(item.get("process"))
        original_material = self._to_text(item.get("material"))
        process_guess_raw = self._to_text(ai_result.get("process_guess"))
        material_guess_raw = self._to_text(ai_result.get("material_guess"))
        process_guess = process_guess_raw if not original_process else ""
        material_guess = material_guess_raw if not original_material else ""
        return {
            "ai_route_unit_price": unit_price,
            "ai_route_source": self.ai_route_service.SOURCE_NAME,
            "ai_route_status": "AI独立估价" if unit_price > 0 else self._classify_ai_route_exception(reasoning)[1],
            "ai_route_confidence": confidence,
            "ai_route_reasoning": reasoning,
            "ai_route_process_guess": process_guess,
            "ai_route_material_guess": material_guess,
            "ai_route_ready": unit_price > 0,
        }

    @staticmethod
    def _empty_ai_route_result(message: str, status: str = "AI待补") -> dict:
        return {
            "ai_route_unit_price": 0.0,
            "ai_route_source": AIRouteQuoteService.SOURCE_NAME,
            "ai_route_status": status,
            "ai_route_confidence": 0.0,
            "ai_route_reasoning": message,
            "ai_route_process_guess": "",
            "ai_route_material_guess": "",
            "ai_route_ready": False,
        }

    @staticmethod
    def _sanitize_ai_reasoning(message: str) -> str:
        text = str(message or "").strip()
        lower = text.lower()
        if any(keyword in lower for keyword in ("timeout", "timed out", "readtimeout", "connecttimeout", "httpsconnectionpool", "read timed out")):
            return "模型超时：千问接口未在限定时间内返回结果"
        if any(keyword in lower for keyword in ("proxyerror", "sslerror", "max retries exceeded", "newconnectionerror", "connection aborted", "connection reset")):
            return "AI接口异常：千问接口当前连接不稳定，请稍后重试"
        if any(keyword in lower for keyword in ("invalid chat format", "invalid_request_error", "expected 'text' field", "qwen pricing api error 400")):
            return "AI接口请求格式异常：服务已调整，请重新报价"
        return text

    @classmethod
    def _classify_ai_route_exception(cls, message: str) -> tuple[str, str]:
        text = cls._sanitize_ai_reasoning(message)
        raw_text = str(message or "")
        if text.startswith("模型超时"):
            return text, "模型超时"
        if text.startswith("AI接口异常"):
            return text, "AI接口异常"
        if "api key" in raw_text.lower() or "未配置" in raw_text:
            return "AI未配置：未配置千问 API key，已跳过 AI 报价", "AI未配置"
        return (f"AI待补：{text}" if text else "AI待补：未形成有效报价", "AI待补")

    def _analyze_price_gap(self, item: dict) -> str:
        finance_unit = self._to_number(item.get("finance_route_unit_price"))
        ai_unit = self._to_number(item.get("ai_route_unit_price"))
        finance_source = self._to_text(item.get("finance_route_source"))
        ai_status = self._to_text(item.get("ai_route_status"))
        ai_confidence = self._to_number(item.get("ai_route_confidence"))
        changjiang_unit = self._to_number(item.get("changjiang_route_unit_price"))
        name = self._to_text(item.get("name"))
        process = self._to_text(item.get("process"))
        weight_kg = self._to_number(item.get("weight_kg"))
        material = self._to_text(item.get("material"))

        if finance_unit <= 0 and ai_unit <= 0:
            return "传统参考和 AI 路线都未形成有效价格，通常是重量、材质或采购参考不足。"
        if finance_unit <= 0:
            return "传统路线缺少采购参考，AI 基于材质、重量和工艺复杂度给出独立估价。"
        if ai_unit <= 0:
            return f"AI 路线当前未形成有效报价，原因：{self._to_text(item.get('ai_route_reasoning')) or ai_status or '接口未返回结果'}。"

        gap = ai_unit - finance_unit
        gap_ratio = abs(gap) / max(finance_unit, ai_unit, 1.0)
        reasons: list[str] = []

        if gap_ratio < 0.08:
            reasons.append("两条路线价格接近，说明 AI 判断和传统采购参考基本一致。")
        else:
            if finance_source in {"Excel表格采购价", "金蝶最近采购价"}:
                reasons.append("传统路线来自历史采购价，可能受到批量、供应商、时间点和税口径影响。")
            elif finance_source == "Excel目标价":
                reasons.append("传统路线采用目标价口径，和 AI 的经验成交价判断天然存在差异。")

            if any(token in f"{name} {process}" for token in ("总成", "组件", "绕线", "叠压", "焊接", "测试")):
                reasons.append("AI 将当前物料识别为多工序或总成件，通常会更强调装配、测试和管理加成。")

            if changjiang_unit > 0:
                if ai_unit > max(finance_unit, changjiang_unit) * 1.1:
                    reasons.append("AI 对工艺复杂度或供应商加成判断更高。")
                elif ai_unit < min(finance_unit, changjiang_unit) * 0.9:
                    reasons.append("AI 更偏向批量化采购或标准件价格区间。")

        if not material or material == "未识别" or weight_kg <= 0:
            reasons.append("BOM 基础字段不完整，会放大两条路线的口径差异。")

        if ai_confidence and ai_confidence < 0.55:
            reasons.append("AI 置信度偏低，这条物料建议人工复核。")

        if not reasons:
            reasons.append("两条路线的差异暂不明显，建议结合供应商和采购批量再确认。")
        return " ".join(reasons[:3])

    def _item_status(self, item: dict) -> str:
        finance_unit = self._to_number(item.get("finance_route_unit_price"))
        ai_unit = self._to_number(item.get("ai_route_unit_price"))
        if finance_unit <= 0 and ai_unit <= 0:
            return "待补参数"
        if finance_unit <= 0:
            return "缺传统参考"
        if ai_unit <= 0:
            return "AI待补"
        gap_ratio = abs(ai_unit - finance_unit) / max(finance_unit, ai_unit, 1.0)
        if gap_ratio >= 0.15:
            return "价差待复核"
        return "双路线可比"

    def _fetch_purchase_refs(self, material_codes: list[str]) -> dict[str, dict]:
        refs: dict[str, dict] = {}
        if not self.config.kingdee.is_ready:
            return refs

        for code in list(dict.fromkeys(material_codes))[:80]:
            if not code:
                continue
            try:
                result = self.kingdee_service.purchase_orders(code, limit=1)
            except Exception:
                continue
            if result.ok and result.data.get("rows"):
                refs[code] = copy.deepcopy(result.data["rows"][0])
        return refs

    def _build_summary(self, items: list[dict]) -> dict:
        total_weight = sum(self._to_number(item.get("weight_kg")) * (self._to_number(item.get("qty")) or 1.0) for item in items)
        finance_total = sum(self._to_number(item.get("finance_route_unit_price")) * (self._to_number(item.get("qty")) or 1.0) for item in items)
        ai_total = sum(self._to_number(item.get("ai_route_unit_price")) * (self._to_number(item.get("qty")) or 1.0) for item in items)
        kingdee_reference_count = sum(1 for item in items if self._to_number(item.get("kingdee_reference_price")) > 0)
        uploaded_price_count = sum(1 for item in items if self._to_number(item.get("current_unit_price")) > 0)
        target_price_count = sum(1 for item in items if self._to_number(item.get("target_unit_price")) > 0)
        finance_reference_count = sum(1 for item in items if bool(item.get("finance_route_has_reference")))
        finance_missing_count = sum(1 for item in items if self._to_number(item.get("finance_route_unit_price")) <= 0)
        ai_ready_count = sum(1 for item in items if self._to_number(item.get("ai_route_unit_price")) > 0)
        ai_unavailable_count = sum(1 for item in items if self._to_number(item.get("ai_route_unit_price")) <= 0)
        pending_count = sum(1 for item in items if item.get("status") in {"待补参数", "缺传统参考", "AI待补"})
        high_gap_count = sum(
            1
            for item in items
            if self._to_number(item.get("finance_route_unit_price")) > 0
            and self._to_number(item.get("ai_route_unit_price")) > 0
            and abs(self._to_number(item.get("route_gap_unit_price"))) / max(
                self._to_number(item.get("finance_route_unit_price")),
                self._to_number(item.get("ai_route_unit_price")),
                1.0,
            ) >= 0.15
        )
        return {
            "total_weight": total_weight,
            "finance_total": finance_total,
            "ai_total": ai_total,
            "route_gap_total": ai_total - finance_total,
            "kingdee_reference_count": kingdee_reference_count,
            "uploaded_price_count": uploaded_price_count,
            "target_price_count": target_price_count,
            "finance_reference_count": finance_reference_count,
            "finance_missing_count": finance_missing_count,
            "ai_ready_count": ai_ready_count,
            "ai_unavailable_count": ai_unavailable_count,
            "high_gap_count": high_gap_count,
            "pending_count": pending_count,
        }

    def _estimate_rule_unit_price(self, item: dict) -> float:
        route = self._estimate_changjiang_route(item)
        return self._to_number(route.get("changjiang_route_unit_price"))

    def _source_tag(self, item: dict) -> str:
        tags: list[str] = ["导入"]
        finance_refs: list[str] = []
        if self._to_number(item.get("current_unit_price")) > 0:
            finance_refs.append("Excel")
        if self._to_number(item.get("kingdee_reference_price")) > 0:
            finance_refs.append("金蝶")
        if self._to_number(item.get("target_unit_price")) > 0:
            finance_refs.append("目标")
        if finance_refs:
            tags.append(f"财务:{'/'.join(finance_refs)}")
        if self._to_number(item.get("changjiang_route_unit_price")) > 0:
            tags.append("规则")
        return " / ".join(tags)

    def _estimate_changjiang_route(self, item: dict) -> dict:
        process = self._to_text(item.get("process"))
        defaults = self.PROCESS_DEFAULTS.get(process, {"lossRate": 0.05, "processFactor": 1.1})
        weight_kg = self._to_number(item.get("weight_kg"))
        material_price = self._to_number(item.get("material_price_used")) or self._to_number(item.get("material_price"))
        loss = self._to_number(item.get("loss")) or defaults["lossRate"]
        extra = self._to_number(item.get("extra"))
        material_cost = material_price * weight_kg * (1 + loss)
        process_cost, process_source, process_note = self._estimate_process_cost(item, material_cost, defaults)
        unit_price = material_cost + process_cost + extra
        status = "规则可用" if unit_price > 0 else "待补参数"
        return {
            "material_cost_est": material_cost,
            "process_cost_est": process_cost,
            "rule_estimate_unit_price": unit_price,
            "changjiang_route_unit_price": unit_price,
            "changjiang_route_source": process_source,
            "changjiang_route_note": process_note,
            "changjiang_route_status": status,
        }

    def _estimate_process_cost(self, item: dict, material_cost: float, defaults: dict) -> tuple[float, str, str]:
        process = self._to_text(item.get("process"))
        item_name = self._to_text(item.get("name"))
        combined = f"{item_name} {process}".strip()
        weight_kg = self._to_number(item.get("weight_kg"))

        if any(token in combined for token in ("绕组", "嵌线")):
            return 0.0, "长江规则-绕组差异口径", "单个定子组件不默认套用 22 元/kg，绕组差额仅用于同系列不同 L 规格差异分析"
        if any(token in combined for token in ("模具冲压叠装焊接", "叠装焊接", "铁芯冲压叠压")):
            return weight_kg * 17.0, "长江规则-铁芯叠装", "铁芯冲压叠装焊接按 17 元/kg 粗估"
        if any(token in combined for token in ("烧结切片研磨", "磁钢烧结")):
            return weight_kg * 35.0, "长江规则-磁钢加工", "磁钢烧结切片研磨按 35 元/kg 粗估"
        if any(token in combined for token in ("高压铸造", "低压铸造", "压铸", "铸造")):
            return weight_kg * 8.0, "长江规则-铸造", "铸造类工艺按 8 元/kg 粗估"
        if "冲压" in combined:
            return weight_kg * 4.5, "长江规则-冲压", "常规冲压按 4.5 元/kg 粗估"
        if ("拉伸" in combined and ("机加工" in combined or "焊" in combined)) or any(token in combined for token in ("摩擦焊",)):
            return weight_kg * 18.0, "长江规则-结构件综合", "拉伸+机加工+焊接综合按 18 元/kg 粗估"
        if any(token in combined for token in ("拉伸件", "拉伸开模件")):
            return weight_kg * 18.0, "长江规则-拉伸综合", "拉伸结构件按 18 元/kg 粗估"

        factor_delta = max(self._to_number(defaults.get("processFactor")) - 1.0, 0.0)
        if material_cost > 0 and factor_delta > 0:
            return material_cost * factor_delta, "长江规则-默认系数兜底", "未命中明确工艺规则，按默认系数估算工艺附加"
        return 0.0, "长江规则-待补工艺", "未命中工艺规则，当前只保留材料成本"

    def _infer_process_and_material(self, name: str, spec: str, material_text: str, process_text: str) -> tuple[str, str]:
        if process_text.strip():
            return process_text.strip(), self._material_alias(material_text or name or spec)
        combined = " ".join(part for part in (name, spec, material_text) if part).lower()
        for rule in self.PROCESS_RULES:
            if any(keyword.lower() in combined for keyword in rule["keywords"]):
                return rule["process"], rule["material"]
        return "", self._material_alias(material_text or name or spec)

    def _latest_raw_price(self, material_input: str) -> dict:
        payload = self._load_demo_payload()
        alias = self._material_alias(material_input)
        live = payload.get("market_prices_cny_per_kg", {}).get(alias)
        if live:
            return {"price": self._to_number(live), "source": "live_or_merged"}

        trend_items = payload.get("trend_items", [])
        matched = []
        for item in trend_items:
            if self._material_alias(item.get("material", "")) == alias and self._to_number(item.get("raw_price_26_jan_feb")) > 0:
                matched.append(item)
        if not matched:
            return {"price": 0.0, "source": "none"}
        picked = max(matched, key=lambda entry: self._to_number(entry.get("raw_price_26_jan_feb")))
        return {"price": self._to_number(picked.get("raw_price_26_jan_feb")), "source": "trend"}

    def _load_demo_payload(self) -> dict:
        if self._demo_payload is None:
            with self.config.demo_data_path.open("r", encoding="utf-8") as handle:
                self._demo_payload = json.load(handle)
        return self._demo_payload

    @staticmethod
    def _material_alias(text: str) -> str:
        normalized = str(text or "").lower()
        if not normalized.strip():
            return ""
        if "铜" in normalized:
            return "铜"
        if any(token in normalized for token in ("硅钢", "35w", "b30")):
            return "硅钢"
        if "铝" in normalized:
            return "铝"
        if any(token in normalized for token in ("镨钕", "钕铁硼", "n35", "n38", "n40", "n42", "n45", "n48", "n50", "n52", "磁")):
            return "磁材"
        if any(token in normalized for token in ("20cr", "钢")):
            return "钢材"
        return "未识别"

    @staticmethod
    def _to_text(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _to_ratio(value, default: float) -> float:
        if value is None or value == "":
            return default
        text = str(value).strip()
        if not text:
            return default
        if text.endswith("%"):
            return ExcelQuoteService._to_number(text[:-1]) / 100.0
        numeric = ExcelQuoteService._to_number(text)
        if numeric > 1:
            return numeric / 100.0
        return numeric or default

    @staticmethod
    def _to_number(value) -> float:
        if value is None or value == "":
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(",", "").replace("，", "")
        text = text.replace("¥", "").replace("元", "")
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            match = re.search(r"-?\d+(?:\.\d+)?", text)
            return float(match.group(0)) if match else 0.0
