#!/usr/bin/env python3
import argparse
import csv
from collections import defaultdict
from pathlib import Path

try:
    import xlsxwriter
except ImportError:  # pragma: no cover - optional dependency
    xlsxwriter = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None


COL_PRODUCT = "\u4ea7\u54c1"
COL_ITEM = "\u7269\u6599"
COL_CODE = "\u7269\u6599\u7f16\u7801"
COL_MATERIAL = "\u6750\u8d28"
COL_PROCESS = "\u5de5\u827a"
COL_BASE_SOURCE = "\u57fa\u7840\u4ef7\u683c\u6765\u6e90"
COL_BASE_TOTAL = "\u57fa\u7840\u91d1\u989d"
COL_PROCESS_LABEL = "\u5de5\u827a\u5355\u4ef7"
COL_PROCESS_TOTAL = "\u5de5\u827a\u91d1\u989d"
COL_LINE_TOTAL = "\u884c\u603b\u4ef7"


SCENARIO_LABELS = {
    "baseline": "\u57fa\u51c6\u7248",
    "conservative": "\u4fdd\u5b88\u91cf\u4ea7\u7248",
    "aggressive": "\u6fc0\u8fdb\u91cf\u4ea7\u7248",
}


VOLUME_TIERS = (
    {
        "key": "tier_1",
        "label": "≤300套/年",
        "min": 0,
        "max": 300,
        "material_discounts": {
            "baseline": {},
            "conservative": {
                "外购标准件": 0.01,
                "漆包线/铜导体": 0.005,
                "铜材": 0.005,
                "铝材": 0.005,
                "铸铝": 0.005,
                "硅钢": 0.005,
                "钢材": 0.005,
                "不锈钢": 0.005,
                "磁材原料": 0.003,
            },
            "aggressive": {
                "外购标准件": 0.03,
                "漆包线/铜导体": 0.01,
                "铜材": 0.01,
                "铝材": 0.01,
                "铸铝": 0.01,
                "硅钢": 0.01,
                "钢材": 0.01,
                "不锈钢": 0.01,
                "磁材原料": 0.005,
            },
        },
        "process_discounts": {
            "baseline": {},
            "conservative": {
                "模具/批量工艺": 0.03,
                "通用批量工艺": 0.02,
                "机加工/综合成形": 0.02,
                "烧结/磁材后加工": 0.01,
                "定子成品工序包": 0.03,
            },
            "aggressive": {
                "模具/批量工艺": 0.06,
                "通用批量工艺": 0.05,
                "机加工/综合成形": 0.04,
                "烧结/磁材后加工": 0.03,
                "定子成品工序包": 0.05,
            },
        },
    },
    {
        "key": "tier_2",
        "label": "301-1000套/年",
        "min": 301,
        "max": 1000,
        "material_discounts": {
            "baseline": {},
            "conservative": {
                "外购标准件": 0.02,
                "漆包线/铜导体": 0.01,
                "铜材": 0.01,
                "铝材": 0.01,
                "铸铝": 0.01,
                "硅钢": 0.01,
                "钢材": 0.01,
                "不锈钢": 0.01,
                "磁材原料": 0.005,
            },
            "aggressive": {
                "外购标准件": 0.04,
                "漆包线/铜导体": 0.02,
                "铜材": 0.02,
                "铝材": 0.02,
                "铸铝": 0.02,
                "硅钢": 0.02,
                "钢材": 0.02,
                "不锈钢": 0.02,
                "磁材原料": 0.01,
            },
        },
        "process_discounts": {
            "baseline": {},
            "conservative": {
                "模具/批量工艺": 0.04,
                "通用批量工艺": 0.03,
                "机加工/综合成形": 0.03,
                "烧结/磁材后加工": 0.02,
                "定子成品工序包": 0.04,
            },
            "aggressive": {
                "模具/批量工艺": 0.08,
                "通用批量工艺": 0.06,
                "机加工/综合成形": 0.05,
                "烧结/磁材后加工": 0.04,
                "定子成品工序包": 0.08,
            },
        },
    },
    {
        "key": "tier_3",
        "label": "1001-3000套/年",
        "min": 1001,
        "max": 3000,
        "material_discounts": {
            "baseline": {},
            "conservative": {
                "外购标准件": 0.03,
                "漆包线/铜导体": 0.01,
                "铜材": 0.01,
                "铝材": 0.01,
                "铸铝": 0.01,
                "硅钢": 0.01,
                "钢材": 0.01,
                "不锈钢": 0.01,
                "磁材原料": 0.005,
            },
            "aggressive": {
                "外购标准件": 0.06,
                "漆包线/铜导体": 0.02,
                "铜材": 0.02,
                "铝材": 0.02,
                "铸铝": 0.02,
                "硅钢": 0.02,
                "钢材": 0.02,
                "不锈钢": 0.02,
                "磁材原料": 0.01,
            },
        },
        "process_discounts": {
            "baseline": {},
            "conservative": {
                "模具/批量工艺": 0.05,
                "通用批量工艺": 0.04,
                "机加工/综合成形": 0.04,
                "烧结/磁材后加工": 0.03,
                "定子成品工序包": 0.05,
            },
            "aggressive": {
                "模具/批量工艺": 0.10,
                "通用批量工艺": 0.08,
                "机加工/综合成形": 0.06,
                "烧结/磁材后加工": 0.05,
                "定子成品工序包": 0.10,
            },
        },
    },
    {
        "key": "tier_4",
        "label": "3001-5000套/年",
        "min": 3001,
        "max": 5000,
        "material_discounts": {
            "baseline": {},
            "conservative": {
                "外购标准件": 0.04,
                "漆包线/铜导体": 0.02,
                "铜材": 0.02,
                "铝材": 0.02,
                "铸铝": 0.02,
                "硅钢": 0.02,
                "钢材": 0.02,
                "不锈钢": 0.02,
                "磁材原料": 0.01,
            },
            "aggressive": {
                "外购标准件": 0.08,
                "漆包线/铜导体": 0.03,
                "铜材": 0.03,
                "铝材": 0.03,
                "铸铝": 0.03,
                "硅钢": 0.03,
                "钢材": 0.03,
                "不锈钢": 0.03,
                "磁材原料": 0.015,
            },
        },
        "process_discounts": {
            "baseline": {},
            "conservative": {
                "模具/批量工艺": 0.06,
                "通用批量工艺": 0.05,
                "机加工/综合成形": 0.04,
                "烧结/磁材后加工": 0.03,
                "定子成品工序包": 0.06,
            },
            "aggressive": {
                "模具/批量工艺": 0.12,
                "通用批量工艺": 0.10,
                "机加工/综合成形": 0.08,
                "烧结/磁材后加工": 0.06,
                "定子成品工序包": 0.12,
            },
        },
    },
    {
        "key": "tier_5",
        "label": "5001-20000套/年",
        "min": 5001,
        "max": 20000,
        "material_discounts": {
            "baseline": {},
            "conservative": {
                "外购标准件": 0.06,
                "漆包线/铜导体": 0.03,
                "铜材": 0.03,
                "铝材": 0.03,
                "铸铝": 0.03,
                "硅钢": 0.03,
                "钢材": 0.03,
                "不锈钢": 0.03,
                "磁材原料": 0.02,
            },
            "aggressive": {
                "外购标准件": 0.12,
                "漆包线/铜导体": 0.05,
                "铜材": 0.05,
                "铝材": 0.05,
                "铸铝": 0.05,
                "硅钢": 0.05,
                "钢材": 0.05,
                "不锈钢": 0.05,
                "磁材原料": 0.04,
            },
        },
        "process_discounts": {
            "baseline": {},
            "conservative": {
                "模具/批量工艺": 0.12,
                "通用批量工艺": 0.10,
                "机加工/综合成形": 0.08,
                "烧结/磁材后加工": 0.06,
                "定子成品工序包": 0.10,
            },
            "aggressive": {
                "模具/批量工艺": 0.22,
                "通用批量工艺": 0.18,
                "机加工/综合成形": 0.14,
                "烧结/磁材后加工": 0.10,
                "定子成品工序包": 0.18,
            },
        },
    },
    {
        "key": "tier_6",
        "label": "20001-50000套/年",
        "min": 20001,
        "max": 50000,
        "material_discounts": {
            "baseline": {},
            "conservative": {
                "外购标准件": 0.08,
                "漆包线/铜导体": 0.04,
                "铜材": 0.04,
                "铝材": 0.04,
                "铸铝": 0.04,
                "硅钢": 0.04,
                "钢材": 0.04,
                "不锈钢": 0.04,
                "磁材原料": 0.03,
            },
            "aggressive": {
                "外购标准件": 0.15,
                "漆包线/铜导体": 0.07,
                "铜材": 0.07,
                "铝材": 0.07,
                "铸铝": 0.07,
                "硅钢": 0.07,
                "钢材": 0.07,
                "不锈钢": 0.07,
                "磁材原料": 0.05,
            },
        },
        "process_discounts": {
            "baseline": {},
            "conservative": {
                "模具/批量工艺": 0.15,
                "通用批量工艺": 0.12,
                "机加工/综合成形": 0.10,
                "烧结/磁材后加工": 0.08,
                "定子成品工序包": 0.12,
            },
            "aggressive": {
                "模具/批量工艺": 0.26,
                "通用批量工艺": 0.22,
                "机加工/综合成形": 0.18,
                "烧结/磁材后加工": 0.14,
                "定子成品工序包": 0.22,
            },
        },
    },
    {
        "key": "tier_7",
        "label": ">50000套/年",
        "min": 50001,
        "max": None,
        "material_discounts": {
            "baseline": {},
            "conservative": {
                "外购标准件": 0.10,
                "漆包线/铜导体": 0.05,
                "铜材": 0.05,
                "铝材": 0.05,
                "铸铝": 0.05,
                "硅钢": 0.05,
                "钢材": 0.05,
                "不锈钢": 0.05,
                "磁材原料": 0.04,
            },
            "aggressive": {
                "外购标准件": 0.18,
                "漆包线/铜导体": 0.08,
                "铜材": 0.08,
                "铝材": 0.08,
                "铸铝": 0.08,
                "硅钢": 0.08,
                "钢材": 0.08,
                "不锈钢": 0.08,
                "磁材原料": 0.06,
            },
        },
        "process_discounts": {
            "baseline": {},
            "conservative": {
                "模具/批量工艺": 0.18,
                "通用批量工艺": 0.15,
                "机加工/综合成形": 0.12,
                "烧结/磁材后加工": 0.10,
                "定子成品工序包": 0.15,
            },
            "aggressive": {
                "模具/批量工艺": 0.32,
                "通用批量工艺": 0.26,
                "机加工/综合成形": 0.22,
                "烧结/磁材后加工": 0.18,
                "定子成品工序包": 0.26,
            },
        },
    },
)



def to_float(value: str | None) -> float:
    text = (value or "").strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def pct_text(value: float) -> str:
    return f"{value * 100:.1f}%"


def money_text(value: float) -> str:
    return f"{value:,.2f}"


def resolve_volume_tier(annual_volume: int) -> dict:
    for tier in VOLUME_TIERS:
        max_volume = tier["max"]
        if annual_volume >= tier["min"] and (max_volume is None or annual_volume <= max_volume):
            return tier
    return VOLUME_TIERS[-1]


def read_rows(path: Path) -> list[dict]:
    for encoding in ("utf-8-sig", "gb18030", "gbk"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                return list(csv.DictReader(f))
        except Exception:
            continue
    raise RuntimeError(f"Failed to read csv: {path}")


def write_csv(path: Path, headers: list[str], rows: list[list]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def contains_any(text: str, keywords: list[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def classify_material(row: dict) -> str:
    source = (row.get(COL_BASE_SOURCE) or "").strip()
    material = (row.get(COL_MATERIAL) or "").strip()
    item = (row.get(COL_ITEM) or "").strip()
    base_total = to_float(row.get(COL_BASE_TOTAL))

    if base_total <= 0:
        return "\u4e0d\u8ba1\u5165"
    if source == "\u6807\u51c6\u4ef6\u4ef7\u683c":
        return "\u5916\u8d2d\u6807\u51c6\u4ef6"
    if source.startswith("SMM \u957f") or source.startswith("SMM \u955d") or "\u9568\u94d5\u91d1\u5c5e" in source:
        return "\u78c1\u6750\u539f\u6599"
    if source.startswith("\u957f\u6c5f \u6f06\u5305\u7ebf"):
        return "\u6f06\u5305\u7ebf/\u94dc\u5bfc\u4f53"
    if source.startswith("\u957f\u6c5f 1#\u7535\u89e3\u94dc"):
        return "\u94dc\u6750"
    if source.startswith("\u957f\u6c5f \u94ddA00"):
        return "\u94dd\u6750"
    if source.startswith("\u957f\u6c5f A356.2"):
        return "\u94f8\u94dd"
    if source.startswith("SMM \u4e0a\u6d77\u5b9d\u94a2 B50A470"):
        return "\u7845\u94a2"
    if source == "\u94a2\u6750\u4ee3\u7406\u4ef7":
        return "\u94a2\u6750"
    if source == "304/2B\u4ee3\u7406":
        return "\u4e0d\u9508\u94a2"
    if material in {"T2", "\u7d2b\u94dc"} or "\u94dc" in material:
        return "\u94dc\u6750"
    if contains_any(item, ["\u6cb9\u6cf5", "\u6362\u70ed\u5668", "\u5feb\u63d2\u8fde\u63a5\u5668"]):
        return "\u5916\u8d2d\u6807\u51c6\u4ef6"
    return "\u5176\u4ed6\u6750\u6599"


def classify_process(row: dict) -> str:
    process = (row.get(COL_PROCESS) or "").strip()
    process_label = (row.get(COL_PROCESS_LABEL) or "").strip()
    process_total = to_float(row.get(COL_PROCESS_TOTAL))

    if process_total <= 0:
        return "\u65e0\u5de5\u827a"
    if process_label == "\u5b9a\u5b50\u6210\u54c1\u5de5\u5e8f\u5305":
        return "\u5b9a\u5b50\u6210\u54c1\u5de5\u5e8f\u5305"
    if process in {
        "\u51b2\u538b",
        "\u51b7\u9566",
        "\u51b2\u538b+\u6298\u5f2f",
        "\u51b2\u538b\u3001\u6298\u5f2f\u7b49",
        "\u6ce8\u5851",
        "\u6324\u538b\u6210\u578b",
        "\u538b\u94f8",
        "\u538b\u94f8+\u786b\u5316",
        "\u9ad8\u538b\u94f8\u9020",
        "\u4f4e\u538b\u94f8\u9020",
    }:
        return "\u6a21\u5177/\u6279\u91cf\u5de5\u827a"
    if process in {
        "\u6fc0\u5149\u5207\u5272+\u6298\u5f2f",
        "\u7ebf\u5207\u5272+\u6298\u5f2f",
        "\u5706\u7ebf",
        "\u6d82\u8986",
    }:
        return "\u901a\u7528\u6279\u91cf\u5de5\u827a"
    if contains_any(process, ["\u673a\u52a0\u5de5", "\u6469\u64e6\u710a", "\u62c9\u4f38"]):
        return "\u673a\u52a0\u5de5/\u7efc\u5408\u6210\u5f62"
    if contains_any(process, ["\u70e7\u7ed3"]):
        return "\u70e7\u7ed3/\u78c1\u6750\u540e\u52a0\u5de5"
    return "\u5176\u4ed6\u5de5\u827a"


def apply_scenario(row: dict, scenario: str, annual_volume: int) -> dict:
    base_material = to_float(row.get(COL_BASE_TOTAL))
    base_process = to_float(row.get(COL_PROCESS_TOTAL))
    material_category = classify_material(row)
    process_category = classify_process(row)
    tier = resolve_volume_tier(annual_volume)
    material_discount = tier["material_discounts"][scenario].get(material_category, 0.0)
    process_discount = tier["process_discounts"][scenario].get(process_category, 0.0)
    scenario_material = base_material * (1.0 - material_discount)
    scenario_process = base_process * (1.0 - process_discount)
    scenario_total = scenario_material + scenario_process
    return {
        "volume_tier_key": tier["key"],
        "volume_tier_label": tier["label"],
        "material_category": material_category,
        "process_category": process_category,
        "material_discount": material_discount,
        "process_discount": process_discount,
        "material_total": scenario_material,
        "process_total": scenario_process,
        "line_total": scenario_total,
        "saving_total": (base_material + base_process) - scenario_total,
    }


def aggregate_by_product(rows: list[dict], annual_volume: int) -> list[dict]:
    results: list[dict] = []
    by_product: dict[str, list[dict]] = defaultdict(list)
    tier = resolve_volume_tier(annual_volume)
    for row in rows:
        by_product[row[COL_PRODUCT]].append(row)

    for product, product_rows in sorted(by_product.items()):
        baseline_material = sum(to_float(row.get(COL_BASE_TOTAL)) for row in product_rows)
        baseline_process = sum(to_float(row.get(COL_PROCESS_TOTAL)) for row in product_rows)
        baseline_total = sum(to_float(row.get(COL_LINE_TOTAL)) for row in product_rows)
        entry = {
            COL_PRODUCT: product,
            "baseline_material": baseline_material,
            "baseline_process": baseline_process,
            "baseline_total": baseline_total,
            "annual_volume": annual_volume,
            "volume_tier_key": tier["key"],
            "volume_tier_label": tier["label"],
        }
        for scenario in ("conservative", "aggressive"):
            scenario_rows = [apply_scenario(row, scenario, annual_volume) for row in product_rows]
            mat_total = sum(row["material_total"] for row in scenario_rows)
            proc_total = sum(row["process_total"] for row in scenario_rows)
            total = sum(row["line_total"] for row in scenario_rows)
            saving = baseline_total - total
            entry[f"{scenario}_material"] = mat_total
            entry[f"{scenario}_process"] = proc_total
            entry[f"{scenario}_total"] = total
            entry[f"{scenario}_saving"] = saving
            entry[f"{scenario}_saving_rate"] = (saving / baseline_total) if baseline_total else 0.0
            entry[f"{scenario}_annual_total"] = total * annual_volume
            entry[f"{scenario}_annual_saving"] = saving * annual_volume
        entry["baseline_annual_total"] = baseline_total * annual_volume
        results.append(entry)
    return results


def build_mix_row(product_rows: list[dict], annual_volume: int) -> dict:
    count = len(product_rows) or 1
    tier = resolve_volume_tier(annual_volume)
    mix = {
        COL_PRODUCT: "\u7b49\u6bd4\u4f8b\u6df7\u4ea7\u5e73\u5747",
        "annual_volume": annual_volume,
        "volume_tier_key": tier["key"],
        "volume_tier_label": tier["label"],
    }
    for key in (
        "baseline_material",
        "baseline_process",
        "baseline_total",
        "baseline_annual_total",
        "conservative_material",
        "conservative_process",
        "conservative_total",
        "conservative_saving",
        "conservative_saving_rate",
        "conservative_annual_total",
        "conservative_annual_saving",
        "aggressive_material",
        "aggressive_process",
        "aggressive_total",
        "aggressive_saving",
        "aggressive_saving_rate",
        "aggressive_annual_total",
        "aggressive_annual_saving",
    ):
        if key.endswith("_rate"):
            continue
        mix[key] = sum(row.get(key, 0.0) for row in product_rows) / count
    baseline_total = mix["baseline_total"]
    mix["conservative_saving_rate"] = (mix["conservative_saving"] / baseline_total) if baseline_total else 0.0
    mix["aggressive_saving_rate"] = (mix["aggressive_saving"] / baseline_total) if baseline_total else 0.0
    mix["baseline_annual_total"] = mix["baseline_total"] * annual_volume
    mix["conservative_annual_total"] = mix["conservative_total"] * annual_volume
    mix["conservative_annual_saving"] = mix["conservative_saving"] * annual_volume
    mix["aggressive_annual_total"] = mix["aggressive_total"] * annual_volume
    mix["aggressive_annual_saving"] = mix["aggressive_saving"] * annual_volume
    return mix


def build_saving_categories(rows: list[dict], scenario: str, annual_volume: int) -> list[tuple[str, float]]:
    totals: dict[str, float] = defaultdict(float)
    for row in rows:
        result = apply_scenario(row, scenario, annual_volume)
        if result["saving_total"] <= 0:
            continue
        totals[result["material_category"]] += to_float(row.get(COL_BASE_TOTAL)) - result["material_total"]
        totals[result["process_category"]] += to_float(row.get(COL_PROCESS_TOTAL)) - result["process_total"]
    return sorted(totals.items(), key=lambda item: item[1], reverse=True)


def build_detail_rows(rows: list[dict], annual_volume: int) -> list[list]:
    detail_rows = []
    for row in rows:
        conservative = apply_scenario(row, "conservative", annual_volume)
        aggressive = apply_scenario(row, "aggressive", annual_volume)
        baseline_total = to_float(row.get(COL_LINE_TOTAL))
        detail_rows.append(
            [
                row.get(COL_PRODUCT, ""),
                row.get(COL_ITEM, ""),
                row.get(COL_CODE, ""),
                conservative["volume_tier_label"],
                row.get(COL_MATERIAL, ""),
                row.get(COL_PROCESS, ""),
                conservative["material_category"],
                conservative["process_category"],
                to_float(row.get(COL_BASE_TOTAL)),
                to_float(row.get(COL_PROCESS_TOTAL)),
                baseline_total,
                conservative["material_discount"],
                conservative["process_discount"],
                conservative["line_total"],
                baseline_total - conservative["line_total"],
                aggressive["material_discount"],
                aggressive["process_discount"],
                aggressive["line_total"],
                baseline_total - aggressive["line_total"],
            ]
        )
    return detail_rows


def fit_columns(worksheet, rows: list[list[str]], min_width: int = 10, max_width: int = 36) -> None:
    if not rows:
        return
    width_count = max(len(row) for row in rows)
    for col in range(width_count):
        width = min_width
        for row in rows:
            if col >= len(row):
                continue
            width = max(width, len(str(row[col])) + 2)
        worksheet.set_column(col, col, min(width, max_width))


def write_table(workbook, worksheet, headers: list[str], rows: list[list], title: str) -> None:
    title_fmt = workbook.add_format({"bold": True, "font_size": 14, "font_name": "Microsoft YaHei"})
    header_fmt = workbook.add_format(
        {
            "bold": True,
            "bg_color": "#DCE6F1",
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
            "font_name": "Microsoft YaHei",
        }
    )
    text_fmt = workbook.add_format({"border": 1, "font_name": "Microsoft YaHei"})
    money_fmt = workbook.add_format({"border": 1, "num_format": "#,##0.00", "font_name": "Microsoft YaHei"})
    percent_fmt = workbook.add_format({"border": 1, "num_format": "0.0%", "font_name": "Microsoft YaHei"})

    worksheet.write(0, 0, title, title_fmt)
    for col, header in enumerate(headers):
        worksheet.write(2, col, header, header_fmt)
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, value in enumerate(row):
            if isinstance(value, float):
                if 0 <= value <= 1 and headers[col_idx].endswith("\u964d\u5e45"):
                    worksheet.write_number(row_idx, col_idx, value, percent_fmt)
                elif headers[col_idx].endswith("\u6298\u6263"):
                    worksheet.write_number(row_idx, col_idx, value, percent_fmt)
                else:
                    worksheet.write_number(row_idx, col_idx, value, money_fmt)
            elif isinstance(value, int):
                worksheet.write_number(row_idx, col_idx, value, text_fmt)
            else:
                worksheet.write(row_idx, col_idx, value, text_fmt)
    worksheet.freeze_panes(3, 0)
    worksheet.autofilter(2, 0, 2 + len(rows), len(headers) - 1)
    fit_columns(worksheet, [headers] + [[str(v) for v in row] for row in rows])


def export_xlsx(summary_rows: list[dict], mix_row: dict, detail_rows: list[list], output_path: Path, annual_volume: int) -> None:
    if xlsxwriter is None:
        _export_xlsx_openpyxl(summary_rows, mix_row, detail_rows, output_path, annual_volume)
        return

    workbook = xlsxwriter.Workbook(str(output_path))
    overview = workbook.add_worksheet("\u60c5\u666f\u603b\u89c8")
    product_ws = workbook.add_worksheet("\u4ea7\u54c1\u5bf9\u6bd4")
    detail_ws = workbook.add_worksheet("\u884c\u7ea7\u964d\u672c")

    title_fmt = workbook.add_format({"bold": True, "font_size": 18, "font_name": "Microsoft YaHei", "font_color": "#17365D"})
    label_fmt = workbook.add_format({"bold": True, "border": 1, "bg_color": "#EAF2F8", "font_name": "Microsoft YaHei"})
    money_fmt = workbook.add_format({"border": 1, "num_format": "#,##0.00", "font_name": "Microsoft YaHei"})
    percent_fmt = workbook.add_format({"border": 1, "num_format": "0.0%", "font_name": "Microsoft YaHei"})
    note_fmt = workbook.add_format({"border": 1, "text_wrap": True, "valign": "top", "font_name": "Microsoft YaHei"})

    tier = resolve_volume_tier(annual_volume)
    overview.write(0, 0, f"\u4e09\u4e00380\u7535\u673a {annual_volume}\u5957/\u5e74\u91cf\u4ea7\u964d\u672c\u6a21\u578b", title_fmt)
    overview.write(2, 0, "\u5e74\u4ea7\u91cf\u53e3\u5f84", label_fmt)
    overview.write(2, 1, annual_volume, money_fmt)
    overview.write(3, 0, "\u91cf\u4ea7\u5206\u6863", label_fmt)
    overview.write(3, 1, tier["label"], note_fmt)
    overview.write(4, 0, "\u57fa\u51c6\u7248\u5e73\u5747\u5355\u5957", label_fmt)
    overview.write_number(4, 1, mix_row["baseline_total"], money_fmt)
    overview.write(5, 0, "\u4fdd\u5b88\u7248\u5e73\u5747\u5355\u5957", label_fmt)
    overview.write_number(5, 1, mix_row["conservative_total"], money_fmt)
    overview.write(6, 0, "\u4fdd\u5b88\u7248\u964d\u5e45", label_fmt)
    overview.write_number(6, 1, mix_row["conservative_saving_rate"], percent_fmt)
    overview.write(7, 0, "\u6fc0\u8fdb\u7248\u5e73\u5747\u5355\u5957", label_fmt)
    overview.write_number(7, 1, mix_row["aggressive_total"], money_fmt)
    overview.write(8, 0, "\u6fc0\u8fdb\u7248\u964d\u5e45", label_fmt)
    overview.write_number(8, 1, mix_row["aggressive_saving_rate"], percent_fmt)

    overview.write(2, 3, "\u8bf4\u660e", label_fmt)
    overview.merge_range(
        3,
        3,
        9,
        8,
        f"\u57fa\u51c6\u7248\u4fdd\u7559\u5f53\u524d BOM \u4f30\u4ef7\u7ed3\u679c\u3002\u4fdd\u5b88\u7248\u53ea\u5bf9\u6613\u8bae\u4ef7\u7684\u5916\u8d2d\u4ef6\u548c\u6279\u91cf\u5de5\u827a\u505a\u5c0f\u5e45\u4e0b\u8c03\u3002\u6fc0\u8fdb\u7248\u8fdb\u4e00\u6b65\u8ba1\u5165\u5e74\u6846\u91c7\u8d2d\u3001\u6a21\u5177/\u5de5\u88c5\u644a\u8584\u4e0e\u8282\u62cd\u4f18\u5316\u7684\u7efc\u5408\u964d\u672c\u6548\u679c\u3002\u5f53\u524d\u5e74\u4ea7\u91cf {annual_volume}\u5957/\u5e74\uff0c\u547d\u4e2d {tier['label']} \u5206\u6863\u3002",
        note_fmt,
    )

    product_headers = [
        "\u4ea7\u54c1",
        "\u91cf\u4ea7\u5206\u6863",
        "\u57fa\u51c6\u5355\u5957\u603b\u6210\u672c",
        "\u4fdd\u5b88\u5355\u5957\u603b\u6210\u672c",
        "\u4fdd\u5b88\u5355\u5957\u964d\u672c\u989d",
        "\u4fdd\u5b88\u964d\u5e45",
        "\u6fc0\u8fdb\u5355\u5957\u603b\u6210\u672c",
        "\u6fc0\u8fdb\u5355\u5957\u964d\u672c\u989d",
        "\u6fc0\u8fdb\u964d\u5e45",
        "\u57fa\u51c6\u5e74\u603b\u6210\u672c",
        "\u4fdd\u5b88\u5e74\u603b\u6210\u672c",
        "\u4fdd\u5b88\u5e74\u964d\u672c\u989d",
        "\u6fc0\u8fdb\u5e74\u603b\u6210\u672c",
        "\u6fc0\u8fdb\u5e74\u964d\u672c\u989d",
    ]
    product_table = []
    for row in summary_rows + [mix_row]:
        product_table.append(
            [
                row[COL_PRODUCT],
                row.get("volume_tier_label", ""),
                row["baseline_total"],
                row["conservative_total"],
                row["conservative_saving"],
                row["conservative_saving_rate"],
                row["aggressive_total"],
                row["aggressive_saving"],
                row["aggressive_saving_rate"],
                row["baseline_annual_total"],
                row["conservative_annual_total"],
                row["conservative_annual_saving"],
                row["aggressive_annual_total"],
                row["aggressive_annual_saving"],
            ]
        )
    write_table(workbook, product_ws, product_headers, product_table, "\u5355\u5957\u4e0e\u5e74\u6210\u672c\u5bf9\u6bd4")

    detail_headers = [
        "\u4ea7\u54c1",
        "\u7269\u6599",
        "\u7269\u6599\u7f16\u7801",
        "\u91cf\u4ea7\u5206\u6863",
        "\u6750\u8d28",
        "\u5de5\u827a",
        "\u6750\u6599\u7c7b\u522b",
        "\u5de5\u827a\u7c7b\u522b",
        "\u57fa\u51c6\u6750\u6599",
        "\u57fa\u51c6\u5de5\u827a",
        "\u57fa\u51c6\u603b\u4ef7",
        "\u4fdd\u5b88\u6750\u6599\u6298\u6263",
        "\u4fdd\u5b88\u5de5\u827a\u6298\u6263",
        "\u4fdd\u5b88\u603b\u4ef7",
        "\u4fdd\u5b88\u964d\u672c\u989d",
        "\u6fc0\u8fdb\u6750\u6599\u6298\u6263",
        "\u6fc0\u8fdb\u5de5\u827a\u6298\u6263",
        "\u6fc0\u8fdb\u603b\u4ef7",
        "\u6fc0\u8fdb\u964d\u672c\u989d",
    ]
    write_table(workbook, detail_ws, detail_headers, detail_rows, "\u884c\u7ea7\u964d\u672c\u6620\u5c04")

    workbook.close()


def _export_xlsx_openpyxl(summary_rows: list[dict], mix_row: dict, detail_rows: list[list], output_path: Path, annual_volume: int) -> None:
    if Workbook is None or PatternFill is None or Border is None or Side is None or Font is None or Alignment is None:
        raise RuntimeError("量产测算导出需要 xlsxwriter 或 openpyxl")

    workbook = Workbook()
    overview = workbook.active
    overview.title = "情景总览"
    product_ws = workbook.create_sheet("产品对比")
    detail_ws = workbook.create_sheet("行级降本")

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    label_fill = PatternFill("solid", fgColor="EAF2F8")
    title_font = Font(bold=True, size=18, color="17365D")
    section_font = Font(bold=True, size=14)

    def apply_cell(cell, *, bold=False, fill=None):
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if bold:
            cell.font = Font(bold=True)
        if fill is not None:
            cell.fill = fill

    tier = resolve_volume_tier(annual_volume)
    overview["A1"] = f"三一380电机 {annual_volume}套/年量产降本模型"
    overview["A1"].font = title_font
    labels = [
        ("A3", "年产量口径", annual_volume, "0"),
        ("A4", "量产分档", tier["label"], "@"),
        ("A5", "基准版平均单套", mix_row["baseline_total"], "#,##0.00"),
        ("A6", "保守版平均单套", mix_row["conservative_total"], "#,##0.00"),
        ("A7", "保守版降幅", mix_row["conservative_saving_rate"], "0.0%"),
        ("A8", "激进版平均单套", mix_row["aggressive_total"], "#,##0.00"),
        ("A9", "激进版降幅", mix_row["aggressive_saving_rate"], "0.0%"),
    ]
    for label_ref, label_text, value, num_fmt in labels:
        value_ref = label_ref.replace("A", "B")
        overview[label_ref] = label_text
        apply_cell(overview[label_ref], bold=True, fill=label_fill)
        overview[value_ref] = value
        overview[value_ref].number_format = num_fmt
        apply_cell(overview[value_ref])
    overview["D3"] = "说明"
    apply_cell(overview["D3"], bold=True, fill=label_fill)
    overview["D4"] = f"基准版保留当前 BOM 估价。保守版只对易议价材料和批量工艺小幅下调。激进版进一步计入年框采购、工装摊薄与节拍优化。当前年产量 {annual_volume} 套/年，命中 {tier['label']} 分档。"
    apply_cell(overview["D4"])

    product_headers = [
        "产品", "量产分档", "基准单套总成本", "保守单套总成本", "保守单套降本额", "保守降幅",
        "激进单套总成本", "激进单套降本额", "激进降幅", "基准年总成本",
        "保守年总成本", "保守年降本额", "激进年总成本", "激进年降本额",
    ]
    product_table = []
    for row in summary_rows + [mix_row]:
        product_table.append([
            row[COL_PRODUCT],
            row.get("volume_tier_label", ""),
            row["baseline_total"],
            row["conservative_total"],
            row["conservative_saving"],
            row["conservative_saving_rate"],
            row["aggressive_total"],
            row["aggressive_saving"],
            row["aggressive_saving_rate"],
            row["baseline_annual_total"],
            row["conservative_annual_total"],
            row["conservative_annual_saving"],
            row["aggressive_annual_total"],
            row["aggressive_annual_saving"],
        ])
    _write_openpyxl_table(product_ws, product_headers, product_table, "单套与年成本对比", header_fill, border)

    detail_headers = [
        "产品", "物料", "物料编码", "量产分档", "材质", "工艺", "材料类别", "工艺类别", "基准材料", "基准工艺",
        "基准总价", "保守材料折扣", "保守工艺折扣", "保守总价", "保守降本额",
        "激进材料折扣", "激进工艺折扣", "激进总价", "激进降本额",
    ]
    _write_openpyxl_table(detail_ws, detail_headers, detail_rows, "行级降本映射", header_fill, border)

    workbook.save(output_path)


def _write_openpyxl_table(worksheet, headers: list[str], rows: list[list], title: str, header_fill, border) -> None:
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=14)
    for idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=3, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, float):
                header_text = headers[col_idx - 1]
                if "降幅" in header_text or "折扣" in header_text:
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "#,##0.00"
    worksheet.freeze_panes = "A4"


def render_markdown(summary_rows: list[dict], mix_row: dict, conservative_categories: list[tuple[str, float]], aggressive_categories: list[tuple[str, float]], annual_volume: int) -> str:
    tier = resolve_volume_tier(annual_volume)
    lines = [
        f"# \u4e09\u4e00380\u7535\u673a {annual_volume}\u5957/\u5e74\u91cf\u4ea7\u964d\u672c\u6a21\u578b",
        "",
        "## \u6a21\u578b\u53e3\u5f84",
        "",
        "- `\u57fa\u51c6\u7248`\uff1a\u4fdd\u7559\u5f53\u524d BOM \u8054\u7f51\u4f30\u4ef7\u7ed3\u679c\uff0c\u4e0d\u989d\u5916\u52a0\u5165\u91cf\u4ea7\u6298\u6263\u3002",
        "- `\u4fdd\u5b88\u91cf\u4ea7\u7248`\uff1a\u53ea\u5bf9\u6613\u8bae\u4ef7\u7684\u5916\u8d2d\u4ef6\u548c\u90e8\u5206\u6279\u91cf\u5de5\u827a\u505a\u5c0f\u5e45\u4e0b\u8c03\u3002",
        "- `\u6fc0\u8fdb\u91cf\u4ea7\u7248`\uff1a\u8fdb\u4e00\u6b65\u53cd\u6620\u5e74\u6846\u91c7\u8d2d\u3001\u6a21\u5177/\u5de5\u88c5\u644a\u8584\u3001\u751f\u4ea7\u8282\u62cd\u4f18\u5316\u5bf9\u6210\u672c\u7684\u5f71\u54cd\u3002",
        f"- \u5e74\u4ea7\u91cf\u53e3\u5f84\uff1a`{annual_volume}` \u5957/\u5e74\u3002",
        f"- \u5f53\u524d\u91cf\u4ea7\u5206\u6863\uff1a`{tier['label']}`\u3002",
        "",
        "## \u6838\u5fc3\u7ed3\u8bba",
        "",
        f"- \u7b49\u6bd4\u4f8b\u6df7\u4ea7\u5e73\u5747\u5355\u5957\u57fa\u51c6\u6210\u672c\uff1a`{money_text(mix_row['baseline_total'])} \u5143`\u3002",
        f"- 保守版平均单套成本：`{money_text(mix_row['conservative_total'])} 元`，比基准版下降 `{money_text(mix_row['conservative_saving'])} 元/套`，降幅 `{pct_text(mix_row['conservative_saving_rate'])}`。",
        f"- 激进版平均单套成本：`{money_text(mix_row['aggressive_total'])} 元`，比基准版下降 `{money_text(mix_row['aggressive_saving'])} 元/套`，降幅 `{pct_text(mix_row['aggressive_saving_rate'])}`。",
        f"- 如果 4 个叠高型号等比例混产、全年合计 `{annual_volume}` 套，保守版年降本约 `{money_text(mix_row['conservative_annual_saving'])} 元`，激进版年降本约 `{money_text(mix_row['aggressive_annual_saving'])} 元`。",
        "",
        "## \u4ea7\u54c1\u5bf9\u6bd4",
        "",
        "| \u4ea7\u54c1 | \u57fa\u51c6\u7248 | \u4fdd\u5b88\u7248 | \u4fdd\u5b88\u964d\u5e45 | \u6fc0\u8fdb\u7248 | \u6fc0\u8fdb\u964d\u5e45 | |",
        "| --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for row in summary_rows:
        lines.append(
            f"| {row[COL_PRODUCT]} | {money_text(row['baseline_total'])} | {money_text(row['conservative_total'])} | {pct_text(row['conservative_saving_rate'])} | {money_text(row['aggressive_total'])} | {pct_text(row['aggressive_saving_rate'])} |"
        )
    lines.extend(
        [
            "",
            "## \u4e3b\u8981\u964d\u672c\u6765\u6e90",
            "",
            "### \u4fdd\u5b88\u7248",
            "",
        ]
    )
    for name, amount in conservative_categories[:8]:
        lines.append(f"- {name}\uff1a`{money_text(amount)} \u5143`")
    lines.extend(
        [
            "",
            "### \u6fc0\u8fdb\u7248",
            "",
        ]
    )
    for name, amount in aggressive_categories[:8]:
        lines.append(f"- {name}\uff1a`{money_text(amount)} \u5143`")
    lines.extend(
        [
            "",
            "## \u6298\u6263\u5047\u8bbe",
            "",
            f"\u5f53\u524d\u5e74\u4ea7\u91cf `{annual_volume}` \u5957/\u5e74\uff0c\u91c7\u7528 `{tier['label']}` \u91cf\u4ea7\u5206\u6863\u3002",
            "",
            "### \u6750\u6599",
            "",
            f"- \u4fdd\u5b88\u7248\uff1a\u5916\u8d2d\u6807\u51c6\u4ef6 `-{pct_text(tier['material_discounts']['conservative'].get('外购标准件', 0.0))}`\uff0c\u94dc/\u94dd/\u94f8\u94dd/\u94a2/\u7845\u94a2/\u6f06\u5305\u7ebf `-{pct_text(tier['material_discounts']['conservative'].get('铜材', 0.0))}`\uff0c\u78c1\u6750\u539f\u6599 `-{pct_text(tier['material_discounts']['conservative'].get('磁材原料', 0.0))}`\u3002",
            f"- \u6fc0\u8fdb\u7248\uff1a\u5916\u8d2d\u6807\u51c6\u4ef6 `-{pct_text(tier['material_discounts']['aggressive'].get('外购标准件', 0.0))}`\uff0c\u94dc/\u94dd/\u94f8\u94dd/\u94a2/\u7845\u94a2/\u6f06\u5305\u7ebf `-{pct_text(tier['material_discounts']['aggressive'].get('铜材', 0.0))}`\uff0c\u78c1\u6750\u539f\u6599 `-{pct_text(tier['material_discounts']['aggressive'].get('磁材原料', 0.0))}`\u3002",
            "",
            "### \u5de5\u827a",
            "",
            f"- \u4fdd\u5b88\u7248\uff1a\u6a21\u5177/\u6279\u91cf\u5de5\u827a `-{pct_text(tier['process_discounts']['conservative'].get('模具/批量工艺', 0.0))}`\uff0c\u901a\u7528\u6279\u91cf\u5de5\u827a `-{pct_text(tier['process_discounts']['conservative'].get('通用批量工艺', 0.0))}`\uff0c\u673a\u52a0\u5de5/\u7efc\u5408\u6210\u5f62 `-{pct_text(tier['process_discounts']['conservative'].get('机加工/综合成形', 0.0))}`\uff0c\u70e7\u7ed3/\u78c1\u6750\u540e\u52a0\u5de5 `-{pct_text(tier['process_discounts']['conservative'].get('烧结/磁材后加工', 0.0))}`\uff0c\u5b9a\u5b50\u6210\u54c1\u5de5\u5e8f\u5305 `-{pct_text(tier['process_discounts']['conservative'].get('定子成品工序包', 0.0))}`\u3002",
            f"- \u6fc0\u8fdb\u7248\uff1a\u6a21\u5177/\u6279\u91cf\u5de5\u827a `-{pct_text(tier['process_discounts']['aggressive'].get('模具/批量工艺', 0.0))}`\uff0c\u901a\u7528\u6279\u91cf\u5de5\u827a `-{pct_text(tier['process_discounts']['aggressive'].get('通用批量工艺', 0.0))}`\uff0c\u673a\u52a0\u5de5/\u7efc\u5408\u6210\u5f62 `-{pct_text(tier['process_discounts']['aggressive'].get('机加工/综合成形', 0.0))}`\uff0c\u70e7\u7ed3/\u78c1\u6750\u540e\u52a0\u5de5 `-{pct_text(tier['process_discounts']['aggressive'].get('烧结/磁材后加工', 0.0))}`\uff0c\u5b9a\u5b50\u6210\u54c1\u5de5\u5e8f\u5305 `-{pct_text(tier['process_discounts']['aggressive'].get('定子成品工序包', 0.0))}`\u3002",
            "",
            "## \u4f7f\u7528\u5efa\u8bae",
            "",
            "- \u5982\u679c\u4f60\u73b0\u5728\u8981\u505a\u5e74\u5ea6\u9884\u7b97\u6216\u5185\u90e8\u7acb\u9879\uff0c\u53ef\u5148\u7528 `\u4fdd\u5b88\u7248`\u3002",
            "- \u5982\u679c\u4f60\u5728\u8c08\u5e74\u6846\u91c7\u8d2d\u3001\u6b63\u5728\u505a\u91cf\u4ea7\u722c\u5761\u76ee\u6807\uff0c\u53ef\u4ee5\u540c\u65f6\u53c2\u8003 `\u6fc0\u8fdb\u7248`\u3002",
            "- \u7531\u4e8e\u5f53\u524d BOM \u91cc\u6ca1\u6709\u5355\u5217\u6a21\u5177\u6295\u8d44\uff0c\u6fc0\u8fdb\u7248\u662f\u7528\u66f4\u9ad8\u7684\u6279\u91cf\u5de5\u827a\u964d\u672c\u7cfb\u6570\u53bb\u8fd1\u4f3c\u53cd\u6620\u5de5\u88c5/\u8282\u62cd\u644a\u8584\u6548\u679c\uff0c\u4e0d\u662f\u4ece BOM \u91cc\u76f4\u63a5\u6263\u9664\u4e00\u6761\u663e\u5f0f\u7684\u6a21\u5177\u8d39\u3002",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Build annual volume cost-down scenarios from line pricing csv.")
    parser.add_argument("line_csv", help="Path to line pricing csv")
    parser.add_argument("--annual-volume", type=int, default=20000, help="Annual production volume")
    parser.add_argument("--output-dir", help="Directory for generated files")
    args = parser.parse_args()

    line_csv = Path(args.line_csv).resolve()
    output_dir = Path(args.output_dir).resolve() if args.output_dir else line_csv.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    stem = line_csv.stem.replace("-\u884c\u7ea7\u62a5\u4ef7", "")
    summary_csv = output_dir / f"{stem}-\u91cf\u4ea7\u964d\u672c\u6c47\u603b.csv"
    detail_csv = output_dir / f"{stem}-\u91cf\u4ea7\u964d\u672c\u660e\u7ec6.csv"
    summary_md = output_dir / f"{stem}-\u91cf\u4ea7\u964d\u672c\u8bf4\u660e.md"
    summary_xlsx = output_dir / f"{stem}-\u91cf\u4ea7\u964d\u672c\u5bf9\u6bd4.xlsx"

    rows = read_rows(line_csv)
    product_rows = aggregate_by_product(rows, args.annual_volume)
    mix_row = build_mix_row(product_rows, args.annual_volume)
    conservative_categories = build_saving_categories(rows, "conservative", args.annual_volume)
    aggressive_categories = build_saving_categories(rows, "aggressive", args.annual_volume)
    detail_rows = build_detail_rows(rows, args.annual_volume)

    summary_headers = [
        "\u4ea7\u54c1",
        "\u91cf\u4ea7\u5206\u6863",
        "\u57fa\u51c6\u5355\u5957\u6750\u6599",
        "\u57fa\u51c6\u5355\u5957\u5de5\u827a",
        "\u57fa\u51c6\u5355\u5957\u603b\u6210\u672c",
        "\u4fdd\u5b88\u5355\u5957\u603b\u6210\u672c",
        "\u4fdd\u5b88\u5355\u5957\u964d\u672c\u989d",
        "\u4fdd\u5b88\u964d\u5e45",
        "\u6fc0\u8fdb\u5355\u5957\u603b\u6210\u672c",
        "\u6fc0\u8fdb\u5355\u5957\u964d\u672c\u989d",
        "\u6fc0\u8fdb\u964d\u5e45",
        f"\u57fa\u51c6\u5e74\u603b\u6210\u672c_{args.annual_volume}\u5957",
        f"\u4fdd\u5b88\u5e74\u603b\u6210\u672c_{args.annual_volume}\u5957",
        f"\u4fdd\u5b88\u5e74\u964d\u672c\u989d_{args.annual_volume}\u5957",
        f"\u6fc0\u8fdb\u5e74\u603b\u6210\u672c_{args.annual_volume}\u5957",
        f"\u6fc0\u8fdb\u5e74\u964d\u672c\u989d_{args.annual_volume}\u5957",
    ]
    summary_table = []
    for row in product_rows + [mix_row]:
        summary_table.append(
            [
                row[COL_PRODUCT],
                row.get("volume_tier_label", ""),
                row["baseline_material"],
                row["baseline_process"],
                row["baseline_total"],
                row["conservative_total"],
                row["conservative_saving"],
                row["conservative_saving_rate"],
                row["aggressive_total"],
                row["aggressive_saving"],
                row["aggressive_saving_rate"],
                row["baseline_annual_total"],
                row["conservative_annual_total"],
                row["conservative_annual_saving"],
                row["aggressive_annual_total"],
                row["aggressive_annual_saving"],
            ]
        )
    write_csv(summary_csv, summary_headers, summary_table)

    detail_headers = [
        "\u4ea7\u54c1",
        "\u7269\u6599",
        "\u7269\u6599\u7f16\u7801",
        "\u91cf\u4ea7\u5206\u6863",
        "\u6750\u8d28",
        "\u5de5\u827a",
        "\u6750\u6599\u7c7b\u522b",
        "\u5de5\u827a\u7c7b\u522b",
        "\u57fa\u51c6\u6750\u6599",
        "\u57fa\u51c6\u5de5\u827a",
        "\u57fa\u51c6\u603b\u4ef7",
        "\u4fdd\u5b88\u6750\u6599\u6298\u6263",
        "\u4fdd\u5b88\u5de5\u827a\u6298\u6263",
        "\u4fdd\u5b88\u603b\u4ef7",
        "\u4fdd\u5b88\u964d\u672c\u989d",
        "\u6fc0\u8fdb\u6750\u6599\u6298\u6263",
        "\u6fc0\u8fdb\u5de5\u827a\u6298\u6263",
        "\u6fc0\u8fdb\u603b\u4ef7",
        "\u6fc0\u8fdb\u964d\u672c\u989d",
    ]
    write_csv(detail_csv, detail_headers, detail_rows)

    summary_md.write_text(
        render_markdown(product_rows, mix_row, conservative_categories, aggressive_categories, args.annual_volume),
        encoding="utf-8-sig",
    )
    export_xlsx(product_rows, mix_row, detail_rows, summary_xlsx, args.annual_volume)

    print(f"SCENARIO_SUMMARY_CSV_EXPORTED={summary_csv}")
    print(f"SCENARIO_DETAIL_CSV_EXPORTED={detail_csv}")
    print(f"SCENARIO_SUMMARY_MD_EXPORTED={summary_md}")
    print(f"SCENARIO_XLSX_EXPORTED={summary_xlsx}")


if __name__ == "__main__":
    main()



