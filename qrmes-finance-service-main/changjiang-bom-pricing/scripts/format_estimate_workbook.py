#!/usr/bin/env python3
import argparse
import csv
from collections import defaultdict
from pathlib import Path

try:
    import xlsxwriter
except ImportError:  # pragma: no cover - optional dependency
    xlsxwriter = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None


def read_csv_rows(path: Path) -> list[dict]:
    for encoding in ("utf-8-sig", "gb18030", "gbk"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                return list(csv.DictReader(f))
        except Exception:
            continue
    raise RuntimeError(f"Failed to read csv: {path}")


def to_float(value: str | None) -> float:
    text = (value or "").strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def trim_markdown_prefix(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith("- "):
        return stripped[2:]
    if stripped.startswith("#"):
        return stripped.lstrip("#").strip()
    return stripped


def read_md_sections(path: Path) -> list[tuple[str, list[str]]]:
    lines = path.read_text(encoding="utf-8-sig").splitlines()
    sections: list[tuple[str, list[str]]] = []
    current_title = "说明"
    current_lines: list[str] = []
    for line in lines:
        if line.startswith("## "):
            if current_lines:
                sections.append((current_title, current_lines))
            current_title = line[3:].strip()
            current_lines = []
            continue
        if line.strip():
            current_lines.append(trim_markdown_prefix(line))
    if current_lines:
        sections.append((current_title, current_lines))
    return sections


def build_source_summary(line_rows: list[dict]) -> list[tuple[str, float]]:
    totals: dict[str, float] = defaultdict(float)
    for row in line_rows:
        totals[row["基础价格来源"]] += to_float(row["基础金额"])
    return sorted(totals.items(), key=lambda item: item[1], reverse=True)


def build_process_summary(line_rows: list[dict]) -> list[tuple[str, float]]:
    totals: dict[str, float] = defaultdict(float)
    for row in line_rows:
        totals[row["工艺"] or "无工艺"] += to_float(row["工艺金额"])
    return sorted(totals.items(), key=lambda item: item[1], reverse=True)


def fit_columns(worksheet, rows: list[list[str]], min_width: int = 10, max_width: int = 48) -> None:
    if not rows:
        return
    col_count = max(len(row) for row in rows)
    for col in range(col_count):
        width = min_width
        for row in rows:
            if col >= len(row):
                continue
            value = "" if row[col] is None else str(row[col])
            width = max(width, len(value) + 2)
        worksheet.set_column(col, col, min(width, max_width))


def write_sheet_table(workbook, worksheet, headers: list[str], rows: list[list], title: str | None = None, autofilter: bool = True) -> None:
    title_fmt = workbook.add_format(
        {"bold": True, "font_size": 14, "font_name": "Microsoft YaHei", "font_color": "#1F1F1F"}
    )
    header_fmt = workbook.add_format(
        {
            "bold": True,
            "bg_color": "#DCE6F1",
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
            "font_name": "Microsoft YaHei",
        }
    )
    text_fmt = workbook.add_format({"border": 1, "valign": "top", "font_name": "Microsoft YaHei"})
    money_fmt = workbook.add_format({"border": 1, "num_format": "#,##0.00", "valign": "top", "font_name": "Microsoft YaHei"})
    int_fmt = workbook.add_format({"border": 1, "num_format": "0", "valign": "top", "font_name": "Microsoft YaHei"})

    start_row = 0
    if title:
        worksheet.write(0, 0, title, title_fmt)
        start_row = 2
    for col, header in enumerate(headers):
        worksheet.write(start_row, col, header, header_fmt)
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row):
            if isinstance(value, int):
                worksheet.write_number(row_idx, col_idx, value, int_fmt)
            elif isinstance(value, float):
                worksheet.write_number(row_idx, col_idx, value, money_fmt)
            else:
                worksheet.write(row_idx, col_idx, value, text_fmt)
    worksheet.freeze_panes(start_row + 1, 0)
    if autofilter and rows:
        worksheet.autofilter(start_row, 0, start_row + len(rows), len(headers) - 1)
    fit_columns(worksheet, [headers] + [[str(cell) for cell in row] for row in rows])


def export_workbook(line_csv: Path, grouped_csv: Path, summary_md: Path, output_xlsx: Path) -> None:
    line_rows = read_csv_rows(line_csv)
    grouped_rows = read_csv_rows(grouped_csv)
    md_sections = read_md_sections(summary_md)

    if xlsxwriter is None:
        _export_workbook_openpyxl(line_rows, grouped_rows, md_sections, line_csv, output_xlsx)
        return

    total_estimate = sum(to_float(row["基础估算总价"]) for row in grouped_rows)
    total_material = sum(to_float(row["基础材料合计"]) for row in grouped_rows)
    total_process = sum(to_float(row["基础工艺合计"]) for row in grouped_rows)
    source_summary = build_source_summary(line_rows)
    process_summary = build_process_summary(line_rows)

    workbook = xlsxwriter.Workbook(str(output_xlsx))
    workbook.set_custom_property("ReportType", "Estimate Workbook")

    title_fmt = workbook.add_format(
        {"bold": True, "font_size": 18, "font_name": "Microsoft YaHei", "font_color": "#17365D"}
    )
    label_fmt = workbook.add_format(
        {"bold": True, "font_name": "Microsoft YaHei", "bg_color": "#EAF2F8", "border": 1}
    )
    money_fmt = workbook.add_format(
        {"font_name": "Microsoft YaHei", "num_format": "#,##0.00", "border": 1, "font_size": 12}
    )
    text_fmt = workbook.add_format({"font_name": "Microsoft YaHei", "border": 1, "valign": "top"})
    note_fmt = workbook.add_format(
        {"font_name": "Microsoft YaHei", "text_wrap": True, "valign": "top", "border": 1}
    )
    section_fmt = workbook.add_format(
        {"bold": True, "font_name": "Microsoft YaHei", "font_size": 13, "font_color": "#17365D"}
    )
    header_fmt = workbook.add_format(
        {
            "bold": True,
            "bg_color": "#DCE6F1",
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
            "font_name": "Microsoft YaHei",
        }
    )
    percent_fmt = workbook.add_format(
        {"font_name": "Microsoft YaHei", "num_format": "0.00%", "border": 1, "font_size": 12}
    )

    overview = workbook.add_worksheet("估价总览")
    overview.write(0, 0, f"{line_csv.stem.replace('-行级报价', '')} 估价结果总览", title_fmt)
    overview.write(2, 0, "估价总额", label_fmt)
    overview.write_number(2, 1, total_estimate, money_fmt)
    overview.write(3, 0, "材料合计", label_fmt)
    overview.write_number(3, 1, total_material, money_fmt)
    overview.write(4, 0, "工艺合计", label_fmt)
    overview.write_number(4, 1, total_process, money_fmt)
    overview.write(5, 0, "工艺占比", label_fmt)
    overview.write_number(5, 1, (total_process / total_estimate) if total_estimate else 0, percent_fmt)

    overview.write(2, 3, "说明", section_fmt)
    overview.merge_range(3, 3, 5, 7, "本表为 AI 估价结果，沿用已有 skill 的材料映射和工艺启发式规则，仅优化呈现方式。", note_fmt)

    overview.write(7, 0, "TOP成本项", section_fmt)
    top_headers = ["物料", "物料编码", "估价总价", "材料", "工艺", "重量kg"]
    top_rows = [
        [
            row["物料"],
            row["物料编码"],
            to_float(row["基础估算总价"]),
            to_float(row["基础材料合计"]),
            to_float(row["基础工艺合计"]),
            to_float(row["总重量kg"]),
        ]
        for row in grouped_rows[:12]
    ]
    for col, header in enumerate(top_headers):
        overview.write(8, col, header, header_fmt)
    for row_idx, row in enumerate(top_rows, start=9):
        for col_idx, value in enumerate(row):
            if isinstance(value, float):
                overview.write_number(row_idx, col_idx, value, money_fmt)
            else:
                overview.write(row_idx, col_idx, value, text_fmt)

    overview.write(7, 8, "材料来源占比", section_fmt)
    for col, header in enumerate(["价格来源", "金额"], start=8):
        overview.write(8, col, header, header_fmt)
    for row_idx, (source, amount) in enumerate(source_summary[:10], start=9):
        overview.write(row_idx, 8, source, text_fmt)
        overview.write_number(row_idx, 9, amount, money_fmt)

    overview.write(7, 11, "工艺成本占比", section_fmt)
    for col, header in enumerate(["工艺", "金额"], start=11):
        overview.write(8, col, header, header_fmt)
    for row_idx, (process, amount) in enumerate(process_summary[:10], start=9):
        overview.write(row_idx, 11, process, text_fmt)
        overview.write_number(row_idx, 12, amount, money_fmt)

    overview.freeze_panes(9, 0)
    overview.set_column(0, 0, 16)
    overview.set_column(1, 1, 14)
    overview.set_column(3, 7, 18)
    overview.set_column(8, 12, 18)

    grouped_ws = workbook.add_worksheet("物料汇总")
    grouped_headers = ["产品", "物料", "物料编码", "估价总价", "材料合计", "工艺合计", "总重量kg", "主要材质/工艺明细"]
    grouped_table_rows = [
        [
            row["产品"],
            row["物料"],
            row["物料编码"],
            to_float(row["基础估算总价"]),
            to_float(row["基础材料合计"]),
            to_float(row["基础工艺合计"]),
            to_float(row["总重量kg"]),
            row["主要材质/工艺明细"],
        ]
        for row in grouped_rows
    ]
    write_sheet_table(workbook, grouped_ws, grouped_headers, grouped_table_rows, title="按物料编码汇总的估价结果")
    grouped_ws.conditional_format(
        3,
        3,
        2 + len(grouped_table_rows),
        3,
        {"type": "3_color_scale", "min_color": "#E2F0D9", "mid_color": "#FFF2CC", "max_color": "#F4CCCC"},
    )

    detail_ws = workbook.add_worksheet("行级明细")
    detail_headers = [
        "产品",
        "物料",
        "物料编码",
        "材质",
        "工艺",
        "数量",
        "单件重量kg",
        "延展重量kg",
        "价格类型",
        "价格来源",
        "基础单价",
        "基础金额",
        "工艺复杂度",
        "工艺单价",
        "工艺金额",
        "行总价",
    ]
    detail_rows = [
        [
            row["产品"],
            row["物料"],
            row["物料编码"],
            row["材质"],
            row["工艺"],
            to_float(row["数量"]),
            to_float(row["重量kg"]),
            to_float(row["延展重量kg"]),
            row["基础价格类型"],
            row["基础价格来源"],
            to_float(row["基础单价"]),
            to_float(row["基础金额"]),
            row["工艺复杂度"],
            row["工艺单价"],
            to_float(row["工艺金额"]),
            to_float(row["行总价"]),
        ]
        for row in line_rows
    ]
    write_sheet_table(workbook, detail_ws, detail_headers, detail_rows, title="逐行估价明细")
    detail_ws.conditional_format(
        3,
        15,
        2 + len(detail_rows),
        15,
        {"type": "3_color_scale", "min_color": "#E2F0D9", "mid_color": "#FFF2CC", "max_color": "#F4CCCC"},
    )

    ai_ws = workbook.add_worksheet("AI结论")
    ai_ws.write(0, 0, "AI分析摘录", title_fmt)
    row_cursor = 2
    for title, lines in md_sections:
        ai_ws.write(row_cursor, 0, title, section_fmt)
        row_cursor += 1
        for line in lines:
            ai_ws.write(row_cursor, 0, line, note_fmt)
            row_cursor += 1
        row_cursor += 1
    ai_ws.set_column(0, 0, 120)

    workbook.close()


def _export_workbook_openpyxl(
    line_rows: list[dict],
    grouped_rows: list[dict],
    md_sections: list[tuple[str, list[str]]],
    line_csv: Path,
    output_xlsx: Path,
) -> None:
    if Workbook is None or PatternFill is None or Border is None or Side is None or Font is None or Alignment is None:
        raise RuntimeError("导出格式化工作簿需要 xlsxwriter 或 openpyxl")

    total_estimate = sum(to_float(row["基础估算总价"]) for row in grouped_rows)
    total_material = sum(to_float(row["基础材料合计"]) for row in grouped_rows)
    total_process = sum(to_float(row["基础工艺合计"]) for row in grouped_rows)
    source_summary = build_source_summary(line_rows)
    process_summary = build_process_summary(line_rows)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "估价总览"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    section_font = Font(bold=True, color="17365D", size=13)
    title_font = Font(bold=True, color="17365D", size=18)
    label_fill = PatternFill("solid", fgColor="EAF2F8")

    def apply_table_cell(cell, *, bold=False, fill=None):
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if bold:
            cell.font = Font(bold=True)
        if fill is not None:
            cell.fill = fill

    overview["A1"] = f"{line_csv.stem.replace('-行级报价', '')} 估价结果总览"
    overview["A1"].font = title_font
    overview["A3"] = "估价总额"
    overview["A4"] = "材料合计"
    overview["A5"] = "工艺合计"
    overview["A6"] = "工艺占比"
    for ref in ("A3", "A4", "A5", "A6"):
        apply_table_cell(overview[ref], bold=True, fill=label_fill)
    overview["B3"] = total_estimate
    overview["B4"] = total_material
    overview["B5"] = total_process
    overview["B6"] = (total_process / total_estimate) if total_estimate else 0
    overview["B3"].number_format = "#,##0.00"
    overview["B4"].number_format = "#,##0.00"
    overview["B5"].number_format = "#,##0.00"
    overview["B6"].number_format = "0.00%"
    for ref in ("B3", "B4", "B5", "B6"):
        apply_table_cell(overview[ref])

    overview["D3"] = "说明"
    overview["D3"].font = section_font
    overview["D4"] = "本表为 AI 估价结果，沿用已有 skill 的材料映射和工艺启发式规则，仅优化呈现方式。"
    apply_table_cell(overview["D4"])

    top_headers = ["物料", "物料编码", "估价总价", "材料", "工艺", "重量kg"]
    overview["A8"] = "TOP成本项"
    overview["A8"].font = section_font
    for idx, header in enumerate(top_headers, start=1):
        cell = overview.cell(row=9, column=idx, value=header)
        apply_table_cell(cell, bold=True, fill=header_fill)
    top_rows = [
        [
            row["物料"],
            row["物料编码"],
            to_float(row["基础估算总价"]),
            to_float(row["基础材料合计"]),
            to_float(row["基础工艺合计"]),
            to_float(row["总重量kg"]),
        ]
        for row in grouped_rows[:12]
    ]
    for row_idx, row in enumerate(top_rows, start=10):
        for col_idx, value in enumerate(row, start=1):
            cell = overview.cell(row=row_idx, column=col_idx, value=value)
            apply_table_cell(cell)
            if isinstance(value, float):
                cell.number_format = "#,##0.00"

    overview["I8"] = "材料来源占比"
    overview["L8"] = "工艺成本占比"
    overview["I8"].font = section_font
    overview["L8"].font = section_font
    for base_col, headers in ((9, ["价格来源", "金额"]), (12, ["工艺", "金额"])):
        for idx, header in enumerate(headers, start=base_col):
            cell = overview.cell(row=9, column=idx, value=header)
            apply_table_cell(cell, bold=True, fill=header_fill)
    for row_idx, (source, amount) in enumerate(source_summary[:10], start=10):
        apply_table_cell(overview.cell(row=row_idx, column=9, value=source))
        amount_cell = overview.cell(row=row_idx, column=10, value=amount)
        amount_cell.number_format = "#,##0.00"
        apply_table_cell(amount_cell)
    for row_idx, (process, amount) in enumerate(process_summary[:10], start=10):
        apply_table_cell(overview.cell(row=row_idx, column=12, value=process))
        amount_cell = overview.cell(row=row_idx, column=13, value=amount)
        amount_cell.number_format = "#,##0.00"
        apply_table_cell(amount_cell)

    grouped_ws = workbook.create_sheet("物料汇总")
    grouped_headers = ["产品", "物料", "物料编码", "估价总价", "材料合计", "工艺合计", "总重量kg", "主要材质/工艺明细"]
    grouped_table_rows = [
        [
            row["产品"],
            row["物料"],
            row["物料编码"],
            to_float(row["基础估算总价"]),
            to_float(row["基础材料合计"]),
            to_float(row["基础工艺合计"]),
            to_float(row["总重量kg"]),
            row["主要材质/工艺明细"],
        ]
        for row in grouped_rows
    ]
    _write_openpyxl_table(grouped_ws, grouped_headers, grouped_table_rows, "按物料编码汇总的估价结果", header_fill, border)

    detail_ws = workbook.create_sheet("行级明细")
    detail_headers = [
        "产品", "物料", "物料编码", "材质", "工艺", "数量", "单件重量kg", "延展重量kg",
        "价格类型", "价格来源", "基础单价", "基础金额", "工艺复杂度", "工艺单价", "工艺金额", "行总价",
    ]
    detail_rows = [
        [
            row["产品"], row["物料"], row["物料编码"], row["材质"], row["工艺"],
            to_float(row["数量"]), to_float(row["重量kg"]), to_float(row["延展重量kg"]),
            row["基础价格类型"], row["基础价格来源"], to_float(row["基础单价"]),
            to_float(row["基础金额"]), row["工艺复杂度"], row["工艺单价"],
            to_float(row["工艺金额"]), to_float(row["行总价"]),
        ]
        for row in line_rows
    ]
    _write_openpyxl_table(detail_ws, detail_headers, detail_rows, "逐行估价明细", header_fill, border)

    ai_ws = workbook.create_sheet("AI结论")
    ai_ws["A1"] = "AI分析摘录"
    ai_ws["A1"].font = title_font
    row_cursor = 3
    for title, lines in md_sections:
        ai_ws.cell(row=row_cursor, column=1, value=title).font = section_font
        row_cursor += 1
        for line in lines:
            cell = ai_ws.cell(row=row_cursor, column=1, value=line)
            apply_table_cell(cell)
            row_cursor += 1
        row_cursor += 1
    ai_ws.column_dimensions["A"].width = 120

    workbook.save(output_xlsx)


def _write_openpyxl_table(worksheet, headers: list[str], rows: list[list], title: str, header_fill, border) -> None:
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=14)
    for idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=3, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, float):
                cell.number_format = "#,##0.00"
    worksheet.freeze_panes = "A4"


def main() -> None:
    parser = argparse.ArgumentParser(description="Reformat estimate csv outputs into a presentation-friendly Excel workbook.")
    parser.add_argument("line_csv", help="Path to the line estimate csv")
    parser.add_argument("grouped_csv", help="Path to the grouped summary csv")
    parser.add_argument("summary_md", help="Path to the markdown summary")
    parser.add_argument("--output-xlsx", required=True, help="Path to write the formatted workbook")
    args = parser.parse_args()

    export_workbook(
        Path(args.line_csv),
        Path(args.grouped_csv),
        Path(args.summary_md),
        Path(args.output_xlsx),
    )
    print(f"FORMATTED_XLSX_EXPORTED={args.output_xlsx}")


if __name__ == "__main__":
    main()



