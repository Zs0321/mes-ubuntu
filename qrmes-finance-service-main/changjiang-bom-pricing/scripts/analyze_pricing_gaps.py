#!/usr/bin/env python3
import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

try:
    import xlsxwriter
except ImportError:  # pragma: no cover - optional dependency
    xlsxwriter = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None


EXPECTED_HEADER = [
    "\u4ea7\u54c1",
    "\u7269\u6599",
    "\u7269\u6599\u7f16\u7801",
    "\u6750\u8d28",
    "\u5de5\u827a",
    "\u6570\u91cf",
    "\u91cd\u91cfkg",
    "\u5ef6\u5c55\u91cd\u91cfkg",
    "\u57fa\u7840\u4ef7\u683c\u7c7b\u578b",
    "\u57fa\u7840\u4ef7\u683c\u6765\u6e90",
    "\u57fa\u7840\u5355\u4ef7",
    "\u57fa\u7840\u91d1\u989d",
    "\u5de5\u827a\u590d\u6742\u5ea6",
    "\u5de5\u827a\u5355\u4ef7",
    "\u5de5\u827a\u91d1\u989d",
    "\u884c\u603b\u4ef7",
    "\u76ee\u524d\u91c7\u8d2d\u603b\u4ef7",
]

COL_PRODUCT = 0
COL_ITEM = 1
COL_CODE = 2
COL_MATERIAL = 3
COL_PROCESS = 4
COL_QTY = 5
COL_WEIGHT = 6
COL_EXT_WEIGHT = 7
COL_PRICE_TYPE = 8
COL_PRICE_SOURCE = 9
COL_BASE_UNIT = 10
COL_BASE_TOTAL = 11
COL_PROCESS_COMPLEXITY = 12
COL_PROCESS_UNIT = 13
COL_PROCESS_TOTAL = 14
COL_LINE_TOTAL = 15
COL_PURCHASE_TOTAL = 16

REASON_LABELS = {
    "purchase_total_reused_across_multiple_codes": "\u91c7\u8d2d\u603b\u4ef7\u7591\u4f3c\u5728\u591a\u4e2a\u7269\u6599\u7f16\u7801\u95f4\u590d\u7528",
    "finished_goods_scope_exceeds_current_raw_material_estimate": "\u5f53\u524d\u4f30\u7b97\u53ea\u8986\u76d6\u539f\u6750\u6599\u6216\u534a\u6210\u54c1\uff0c\u672a\u8986\u76d6\u6210\u54c1\u91c7\u8d2d\u53e3\u5f84",
    "missing_winding_embedding_varnish_test_process": "\u7f3a\u5c11\u7ed5\u7ebf\u3001\u5d4c\u7ebf\u3001\u6d78\u6f06\u3001\u70d8\u5e72\u3001\u6d4b\u8bd5\u7b49\u5b9a\u5b50\u5173\u952e\u5de5\u5e8f",
    "material_proxy_likely_underprices_core_material": "\u6750\u6599\u4f7f\u7528\u4ee3\u7406\u4ef7\uff0c\u53ef\u80fd\u4f4e\u4f30\u4e86\u771f\u5b9e\u6750\u6599\u6210\u672c",
    "minimum_lot_or_minimum_process_charge_not_captured": "\u672a\u8ba1\u5165\u6700\u5c0f\u8d77\u8ba2\u91cf\u6216\u6700\u4f4e\u52a0\u5de5\u8d77\u6b65\u4ef7",
    "piece_rate_or_process_formula_is_far_below_purchase_floor": "\u6309\u4ef6\u6216\u6309\u5de5\u5e8f\u7684\u4f30\u4ef7\u516c\u5f0f\u660e\u663e\u4f4e\u4e8e\u5b9e\u9645\u91c7\u8d2d\u5e95\u4ef7",
    "purchase_scope_or_supplier_markup_exceeds_current_formula": "\u91c7\u8d2d\u8303\u56f4\u6216\u4f9b\u5e94\u5546\u52a0\u4ef7\u663e\u8457\u9ad8\u4e8e\u5f53\u524d\u516c\u5f0f",
    "purchase_price_outlier_vs_other_stack_heights": "\u76f8\u540c\u96f6\u4ef6\u5728\u5176\u4ed6\u53e0\u9ad8\u4ea7\u54c1\u4e2d\u7684\u91c7\u8d2d\u4ef7\u66f4\u96c6\u4e2d\uff0c\u672c\u6761\u91c7\u8d2d\u4ef7\u660e\u663e\u504f\u79bb",
    "peer_purchase_prices_split_into_multiple_bands": "\u540c\u7c7b\u96f6\u4ef6\u5728\u4e0d\u540c\u53e0\u9ad8\u4ea7\u54c1\u95f4\u5b58\u5728\u591a\u6863\u91c7\u8d2d\u4ef7\uff0c\u7591\u4f3c\u7248\u672c\u6216\u53e3\u5f84\u4e0d\u4e00\u81f4",
}


def contains_any(text: str, keywords: list[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def to_float(value: str) -> float:
    value = (value or "").strip().replace(",", "")
    if not value:
        return 0.0
    try:
        return float(value)
    except ValueError:
        return 0.0


def read_csv_rows(path: Path) -> tuple[str, list[str], list[list[str]]]:
    last_error = None
    for encoding in ("utf-8-sig", "gb18030", "gbk"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                rows = list(csv.reader(f))
            if not rows:
                return encoding, [], []
            header = rows[0]
            if len(header) < len(EXPECTED_HEADER):
                raise ValueError(f"Unexpected header width: {len(header)}")
            return encoding, header, rows[1:]
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"Failed to read {path}: {last_error}")


def parse_row(row: list[str]) -> dict:
    padded = row + [""] * max(0, len(EXPECTED_HEADER) - len(row))
    return {
        "product": padded[COL_PRODUCT].strip(),
        "item": padded[COL_ITEM].strip(),
        "code": padded[COL_CODE].strip(),
        "material": padded[COL_MATERIAL].strip(),
        "process": padded[COL_PROCESS].strip(),
        "qty": padded[COL_QTY].strip(),
        "weight_kg": to_float(padded[COL_WEIGHT]),
        "ext_weight_kg": to_float(padded[COL_EXT_WEIGHT]),
        "price_type": padded[COL_PRICE_TYPE].strip(),
        "price_source": padded[COL_PRICE_SOURCE].strip(),
        "base_unit_price": to_float(padded[COL_BASE_UNIT]),
        "base_total": to_float(padded[COL_BASE_TOTAL]),
        "process_complexity": padded[COL_PROCESS_COMPLEXITY].strip(),
        "process_unit_price": padded[COL_PROCESS_UNIT].strip(),
        "process_total": to_float(padded[COL_PROCESS_TOTAL]),
        "line_total": to_float(padded[COL_LINE_TOTAL]),
        "purchase_total_text": padded[COL_PURCHASE_TOTAL].strip(),
        "purchase_total": to_float(padded[COL_PURCHASE_TOTAL]) if padded[COL_PURCHASE_TOTAL].strip() else None,
    }


def build_entries(rows: list[list[str]], code_filter: str | None) -> list[dict]:
    entries: dict[str, dict] = {}
    for raw in rows:
        row = parse_row(raw)
        if not row["code"]:
            continue
        if code_filter and row["code"] != code_filter:
            continue

        entry = entries.setdefault(
            row["code"],
            {
                "product": row["product"],
                "item": row["item"],
                "code": row["code"],
                "weight_kg": 0.0,
                "ext_weight_kg": 0.0,
                "base_total": 0.0,
                "process_total": 0.0,
                "line_total": 0.0,
                "purchase_total": None,
                "rows": [],
            },
        )
        entry["weight_kg"] += row["weight_kg"]
        entry["ext_weight_kg"] += row["ext_weight_kg"]
        entry["base_total"] += row["base_total"]
        entry["process_total"] += row["process_total"]
        entry["line_total"] += row["line_total"]
        if row["purchase_total"] is not None and entry["purchase_total"] is None:
            entry["purchase_total"] = row["purchase_total"]
        entry["rows"].append(row)

    grouped = list(entries.values())
    for entry in grouped:
        buy = entry["purchase_total"]
        line = entry["line_total"]
        entry["gap"] = (buy - line) if buy is not None else None
        entry["coverage"] = (line / buy) if buy else None
    grouped.sort(key=lambda x: (x["gap"] is None, -(x["gap"] or 0), x["code"]))
    return grouped


def build_purchase_repetition_map(entries: list[dict]) -> dict[tuple[str, float], list[tuple[str, str, str]]]:
    repeated: dict[tuple[str, float], list[tuple[str, str, str]]] = defaultdict(list)
    for entry in entries:
        if entry["purchase_total"] is None:
            continue
        repeated[(entry["product"], entry["purchase_total"])].append(
            (entry["product"], entry["item"], entry["code"])
        )
    return repeated


def estimate_stator_finished_goods(entry: dict) -> dict | None:
    if "\u5b9a\u5b50\u7ec4\u4ef6" not in entry["item"]:
        return None

    steel_row = None
    copper_row = None
    for row in entry["rows"]:
        material = row["material"]
        if steel_row is None and contains_any(material, ["\u94c1\u82af", "\u7845\u94a2"]):
            steel_row = row
        if copper_row is None and contains_any(material, ["\u94dc\u7ebf", "\u6241\u94dc", "\u6f06\u5305"]):
            copper_row = row

    if steel_row is None or copper_row is None:
        return None

    silicon_steel_price = 9.50
    winding_process_per_copper_kg = 220.0
    auxiliary_on_copper = 0.12
    auxiliary_per_total_kg = 8.0
    assembly_and_test_fixed = 5500.0

    silicon_correction = max(
        0.0,
        (silicon_steel_price - steel_row["base_unit_price"]) * steel_row["ext_weight_kg"],
    )
    winding_conversion = copper_row["ext_weight_kg"] * winding_process_per_copper_kg
    auxiliaries = (
        copper_row["base_total"] * auxiliary_on_copper
        + entry["ext_weight_kg"] * auxiliary_per_total_kg
    )
    total = (
        entry["line_total"]
        + silicon_correction
        + winding_conversion
        + auxiliaries
        + assembly_and_test_fixed
    )

    return {
        "silicon_steel_price": silicon_steel_price,
        "winding_process_per_copper_kg": winding_process_per_copper_kg,
        "auxiliary_on_copper": auxiliary_on_copper,
        "auxiliary_per_total_kg": auxiliary_per_total_kg,
        "assembly_and_test_fixed": assembly_and_test_fixed,
        "silicon_correction": silicon_correction,
        "winding_conversion": winding_conversion,
        "auxiliaries": auxiliaries,
        "estimated_total": total,
        "coverage_vs_purchase": (total / entry["purchase_total"]) if entry["purchase_total"] else None,
        "gap_vs_purchase": (entry["purchase_total"] - total) if entry["purchase_total"] else None,
    }


def classify_entry(
    entry: dict,
    repeated_map: dict[tuple[str, float], list[tuple[str, str, str]]],
    item_purchase_map: dict[str, list[dict]],
) -> list[str]:
    reasons: list[str] = []
    coverage = entry["coverage"]
    buy = entry["purchase_total"]
    if buy is None:
        return reasons

    same_purchase = repeated_map.get((entry["product"], buy), [])
    if len(same_purchase) > 1 and contains_any(entry["item"], ["\u7ec4\u4ef6", "\u603b\u6210"]):
        reasons.append("purchase_total_reused_across_multiple_codes")

    if coverage is not None and coverage < 0.35 and contains_any(entry["item"], ["\u7ec4\u4ef6", "\u603b\u6210"]):
        reasons.append("finished_goods_scope_exceeds_current_raw_material_estimate")

    copper_without_process = any(
        contains_any(row["material"], ["\u94dc\u7ebf", "\u6241\u94dc", "\u6f06\u5305"]) and not row["process"]
        for row in entry["rows"]
    )
    if copper_without_process:
        reasons.append("missing_winding_embedding_varnish_test_process")

    low_proxy_material = any(
        "\u4ee3\u7406" in row["price_source"]
        and contains_any(row["material"], ["\u94c1\u82af", "\u7845\u94a2", "\u78c1\u94a2", "\u4e0d\u9508\u94a2"])
        for row in entry["rows"]
    )
    if low_proxy_material:
        reasons.append("material_proxy_likely_underprices_core_material")

    peer_ref = get_peer_purchase_reference(entry, item_purchase_map)
    if (
        peer_ref
        and buy is not None
        and abs(buy - peer_ref["reference"]) / peer_ref["reference"] > 0.5
    ):
        reasons.append("purchase_price_outlier_vs_other_stack_heights")

    peer_band = get_peer_purchase_band(entry, item_purchase_map)
    if peer_band:
        reasons.append("peer_purchase_prices_split_into_multiple_bands")

    if (
        entry["weight_kg"] <= 1.0
        and entry["line_total"] > 0
        and buy / entry["line_total"] >= 8.0
        and any(
            contains_any(row["process"], ["\u51b7\u9566", "\u51b2\u538b", "\u6298\u5f2f", "\u673a\u52a0\u5de5", "\u6fc0\u5149"])
            for row in entry["rows"]
        )
    ):
        reasons.append("minimum_lot_or_minimum_process_charge_not_captured")

    if (
        entry["process_total"] > 0
        and entry["line_total"] > 0
        and buy / entry["line_total"] >= 10.0
        and any(
            contains_any(row["process"], ["\u673a\u52a0\u5de5", "\u51b7\u9566", "\u6ce8\u5851", "\u94a3\u91d1"])
            for row in entry["rows"]
        )
    ):
        reasons.append("piece_rate_or_process_formula_is_far_below_purchase_floor")

    if not reasons and coverage is not None and coverage < 0.6:
        reasons.append("purchase_scope_or_supplier_markup_exceeds_current_formula")

    return reasons


def localize_reasons(reasons: list[str]) -> list[str]:
    return [REASON_LABELS.get(reason, reason) for reason in reasons]


def estimate_optimized_total(entry: dict, item_purchase_map: dict[str, list[dict]]) -> dict:
    stator_model = estimate_stator_finished_goods(entry)
    if stator_model:
        return {
            "estimated_total": stator_model["estimated_total"],
            "basis": "\u5b9a\u5b50\u6210\u54c1\u53e3\u5f84\u6a21\u578b\uff08\u7845\u94a2\u7247\u4fee\u6b63+\u7ed5\u7ebf/\u5d4c\u7ebf/\u6d78\u6f06\u5de5\u5e8f+\u8f85\u6599\u635f\u8017+\u88c5\u914d\u6d4b\u8bd5\uff09",
            "gap_vs_purchase": stator_model["gap_vs_purchase"],
            "coverage_vs_purchase": stator_model["coverage_vs_purchase"],
        }

    peer_ref = get_peer_purchase_reference(entry, item_purchase_map)
    if peer_ref and entry["coverage"] is not None and entry["coverage"] < 0.7:
        reference = peer_ref["reference"]
        buy = entry["purchase_total"]
        return {
            "estimated_total": reference,
            "basis": "\u53c2\u8003\u5176\u4ed6\u53e0\u9ad8\u4ea7\u54c1\u4e2d\u540c\u7c7b\u96f6\u4ef6\u7684\u91c7\u8d2d\u4e2d\u4f4d\u4ef7",
            "gap_vs_purchase": (buy - reference) if buy is not None else None,
            "coverage_vs_purchase": (reference / buy) if buy else None,
        }

    buy = entry["purchase_total"]
    return {
        "estimated_total": entry["line_total"],
        "basis": "\u57fa\u7840\u6750\u6599+\u5de5\u827a\u516c\u5f0f",
        "gap_vs_purchase": entry["gap"],
        "coverage_vs_purchase": entry["coverage"],
    }


def explain_entry(
    entry: dict,
    repeated_map: dict[tuple[str, float], list[tuple[str, str, str]]],
    item_purchase_map: dict[str, list[dict]],
) -> list[str]:
    messages: list[str] = []
    buy = entry["purchase_total"]
    if buy is None:
        return messages

    same_purchase = repeated_map.get((entry["product"], buy), [])
    if len(same_purchase) > 1:
        peer_codes = [code for _, _, code in same_purchase if code != entry["code"]]
        if peer_codes:
            messages.append(
                "\u540c\u4e00\u4ea7\u54c1\u5185\u8fd8\u6709\u5176\u4ed6\u7269\u6599\u7f16\u7801\u5171\u7528\u8be5\u91c7\u8d2d\u4ef7\uff1a"
                + "\u3001".join(peer_codes)
            )

    peer_ref = get_peer_purchase_reference(entry, item_purchase_map)
    if peer_ref and abs(buy - peer_ref["reference"]) / peer_ref["reference"] > 0.5:
        messages.append(
            "\u5176\u4ed6\u53e0\u9ad8\u4ea7\u54c1\u4e2d\u540c\u7c7b\u96f6\u4ef6\u7684\u91c7\u8d2d\u4ef7\u5927\u591a\u96c6\u4e2d\u5728 "
            f"{peer_ref['reference']:.2f} \u5143\u9644\u8fd1\uff0c\u672c\u6761\u4e3a {buy:.2f} \u5143"
        )

    peer_band = get_peer_purchase_band(entry, item_purchase_map)
    if peer_band and not peer_ref:
        messages.append(
            "\u540c\u7c7b\u96f6\u4ef6\u5728\u5176\u4ed6\u53e0\u9ad8\u4ea7\u54c1\u4e2d\u5b58\u5728 "
            f"{peer_band['min']:.2f} \u5230 {peer_band['max']:.2f} \u5143\u7684\u591a\u6863\u91c7\u8d2d\u4ef7\uff0c\u9700\u8981\u6838\u5bf9\u7248\u672c\u6216\u53e3\u5f84"
        )

    if "\u5b9a\u5b50\u7ec4\u4ef6" in entry["item"]:
        optimized = estimate_optimized_total(entry, item_purchase_map)
        messages.append(
            "\u5b9a\u5b50\u4ef6\u5f53\u524d\u57fa\u7840\u4f30\u7b97\u6f0f\u6389\u4e86\u7ed5\u7ebf\u3001\u5d4c\u7ebf\u3001\u6d78\u6f06\u3001\u70d8\u5e72\u548c\u6d4b\u8bd5\u7b49\u5173\u952e\u5de5\u5e8f"
        )
        messages.append(
            "\u6309\u6210\u54c1\u53e3\u5f84\u4f18\u5316\u540e\u4f30\u7b97\u7ea6\u4e3a "
            f"{optimized['estimated_total']:.2f} \u5143"
        )

    return messages


def median_value(values: list[float]) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    mid = len(ordered) // 2
    if len(ordered) % 2 == 1:
        return ordered[mid]
    return (ordered[mid - 1] + ordered[mid]) / 2


def build_item_purchase_map(entries: list[dict]) -> dict[str, list[dict]]:
    item_map: dict[str, list[dict]] = defaultdict(list)
    for entry in entries:
        if entry["purchase_total"] is not None:
            item_map[entry["item"]].append(entry)
    return item_map


def get_peer_purchase_values(entry: dict, item_purchase_map: dict[str, list[dict]]) -> list[float]:
    peers = []
    for peer in item_purchase_map.get(entry["item"], []):
        if peer["product"] == entry["product"]:
            continue
        if peer["purchase_total"] is None:
            continue
        peers.append(peer["purchase_total"])
    return sorted(peers)


def get_peer_purchase_reference(entry: dict, item_purchase_map: dict[str, list[dict]]) -> dict | None:
    values = get_peer_purchase_values(entry, item_purchase_map)
    if len(values) < 2 or values[0] <= 0:
        return None
    if values[-1] / values[0] > 1.5:
        return None
    return {
        "reference": median_value(values),
        "min": values[0],
        "max": values[-1],
        "count": len(values),
    }


def get_peer_purchase_band(entry: dict, item_purchase_map: dict[str, list[dict]]) -> dict | None:
    values = get_peer_purchase_values(entry, item_purchase_map)
    if len(values) < 2 or values[0] <= 0:
        return None
    ratio = values[-1] / values[0]
    if ratio < 2.2:
        return None
    return {
        "min": values[0],
        "max": values[-1],
        "count": len(values),
        "ratio": ratio,
    }


def print_row_details(entry: dict) -> None:
    for row in entry["rows"]:
        purchase = row["purchase_total_text"] or "-"
        print(
            "    - material={material} process={process} qty={qty} weight={weight:.4f} "
            "base={base:.4f} process_total={proc:.4f} line_total={line:.4f} purchase={purchase}".format(
                material=row["material"] or "-",
                process=row["process"] or "-",
                qty=row["qty"] or "-",
                weight=row["weight_kg"],
                base=row["base_total"],
                proc=row["process_total"],
                line=row["line_total"],
                purchase=purchase,
            )
        )


def print_summary(
    entries: list[dict],
    repeated_map: dict[tuple[str, float], list[tuple[str, str, str]]],
    item_purchase_map: dict[str, list[dict]],
    show_rows: bool,
) -> None:
    for entry in entries:
        print(f"{entry['code']} | {entry['product']} | {entry['item']}")
        print(
            "  base_estimated={:.4f} base={:.4f} process={:.4f} purchase={} gap={} coverage={}".format(
                entry["line_total"],
                entry["base_total"],
                entry["process_total"],
                f"{entry['purchase_total']:.4f}" if entry["purchase_total"] is not None else "-",
                f"{entry['gap']:.4f}" if entry["gap"] is not None else "-",
                f"{entry['coverage']:.2%}" if entry["coverage"] is not None else "-",
            )
        )
        reasons = classify_entry(entry, repeated_map, item_purchase_map)
        if reasons:
            print("  reasons=" + "；".join(localize_reasons(reasons)))

        optimized = estimate_optimized_total(entry, item_purchase_map)
        if optimized:
            print(
                "  optimized_estimated={:.4f} basis={}".format(
                    optimized["estimated_total"],
                    optimized["basis"],
                )
            )
            print(
                "  optimized_vs_purchase gap={} coverage={}".format(
                    f"{optimized['gap_vs_purchase']:.4f}" if optimized["gap_vs_purchase"] is not None else "-",
                    f"{optimized['coverage_vs_purchase']:.2%}" if optimized["coverage_vs_purchase"] is not None else "-",
                )
            )

        details = explain_entry(entry, repeated_map, item_purchase_map)
        if details:
            print("  detail=" + "；".join(details))

        print(
            "  weight_kg={:.4f} ext_weight_kg={:.4f}".format(
                entry["weight_kg"], entry["ext_weight_kg"]
            )
        )
        if show_rows:
            print_row_details(entry)
        print()


def print_top_anomalies(
    entries: list[dict],
    repeated_map: dict[tuple[str, float], list[tuple[str, str, str]]],
    item_purchase_map: dict[str, list[dict]],
    top_n: int,
) -> None:
    filtered = [entry for entry in entries if entry["gap"] is not None and entry["gap"] > 0]
    filtered.sort(key=lambda x: x["gap"], reverse=True)
    print(f"TOP_POSITIVE_GAPS={min(top_n, len(filtered))}")
    for entry in filtered[:top_n]:
        reasons = classify_entry(entry, repeated_map, item_purchase_map)
        print(
            "{code} | {product} | {item} | gap={gap:.4f} | coverage={coverage:.2%} | reasons={reasons}".format(
                code=entry["code"],
                product=entry["product"],
                item=entry["item"],
                gap=entry["gap"],
                coverage=entry["coverage"] if entry["coverage"] is not None else 0.0,
                reasons="；".join(localize_reasons(reasons)) if reasons else "-",
            )
        )


def format_optional_float(value: float | None) -> str:
    return "" if value is None else f"{value:.4f}"


def export_summary(
    entries: list[dict],
    repeated_map: dict[tuple[str, float], list[tuple[str, str, str]]],
    item_purchase_map: dict[str, list[dict]],
    output_path: Path,
) -> None:
    header = [
        "\u4ea7\u54c1",
        "\u7269\u6599",
        "\u7269\u6599\u7f16\u7801",
        "\u57fa\u7840\u4f30\u7b97\u603b\u4ef7",
        "\u57fa\u7840\u6750\u6599\u5408\u8ba1",
        "\u57fa\u7840\u5de5\u827a\u5408\u8ba1",
        "\u91c7\u8d2d\u603b\u4ef7",
        "\u57fa\u7840\u5dee\u989d",
        "\u57fa\u7840\u8986\u76d6\u7387",
        "\u4f18\u5316\u540e\u4f30\u7b97",
        "\u4f18\u5316\u4f9d\u636e",
        "\u4f18\u5316\u540e\u5dee\u989d",
        "\u4f18\u5316\u540e\u8986\u76d6\u7387",
        "\u603b\u91cd\u91cfkg",
        "\u4e3b\u8981\u6750\u8d28/\u5de5\u827a\u660e\u7ec6",
        "\u5dee\u5f02\u539f\u56e0\u6807\u7b7e",
        "\u5dee\u5f02\u539f\u56e0\u660e\u7ec6",
    ]

    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in build_export_rows(entries, repeated_map, item_purchase_map):
            writer.writerow(row)


def build_export_rows(
    entries: list[dict],
    repeated_map: dict[tuple[str, float], list[tuple[str, str, str]]],
    item_purchase_map: dict[str, list[dict]],
) -> list[list[str]]:
    rows: list[list[str]] = []
    for entry in entries:
        reasons = classify_entry(entry, repeated_map, item_purchase_map)
        reason_text = "；".join(localize_reasons(reasons))
        detail_text = "；".join(explain_entry(entry, repeated_map, item_purchase_map))
        optimized = estimate_optimized_total(entry, item_purchase_map)
        detail = " | ".join(
            "{material}/{process}/base={base:.2f}/proc={proc:.2f}/line={line:.2f}".format(
                material=row["material"] or "-",
                process=row["process"] or "-",
                base=row["base_total"],
                proc=row["process_total"],
                line=row["line_total"],
            )
            for row in entry["rows"]
        )
        rows.append(
            [
                entry["product"],
                entry["item"],
                entry["code"],
                f"{entry['line_total']:.4f}",
                f"{entry['base_total']:.4f}",
                f"{entry['process_total']:.4f}",
                format_optional_float(entry["purchase_total"]),
                format_optional_float(entry["gap"]),
                "" if entry["coverage"] is None else f"{entry['coverage']:.2%}",
                f"{optimized['estimated_total']:.4f}",
                optimized["basis"],
                format_optional_float(optimized["gap_vs_purchase"]),
                "" if optimized["coverage_vs_purchase"] is None else f"{optimized['coverage_vs_purchase']:.2%}",
                f"{entry['ext_weight_kg']:.4f}",
                detail,
                reason_text,
                detail_text,
            ]
        )
    return rows


def export_product_workbook(
    entries: list[dict],
    repeated_map: dict[tuple[str, float], list[tuple[str, str, str]]],
    item_purchase_map: dict[str, list[dict]],
    output_path: Path,
) -> None:
    if xlsxwriter is None:
        _export_product_workbook_openpyxl(entries, repeated_map, item_purchase_map, output_path)
        return

    header = [
        "产品",
        "物料",
        "物料编码",
        "基础估算总价",
        "基础材料合计",
        "基础工艺合计",
        "采购总价",
        "基础差额",
        "基础覆盖率",
        "优化后估算",
        "优化依据",
        "优化后差额",
        "优化后覆盖率",
        "总重量kg",
        "主要材质/工艺明细",
        "差异原因标签",
        "差异原因明细",
    ]
    product_rows: dict[str, list[list[str]]] = defaultdict(list)
    for row in build_export_rows(entries, repeated_map, item_purchase_map):
        product_rows[row[0]].append(row)

    workbook = xlsxwriter.Workbook(str(output_path))
    header_fmt = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "border": 1, "text_wrap": True, "valign": "vcenter"})
    text_fmt = workbook.add_format({"border": 1, "valign": "top"})
    money_fmt = workbook.add_format({"border": 1, "num_format": "0.0000", "valign": "top"})
    percent_fmt = workbook.add_format({"border": 1, "num_format": "0.00%", "valign": "top"})
    stator_text_fmt = workbook.add_format({"border": 1, "bg_color": "#FFF2CC", "valign": "top"})
    stator_money_fmt = workbook.add_format({"border": 1, "bg_color": "#FFF2CC", "num_format": "0.0000", "valign": "top"})
    stator_percent_fmt = workbook.add_format({"border": 1, "bg_color": "#FFF2CC", "num_format": "0.00%", "valign": "top"})
    medium_text_fmt = workbook.add_format({"border": 1, "bg_color": "#FCE5CD", "valign": "top"})
    medium_money_fmt = workbook.add_format({"border": 1, "bg_color": "#FCE5CD", "num_format": "0.0000", "valign": "top"})
    medium_percent_fmt = workbook.add_format({"border": 1, "bg_color": "#FCE5CD", "num_format": "0.00%", "valign": "top"})
    severe_text_fmt = workbook.add_format({"border": 1, "bg_color": "#F4CCCC", "valign": "top"})
    severe_money_fmt = workbook.add_format({"border": 1, "bg_color": "#F4CCCC", "num_format": "0.0000", "valign": "top"})
    severe_percent_fmt = workbook.add_format({"border": 1, "bg_color": "#F4CCCC", "num_format": "0.00%", "valign": "top"})

    widths = [20, 14, 14, 13, 13, 13, 13, 13, 11, 13, 24, 13, 12, 11, 52, 36, 48]
    numeric_cols = {3, 4, 5, 6, 7, 9, 11, 13}
    percent_cols = {8, 12}

    used_sheet_names: set[str] = set()
    for product in sorted(product_rows):
        sheet_name = _build_sheet_name(product, used_sheet_names)
        worksheet = workbook.add_worksheet(sheet_name)
        worksheet.freeze_panes(1, 0)
        worksheet.autofilter(0, 0, max(len(product_rows[product]), 1), len(header) - 1)

        for col, value in enumerate(header):
            worksheet.write(0, col, value, header_fmt)
            worksheet.set_column(col, col, widths[col])

        worksheet.write(0, len(header) + 1, "颜色说明", header_fmt)
        worksheet.write(1, len(header) + 1, "红色：优化后差额>3000或优化后覆盖率<70%，优先核对", severe_text_fmt)
        worksheet.write(2, len(header) + 1, "橙色：优化后差额>1000或优化后覆盖率<90%，建议复核", medium_text_fmt)
        worksheet.write(3, len(header) + 1, "黄色：定子成品口径或重点说明列", stator_text_fmt)
        worksheet.set_column(len(header) + 1, len(header) + 1, 34)

        product_rows[product].sort(key=lambda row: float(row[7] or 0), reverse=True)
        for r_idx, row in enumerate(product_rows[product], start=1):
            is_stator = row[1] == "定子组件"
            optimized_gap = float(row[11] or 0) if row[11] else 0.0
            optimized_coverage = float(row[12].rstrip("%")) / 100 if row[12] else 0.0
            severe = (row[11] and abs(optimized_gap) > 3000) or (row[12] and optimized_coverage < 0.70)
            medium = (row[11] and abs(optimized_gap) > 1000) or (row[12] and optimized_coverage < 0.90)
            for c_idx, value in enumerate(row):
                if severe and c_idx in {6, 7, 8, 9, 10, 11, 12, 15, 16}:
                    text_style = severe_text_fmt
                    money_style = severe_money_fmt
                    percent_style = severe_percent_fmt
                elif medium and c_idx in {6, 7, 8, 9, 10, 11, 12, 15, 16}:
                    text_style = medium_text_fmt
                    money_style = medium_money_fmt
                    percent_style = medium_percent_fmt
                elif is_stator and c_idx in {1, 2, 9, 10, 11, 12, 15, 16}:
                    text_style = stator_text_fmt
                    money_style = stator_money_fmt
                    percent_style = stator_percent_fmt
                else:
                    text_style = text_fmt
                    money_style = money_fmt
                    percent_style = percent_fmt
                if c_idx in numeric_cols:
                    worksheet.write_number(r_idx, c_idx, float(value) if value else 0.0, money_style)
                elif c_idx in percent_cols:
                    percent_value = float(value.rstrip("%")) / 100 if value else 0.0
                    worksheet.write_number(r_idx, c_idx, percent_value, percent_style)
                else:
                    worksheet.write(r_idx, c_idx, value, text_style)

    workbook.close()


def _build_sheet_name(product: str, used_names: set[str]) -> str:
    match = re.search(r"(\d+)叠高", product)
    base = match.group(1) if match else product.replace("三一380电机", "").strip()
    base = (base or "产品").replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "")
    base = base.replace("[", "(").replace("]", ")").replace(":", "：")
    base = base[:31] or "产品"
    candidate = base
    counter = 2
    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = f"{base[: max(1, 31 - len(suffix))]}{suffix}"
        counter += 1
    used_names.add(candidate)
    return candidate


def _export_product_workbook_openpyxl(
    entries: list[dict],
    repeated_map: dict[tuple[str, float], list[tuple[str, str, str]]],
    item_purchase_map: dict[str, list[dict]],
    output_path: Path,
) -> None:
    if Workbook is None or PatternFill is None or Border is None or Side is None or Font is None or Alignment is None:
        raise RuntimeError("导出分产品工作簿需要 xlsxwriter 或 openpyxl")

    header = [
        "产品",
        "物料",
        "物料编码",
        "基础估算总价",
        "基础材料合计",
        "基础工艺合计",
        "采购总价",
        "基础差额",
        "基础覆盖率",
        "优化后估算",
        "优化依据",
        "优化后差额",
        "优化后覆盖率",
        "总重量kg",
        "主要材质/工艺明细",
        "差异原因标签",
        "差异原因明细",
    ]
    widths = [20, 14, 14, 13, 13, 13, 13, 13, 11, 13, 24, 13, 12, 11, 52, 36, 48]
    numeric_cols = {4, 5, 6, 7, 8, 10, 12, 14}
    percent_cols = {9, 13}

    product_rows: dict[str, list[list[str]]] = defaultdict(list)
    for row in build_export_rows(entries, repeated_map, item_purchase_map):
        product_rows[row[0]].append(row)

    workbook = Workbook()
    workbook.remove(workbook.active)

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "header": PatternFill("solid", fgColor="D9EAF7"),
        "stator": PatternFill("solid", fgColor="FFF2CC"),
        "medium": PatternFill("solid", fgColor="FCE5CD"),
        "severe": PatternFill("solid", fgColor="F4CCCC"),
    }

    used_sheet_names: set[str] = set()
    for product in sorted(product_rows):
        sheet_name = _build_sheet_name(product, used_sheet_names)
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.freeze_panes = "A2"

        for col, value in enumerate(header, start=1):
            cell = worksheet.cell(row=1, column=col, value=value)
            cell.font = Font(bold=True)
            cell.fill = fills["header"]
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            worksheet.column_dimensions[cell.column_letter].width = widths[col - 1]

        worksheet.cell(row=1, column=len(header) + 2, value="颜色说明").fill = fills["header"]
        worksheet.cell(row=1, column=len(header) + 2).font = Font(bold=True)
        worksheet.cell(row=2, column=len(header) + 2, value="红色：优化后差额>3000或优化后覆盖率<70%，优先核对").fill = fills["severe"]
        worksheet.cell(row=3, column=len(header) + 2, value="橙色：优化后差额>1000或优化后覆盖率<90%，建议复核").fill = fills["medium"]
        worksheet.cell(row=4, column=len(header) + 2, value="黄色：定子成品口径或重点说明列").fill = fills["stator"]

        product_rows[product].sort(key=lambda row: float(row[7] or 0), reverse=True)
        for r_idx, row in enumerate(product_rows[product], start=2):
            is_stator = row[1] == "定子组件"
            optimized_gap = float(row[11] or 0) if row[11] else 0.0
            optimized_coverage = float(row[12].rstrip("%")) / 100 if row[12] else 0.0
            severe = (row[11] and abs(optimized_gap) > 3000) or (row[12] and optimized_coverage < 0.70)
            medium = (row[11] and abs(optimized_gap) > 1000) or (row[12] and optimized_coverage < 0.90)

            for c_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                if c_idx in numeric_cols:
                    cell.value = float(value) if value else 0.0
                    cell.number_format = "0.0000"
                elif c_idx in percent_cols:
                    cell.value = float(value.rstrip("%")) / 100 if value else 0.0
                    cell.number_format = "0.00%"
                else:
                    cell.value = value

                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if severe and c_idx in {7, 8, 9, 10, 11, 12, 13, 16, 17}:
                    cell.fill = fills["severe"]
                elif medium and c_idx in {7, 8, 9, 10, 11, 12, 13, 16, 17}:
                    cell.fill = fills["medium"]
                elif is_stator and c_idx in {2, 3, 10, 11, 12, 13, 16, 17}:
                    cell.fill = fills["stator"]

    workbook.save(output_path)


def export_summary_doc(
    entries: list[dict],
    repeated_map: dict[tuple[str, float], list[tuple[str, str, str]]],
    item_purchase_map: dict[str, list[dict]],
    output_path: Path,
) -> None:
    lines: list[str] = []
    lines.append("# 三一380电机BOM差异总结")
    lines.append("")
    lines.append("## 结论")
    lines.append("")

    top_entries = [entry for entry in entries if entry["gap"] is not None]
    top_entries.sort(key=lambda x: x["gap"], reverse=True)
    for entry in top_entries[:6]:
        optimized = estimate_optimized_total(entry, item_purchase_map)
        reasons = "；".join(localize_reasons(classify_entry(entry, repeated_map, item_purchase_map)))
        lines.append(
            "- {product} {item} {code}：基础估算 {base:.2f} 元，采购价 {buy:.2f} 元，优化后估算 {opt:.2f} 元，优化后差额 {gap:.2f} 元。原因：{reasons}".format(
                product=entry["product"],
                item=entry["item"],
                code=entry["code"],
                base=entry["line_total"],
                buy=entry["purchase_total"] or 0.0,
                opt=optimized["estimated_total"],
                gap=optimized["gap_vs_purchase"] or 0.0,
                reasons=reasons or "无",
            )
        )

    lines.append("")
    lines.append("## 按产品观察")
    lines.append("")
    by_product: dict[str, list[dict]] = defaultdict(list)
    for entry in entries:
        by_product[entry["product"]].append(entry)

    for product in sorted(by_product):
        product_entries = sorted(
            [entry for entry in by_product[product] if entry["gap"] is not None],
            key=lambda x: x["gap"],
            reverse=True,
        )
        lines.append(f"### {product}")
        for entry in product_entries[:3]:
            optimized = estimate_optimized_total(entry, item_purchase_map)
            lines.append(
                "- {item} {code}：采购价 {buy:.2f} 元，优化后估算 {opt:.2f} 元，优化后覆盖率 {cov}".format(
                    item=entry["item"],
                    code=entry["code"],
                    buy=entry["purchase_total"] or 0.0,
                    opt=optimized["estimated_total"],
                    cov=f"{optimized['coverage_vs_purchase']:.2%}" if optimized["coverage_vs_purchase"] is not None else "-",
                )
            )
        lines.append("")

    lines.append("## 重点原因归纳")
    lines.append("")
    lines.append("- 定子组件的主要问题不是单纯材料价偏差，而是基础估算漏掉了绕线、嵌线、浸漆、烘干、测试等成品工序。")
    lines.append("- 机壳、后端盖等少数零件在其他叠高产品里的采购价明显更集中，本次表中个别采购价更像离群值，需要先回原始采购单核价。")
    lines.append("- 当同类零件在不同叠高之间出现多档采购价时，更可能是版本、供应商或采购口径不同，而不是单一材料差。")
    lines.append("")
    lines.append("## 颜色说明")
    lines.append("")
    lines.append("- 红色：优化后差额大于 3000 元，或优化后覆盖率低于 70%，优先核对。")
    lines.append("- 橙色：优化后差额大于 1000 元，或优化后覆盖率低于 90%，建议复核。")
    lines.append("- 黄色：定子成品口径相关重点列。")
    lines.append("")
    lines.append("## 输出文件")
    lines.append("")
    lines.append("- Excel分表版：`sany-380-motor-review-by-product-optimized-colored.xlsx`")
    lines.append("- 汇总CSV：`sany-380-motor-review-optimized-cn.csv`")

    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Group BOM pricing rows by material code, classify likely gap causes, and add a finished-goods stator estimate."
    )
    parser.add_argument("csv_path", help="Path to the BOM pricing CSV")
    parser.add_argument("--code", help="Only show one material code")
    parser.add_argument("--top", type=int, default=10, help="Show top N positive-gap anomalies")
    parser.add_argument("--no-rows", action="store_true", help="Hide per-row details")
    parser.add_argument("--export-summary", help="Write a grouped summary CSV to this path")
    parser.add_argument("--export-xlsx-by-product", help="Write one XLSX workbook with one worksheet per product")
    parser.add_argument("--export-summary-doc", help="Write a markdown summary document to this path")
    args = parser.parse_args()

    _, header, rows = read_csv_rows(Path(args.csv_path))
    if header[: len(EXPECTED_HEADER)] != EXPECTED_HEADER:
        print("HEADER_NOTE=Header differs from expected names, analysis continues by column position.")

    all_entries = build_entries(rows, None)
    repeated_map = build_purchase_repetition_map(all_entries)
    item_purchase_map = build_item_purchase_map(all_entries)
    entries = [entry for entry in all_entries if not args.code or entry["code"] == args.code]

    if args.code:
        print_summary(entries, repeated_map, item_purchase_map, show_rows=not args.no_rows)
    else:
        print_top_anomalies(entries, repeated_map, item_purchase_map, args.top)

    if args.export_summary:
        export_summary(entries, repeated_map, item_purchase_map, Path(args.export_summary))
        print(f"SUMMARY_EXPORTED={args.export_summary}")
    if args.export_xlsx_by_product:
        export_product_workbook(entries, repeated_map, item_purchase_map, Path(args.export_xlsx_by_product))
        print(f"XLSX_EXPORTED={args.export_xlsx_by_product}")
    if args.export_summary_doc:
        export_summary_doc(entries, repeated_map, item_purchase_map, Path(args.export_summary_doc))
        print(f"SUMMARY_DOC_EXPORTED={args.export_summary_doc}")


if __name__ == "__main__":
    main()



