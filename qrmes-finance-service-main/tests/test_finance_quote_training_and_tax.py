from io import BytesIO
from pathlib import Path
import sys
import unittest
import zipfile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
APP_WEB_ROOT = ROOT / "app_web"
for candidate in (ROOT, APP_WEB_ROOT):
    if str(candidate) not in sys.path:
        sys.path.insert(0, str(candidate))

from app_web.backend.config import load_config
from app_web.backend.services.excel_quote_service import ExcelQuoteService
from app_web.backend.services.finance_skill_quote_service import FinanceSkillQuoteService


TRAINING_DATA = ROOT / "app_web" / "backend" / "data" / "liugong_motor_training_reference.json"


class FinanceQuoteTrainingAndTaxTests(unittest.TestCase):
    def test_load_config_defaults_to_finance_demo_static_directory(self):
        config = load_config()

        self.assertEqual(
            config.static_dir,
            (ROOT / "app_web" / "static" / "finance_demo").resolve(),
        )
        self.assertEqual(
            config.demo_data_path,
            (ROOT / "app_web" / "static" / "finance_demo" / "data" / "demo_data.json").resolve(),
        )

    def test_training_reference_dataset_exists(self):
        self.assertTrue(
            TRAINING_DATA.exists(),
            "柳工双12/15/20电机训练数据还没有落成结构化参考文件，报价系统无法稳定复用该训练文档",
        )

    def test_finance_skill_quote_service_can_lookup_liugong_training_reference_by_code(self):
        service = FinanceSkillQuoteService(config=None, kingdee_service=None, ai_route_service=None)

        hit = service.find_training_reference({
            "code": "T44201229",
            "name": "定子组件",
            "spec": "OD200扁线定子组件",
        })

        self.assertIsNotNone(hit)
        self.assertEqual(hit["material"], "漆包线（铜）")
        self.assertAlmostEqual(float(hit["weight_kg"]), 4.95, places=2)

    def test_allocate_tax_inclusive_prices_keeps_total_unchanged_while_breaking_out_13_percent_tax(self):
        rows = [
            {"code": "A", "qty": 2, "ai_route_unit_price": 113.0},
            {"code": "B", "qty": 1, "ai_route_unit_price": 226.0},
        ]

        result = FinanceSkillQuoteService.allocate_tax_inclusive_prices(rows, price_field="ai_route_unit_price", tax_rate=0.13)

        original_total = 113.0 * 2 + 226.0
        inclusive_total = sum(float(item["tax_inclusive_unit_price"]) * float(item["qty"]) for item in result)
        exclusive_total = sum(float(item["tax_exclusive_unit_price"]) * float(item["qty"]) for item in result)
        tax_total = sum(float(item["tax_amount_unit_price"]) * float(item["qty"]) for item in result)

        self.assertAlmostEqual(inclusive_total, original_total, places=2)
        self.assertAlmostEqual(exclusive_total + tax_total, original_total, places=2)
        self.assertTrue(all(float(item["tax_amount_unit_price"]) >= 0 for item in result))
        self.assertTrue(all(float(item["tax_exclusive_unit_price"]) <= float(item["tax_inclusive_unit_price"]) for item in result))

    def test_annotate_price_tax_breakdown_can_be_applied_only_to_ai_route_fields(self):
        rows = [
            {
                "code": "A",
                "qty": 1,
                "finance_route_unit_price": 100.0,
                "ai_route_unit_price": 113.0,
            }
        ]

        FinanceSkillQuoteService.annotate_price_tax_breakdown(rows, price_field="ai_route_unit_price", prefix="ai_route", tax_rate=0.13)

        self.assertEqual(rows[0]["finance_route_unit_price"], 100.0)
        self.assertNotIn("finance_route_tax_exclusive_unit_price", rows[0])
        self.assertAlmostEqual(rows[0]["ai_route_tax_inclusive_unit_price"], 113.0, places=6)
        self.assertAlmostEqual(rows[0]["ai_route_tax_exclusive_unit_price"], 100.0, places=2)
        self.assertAlmostEqual(rows[0]["ai_route_tax_amount_unit_price"], 13.0, places=2)

    def test_export_quote_workbook_adds_tax_scope_hint_to_summary_sheet(self):
        service = ExcelQuoteService(config=load_config(), kingdee_service=None, ai_route_service=None)
        workbook_bytes = service.export_quote_workbook(
            {
                "model": {"label": "测试报价", "filename": "demo.xlsx", "sheet_name": "Sheet1", "production_mode": "mass", "annual_volume": 10000},
                "summary": {
                    "finance_total": 100.0,
                    "ai_total_tax_inclusive": 113.0,
                    "ai_total_tax_exclusive": 100.0,
                    "ai_total_tax_amount": 13.0,
                    "selected_quote_total_tax_inclusive": 113.0,
                    "selected_quote_total_tax_exclusive": 100.0,
                    "selected_quote_total_tax_amount": 13.0,
                    "volume_baseline_total": 113.0,
                    "volume_conservative_total": 108.0,
                    "volume_aggressive_total": 104.0,
                },
                "items": [],
            }
        )

        wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
        ws = wb["总价概览"]
        rows = list(ws.iter_rows(values_only=True))
        flattened = [str(cell) for row in rows for cell in row if cell is not None]
        joined = "\n".join(flattened)

        self.assertIn("价格口径说明", joined)
        self.assertIn("默认展示为含税价格", joined)
        self.assertIn("仅 AI 报价与最终选用价拆分未税价和税额", joined)
        self.assertIn("量产基准总价(元)", joined)
        self.assertIn("量产保守总价(元)", joined)
        self.assertIn("量产激进总价(元)", joined)

    def test_export_ai_quote_workbook_places_tooling_compare_columns_next_to_ai_price(self):
        service = ExcelQuoteService(config=load_config(), kingdee_service=None, ai_route_service=None)
        workbook_bytes = service.export_ai_quote_workbook(
            {
                "model": {"label": "AI测试报价", "filename": "ai.xlsx", "sheet_name": "Sheet1", "production_mode": "mass", "annual_volume": 10000},
                "items": [
                    {
                        "code": "X1",
                        "name": "前端盖",
                        "spec": "OD180mm前端盖",
                        "qty": 1,
                        "ai_route_unit_price": 90.12,
                        "ai_route_confidence": 0.92,
                        "sample_machining_unit_price": 219.80,
                        "mass_tooling_unit_price": 90.12,
                        "tooling_cost": 41500,
                        "mass_break_even_volume": 321,
                        "ai_route_process_guess": "高压压铸",
                        "ai_route_material_guess": "ADC12",
                        "ai_route_reasoning": "前端盖量产开模口径...",
                        "comparison_reason_summary": "量产开模价显著低于样品机加工价",
                        "finance_route_unit_price": 120.0,
                        "volume_baseline_unit_price": 90.12,
                        "volume_conservative_unit_price": 86.51,
                        "volume_aggressive_unit_price": 82.91,
                        "status": "双路线可比",
                    }
                ],
            }
        )

        wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
        ws = wb["AI分析明细"]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertIn("样品机加工单价(元)", header)
        self.assertIn("量产开模单价(元)", header)
        self.assertIn("开模费(元)", header)
        self.assertIn("开模收益平衡点(套/年)", header)
        self.assertIn("AI分析摘要", header)
        self.assertLess(header.index("AI报价单价(元)"), header.index("样品机加工单价(元)"))
        self.assertLess(header.index("样品机加工单价(元)"), header.index("AI置信度"))
        self.assertIn("BOM明细与异常清单", wb.sheetnames)
        self.assertIn("差异复核工作簿", wb.sheetnames)

    def test_export_quote_workbook_adds_tooling_items_sheet(self):
        service = ExcelQuoteService(config=load_config(), kingdee_service=None, ai_route_service=None)
        workbook_bytes = service.export_quote_workbook(
            {
                "model": {"label": "测试报价", "filename": "demo.xlsx", "sheet_name": "Sheet1", "production_mode": "mass", "annual_volume": 10000},
                "summary": {},
                "items": [
                    {
                        "code": "X1",
                        "name": "前端盖",
                        "qty": 1,
                        "finance_route_unit_price": 10,
                        "ai_route_unit_price": 90.12,
                        "sample_machining_unit_price": 219.8,
                        "mass_tooling_unit_price": 90.12,
                        "tooling_cost": 41500,
                        "mass_break_even_volume": 321,
                        "volume_baseline_unit_price": 90.12,
                        "production_mode": "mass",
                        "annual_volume": 10000,
                    },
                    {
                        "code": "X2",
                        "name": "定子组件",
                        "qty": 1,
                        "finance_route_unit_price": 10,
                        "ai_route_unit_price": 467.53,
                        "production_mode": "mass",
                        "annual_volume": 10000,
                    }
                ],
            }
        )

        wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
        self.assertIn("开模类物料", wb.sheetnames)
        ws = wb["开模类物料"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertGreaterEqual(len(rows), 2)
        header = list(rows[0])
        self.assertIn("样品机加工单价(元)", header)
        self.assertIn("量产开模单价(元)", header)
        self.assertIn("开模费(元)", header)
        body_codes = [row[0] for row in rows[1:] if row and row[0]]
        self.assertIn("X1", body_codes)
        self.assertNotIn("X2", body_codes)

    def test_export_file_prefix_includes_requested_annual_volume_even_without_label_hint(self):
        service = ExcelQuoteService(config=load_config(), kingdee_service=None, ai_route_service=None)
        prefix = service._export_file_prefix({
            "model": {
                "label": "柳工OD220定子总成",
                "filename": "liugong_quote.xlsx",
                "annual_volume": 3000,
            }
        })

        self.assertIn("3000套_年", prefix)

    def test_describe_quote_exports_uses_annual_volume_in_excel_filenames_even_when_label_has_no_quantity(self):
        service = ExcelQuoteService(config=load_config(), kingdee_service=None, ai_route_service=None)
        exports = service.describe_quote_exports({
            "model": {
                "label": "柳工报价",
                "filename": "liugong_quote.xlsx",
                "annual_volume": 300,
            },
            "summary": {},
            "items": [{"code": "A1", "qty": 1}],
        })

        filenames = [item["filename"] for item in exports]
        self.assertTrue(all("300套_年" in name for name in filenames[:4]))
        self.assertTrue(any("系统行级报价总表" in name for name in filenames))

    def test_export_quote_package_writes_annual_volume_into_zip_member_names_even_when_label_has_no_quantity(self):
        service = ExcelQuoteService(config=load_config(), kingdee_service=None, ai_route_service=None)
        package_bytes = service.export_quote_package({
            "model": {
                "label": "柳工报价",
                "filename": "liugong_quote.xlsx",
                "sheet_name": "Sheet1",
                "production_mode": "mass",
                "annual_volume": 5000,
            },
            "summary": {},
            "items": [
                {
                    "code": "A1",
                    "name": "测试件",
                    "qty": 1,
                    "finance_route_unit_price": 10,
                    "ai_route_unit_price": 12,
                    "volume_baseline_unit_price": 11,
                    "volume_conservative_unit_price": 10.5,
                    "volume_aggressive_unit_price": 10,
                    "annual_volume": 5000,
                    "production_mode": "mass",
                }
            ],
        })

        with zipfile.ZipFile(BytesIO(package_bytes), "r") as archive:
            names = archive.namelist()

        self.assertTrue(any("5000套_年" in name for name in names if name.endswith(".xlsx")))
        self.assertTrue(any("系统AI分析明细" in name for name in names if name.endswith(".xlsx")))
        self.assertTrue(any("BOM明细与异常清单" in name for name in names if name.endswith(".xlsx")))

    def test_export_quote_package_batch_includes_multiple_volume_directories(self):
        service = ExcelQuoteService(config=load_config(), kingdee_service=None, ai_route_service=None)
        batch_bytes = service.export_quote_package_batch([
            {
                "model": {"label": "柳工报价", "filename": "liugong_quote.xlsx", "sheet_name": "Sheet1", "production_mode": "mass", "annual_volume": 300, "requested_volume_label": "300套/年（显示档位 <=300；折扣档位 <=300套/年）"},
                "summary": {},
                "items": [{"code": "A1", "qty": 1, "finance_route_unit_price": 10, "ai_route_unit_price": 12}],
            },
            {
                "model": {"label": "柳工报价", "filename": "liugong_quote.xlsx", "sheet_name": "Sheet1", "production_mode": "mass", "annual_volume": 1000, "requested_volume_label": "1000套/年（显示档位 301-1000；折扣档位 301-1000套/年）"},
                "summary": {},
                "items": [{"code": "A1", "qty": 1, "finance_route_unit_price": 10, "ai_route_unit_price": 12}],
            },
        ])

        with zipfile.ZipFile(BytesIO(batch_bytes), "r") as archive:
            names = archive.namelist()

        self.assertTrue(any(name.startswith("柳工报价_300套_年/") for name in names))
        self.assertTrue(any(name.startswith("柳工报价_1000套_年/") for name in names))
        self.assertFalse(any("显示档位" in name or "折扣档位" in name for name in names))
        self.assertIn("manifest.json", names)

    def test_export_quote_workbook_recomputes_missing_mass_volume_prices_before_writing(self):
        service = ExcelQuoteService(config=load_config(), kingdee_service=None, ai_route_service=None)
        workbook_bytes = service.export_quote_workbook(
            {
                "model": {"label": "测试报价", "filename": "demo.xlsx", "sheet_name": "Sheet1", "production_mode": "mass", "annual_volume": 1000},
                "summary": {},
                "items": [
                    {
                        "code": "A1",
                        "name": "油封",
                        "qty": 1,
                        "production_mode": "mass",
                        "annual_volume": 1000,
                        "ai_route_unit_price": 3.85,
                        "volume_baseline_unit_price": 3.85,
                        "volume_conservative_unit_price": 0,
                        "volume_aggressive_unit_price": 0,
                        "volume_pricing_summary": "量产口径（1000套/年，301-1000档）：基准 3.85 / 保守 0.00 / 激进 0.00 元",
                    }
                ],
            }
        )

        wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
        ws = wb["行级报价总表"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertGreater(float(rows[1][20] or 0), 0)
        self.assertGreater(float(rows[1][21] or 0), 0)


if __name__ == "__main__":
    unittest.main()
