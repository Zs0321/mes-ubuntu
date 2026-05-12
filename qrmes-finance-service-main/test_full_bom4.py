#!/usr/bin/env python3
import os
os.environ["DASHSCOPE_API_KEY"] = ""
os.environ["PRICING_QWEN_TIMEOUT"] = "5"
os.environ["PRICING_QWEN_TIMEOUT_RETRY"] = "0"

import sys
sys.path.insert(0, '/Volumes/172.16.30.10/volume2/mes_ubuntu_split_result/qrmes-finance-service/app_web')
sys.path.insert(0, '/Volumes/172.16.30.10/volume2/mes_ubuntu_split_result/qrmes-finance-service')

from backend.services.finance_skill_quote_service import FinanceSkillQuoteService
import csv
import openpyxl
import tempfile
from pathlib import Path

class FakeAIRouteService:
    is_ready = False
    max_workers = 1
    def load_skill_workflow_hint(self):
        return "price_bom mandatory; gap/format/volume on demand."
    def plan_script_usage(self, plan_input):
        return {"selected_scripts": ["price_bom_xlsx.py", "model_volume_pricing.py"], "registry": ["price_bom_xlsx.py", "model_volume_pricing.py"]}
    def prepare_staged_pricing_input(self, item):
        return item
    def estimate_item(self, item):
        return {"unit_price": 0, "confidence": 0, "reasoning": "", "source": "fake"}

rows_data = [
    ("定子总成","T44001177","OD220定子总成（028）","","","",1.0),
    ("定子组件","T44200282.A0","OD220mm，L135mm浸漆定子组件","扁线4.12*1.92",5.6,"",1.0),
    ("定子铁芯","T44210291.A1","OD220mm，L135mm，三V","B30AHV1500",14.3,"",1.0),
]

svc = FinanceSkillQuoteService(ai_route_service=FakeAIRouteService())

with tempfile.TemporaryDirectory(prefix="finance_test_") as temp_dir:
    workdir = Path(temp_dir)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["产品", "产品编码", "物料", "物料编码", "规格", "材质", "重量（kg）", "工艺", "含税单价", "数量"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, it in enumerate(rows_data, 2):
        ws.cell(row=row_idx, column=1, value="高速油泵")
        ws.cell(row=row_idx, column=2, value="Leviticus-Motor-WP-105-D")
        ws.cell(row=row_idx, column=3, value=it[0])
        ws.cell(row=row_idx, column=4, value=it[1])
        ws.cell(row=row_idx, column=5, value=it[2])
        ws.cell(row=row_idx, column=6, value=it[3] if it[3] else "")
        ws.cell(row=row_idx, column=7, value=it[4] if it[4] else "")
        ws.cell(row=row_idx, column=8, value=it[5] if it[5] else "")
        ws.cell(row=row_idx, column=9, value="")
        ws.cell(row=row_idx, column=10, value=it[6])
    input_xlsx = workdir / "input.xlsx"
    wb.save(input_xlsx)
    
    line_csv, grouped_csv, summary_md, snapshot_json = svc._run_price_bom(input_xlsx, workdir)
    print("line_csv:")
    with open(line_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            print(f"  {row['物料']}: 行总价={row['行总价']} 工艺金额={row['工艺金额']} 基础金额={row['基础金额']}")
    
    volume_outputs = svc._run_volume_pricing(line_csv, workdir, 1000)
    print("\nvolume detail:")
    detail_rows = svc._read_volume_detail_rows(volume_outputs)
    for r in detail_rows:
        print(f"  {r['物料']}: 基准总价={r['基准总价']}")
